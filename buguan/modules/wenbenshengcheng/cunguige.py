import json
import os
import re

import chardet
import configparser
import openpyxl
import pymysql

from modules.condition_input.funcs.db_cnt import get_connection
from openpyxl.styles import Alignment, Border, Side, Font

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center')
font_10 = Font(size=10)

db_config1 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': '产品设计活动库'
}

# === 读取 JSON 数据 ===
def load_json_data(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)


# === 从 JSON 中提取指定 section + 名称 的值 ===
def get_value(data, section, name):
    for section_name, section_data in data.get("DictOutDatas", {}).items():
        if section_name == section:
            for item in section_data.get("Datas", []):
                if item.get("Name") == name:
                    try:
                        return float(item["Value"])
                    except:
                        return item["Value"]
    return None


# === 定义各结构件规格的生成逻辑 ===
def generate_spec(component_name, data, product_id=None):
    """
    根据元件名称返回其规格字符串，如：EHA500X10;h=8
    如果无法生成返回 None
    """

    if component_name == "管箱封头":
        dh = get_value(data, "管箱封头", "封头类型代号")
        d = get_value(data, "管箱封头", "椭圆形封头计算内径")
        t = get_value(data, "管箱封头", "椭圆形封头名义厚度")
        t_plus = get_value(data, "管箱封头", "椭圆形封头最小成型厚度")
        h = get_value(data, "管箱封头", "椭圆形封头直边高度")
        if None not in (dh,d, t, h):
            return f"{dh}{d}×{t}({t_plus});h={h}"

    elif component_name == "管箱圆筒":
        id_ = get_value(data, "管箱圆筒", "圆筒内径")
        t = get_value(data, "管箱圆筒", "圆筒名义厚度")
        l = get_value(data, "管箱圆筒", "圆筒长度")
        if None not in (id_, t, l):
            return f"ID{id_}×{t};L={l}"

    elif component_name == "管箱法兰":
        w = get_value(data, "管箱法兰", "法兰名义外径")
        n = get_value(data, "管箱法兰", "法兰名义内径")
        h = get_value(data, "管箱法兰", "法兰颈部高度")+get_value(data, "管箱法兰", "法兰名义厚度")
        if None not in (w, n, h):
            return f"Ø{w}/Ø{n}；H={h}"

    elif component_name == "分程隔板":
        t = get_value(data, "管箱分程隔板", "管箱分程隔板名义厚度")
        if t is not None:
            return f"δ={t}"

    elif component_name == "管箱垫片":
        w = get_value(data, "管箱法兰", "垫片名义外径")
        n = get_value(data, "管箱法兰", "垫片名义内径")
        if None not in (w, n):
            return f"Ø{w}/Ø{n}"

    elif component_name == "U形换热管":
        w = get_value(data, "固定管板", "换热管外径")
        b = get_value(data, "固定管板", "换热管壁厚")
        l = get_pipe_param_value("换热管公称长度LN")
        if None not in (w, b, l):
            return f"Ø{w}×Ø{b};L={l}"
    elif component_name == "旁路挡板":
        w = get_pipe_param_value("旁路挡板厚度")
        if w not in(None,"Null","null"):
            return f"δ={w}"
    elif component_name == "固定管板":
        w = get_value(data, "固定管板", "管板名义厚度")
        if w is not None:
            return f"δ={w}"
    elif component_name == "定距管":
        # w = get_value(data, "管束", "换热管外径")
        # n = get_value(data, "管束", "换热管壁厚")
        # val1 = get_value(data, "管束", "定距管长度1")
        # val2 = get_value(data, "管束", "定距管长度2")
        # l = max(val1, val2)
        # if None not in (w,n,l):
        #     return f"Ø{w}×{n};L={l}"
        w = get_value(data, "固定管板", "换热管外径")
        b = get_value(data, "固定管板", "换热管壁厚")
        l = get_pipe_param_value("换热管公称长度LN")
        if None not in (w, b, l):
            return f"Ø{w}×Ø{b};L={l}"
    elif component_name == "折流板":
        w = get_value(data, "管束", "折流板厚度")
        if w is not None:
            return f"δ={w}"
    elif component_name == "防冲板":
        w = get_pipe_param_value("防冲板厚度")
        if w is not None:
            return f"δ={w}"
    elif component_name == "支持板":
        w = get_value(data, "管束", "支持板厚度")
        if w is not None:
            return f"δ={w}"
    elif component_name == "挡管":
        w = get_value(data, "固定管板", "换热管外径")
        b = get_value(data, "固定管板", "换热管壁厚")
        l = get_value(data, "管束", "中间挡管/挡板长度")
        if None not in (w, b, l):
            return f"Ø{w}×{b};L={l}"
    elif component_name == "拉杆":
        val1 = get_value(data, "管束", "拉杆长度1")
        val2 = get_value(data, "管束", "拉杆长度2")
        w = max(val1, val2)
        l = get_value(data, "固定管板", "换热管外径")
        if l is not None:
            try:
                l = float(l)
                if 10 <= l <= 14:
                    rod_diameter = 10
                elif 14 < l < 25:
                    rod_diameter = 12
                elif 25 <= l <= 32:
                    rod_diameter = 16
                elif 32 < l <= 57:
                    rod_diameter = 27
                else:
                    rod_diameter = "[超出范围]"
                return f"Ø{rod_diameter},L={w}"
            except:
                return ""

    elif component_name == "螺母（拉杆）":
        w = get_value(data, "固定管板", "换热管外径")
        if w is not None:
            try:
                w = float(w)
                if 10 <= w <= 14:
                    rod_diameter = 10
                elif 14 < w < 25:
                    rod_diameter = 12
                elif 25 <= w <= 32:
                    rod_diameter = 16
                elif 32 < w <= 57:
                    rod_diameter = 27
                else:
                    rod_diameter = "[超出范围]"
                return f"{rod_diameter}"
            except:
                return ""

    elif component_name == "管箱侧垫片":
        w = get_value(data, "管箱法兰", "垫片名义外径")
        n = get_value(data, "管箱法兰", "垫片名义内径")
        if None not in (w, n):
            return f"Ø{w}/Ø{n}"
    elif component_name == "头盖法兰":
        w = get_value(data, "头盖法兰", "法兰名义外径")
        n = get_value(data, "头盖法兰", "法兰名义内径")
        h = get_value(data, "头盖法兰", "法兰颈部高度")+get_value(data, "壳体法兰", "法兰名义厚度")
        if None not in (w, n, h):
            return f"Ø{w}/Ø{n}；H={h}"
    elif component_name == "管箱平盖":
        w = get_value(data, "管箱平盖", "法兰名义外径")
        h = get_value(data, "壳体法兰", "法兰名义厚度")
        if None not in (w, h):
            return f"Ø{w}；H={h}"
    elif component_name == "平盖垫片":
        w = get_value(data, "头盖法兰", "垫片名义外径")
        n = get_value(data, "头盖法兰", "垫片名义内径")
        if None not in (w, n):
            return f"Ø{w}/Ø{n}"
    elif component_name == "壳体法兰":
        w = get_value(data, "壳体法兰", "法兰名义外径")
        n = get_value(data, "壳体法兰", "法兰名义内径")
        h = get_value(data, "壳体法兰", "法兰颈部高度")+get_value(data, "壳体法兰", "法兰名义厚度")
        if None not in (w, n, h):
            return f"Ø{w}/Ø{n}；H={h}"

    elif component_name == "壳体圆筒":
        id_ = get_value(data, "壳体圆筒", "圆筒内径")
        t = get_value(data, "壳体圆筒", "圆筒名义厚度")
        l = get_value(data, "壳体圆筒", "圆筒长度")
        if None not in (id_, t, l):
            return f"ID{id_}×{t};L={l}"
    elif component_name == "壳体封头":
        dh = get_value(data, "壳体封头", "封头类型代号")
        d = get_value(data, "壳体封头", "椭圆形封头计算内径")
        t = get_value(data, "壳体封头", "椭圆形封头名义厚度")
        t_plus = get_value(data, "壳体封头", "椭圆形封头最小成型厚度")
        h = get_value(data, "壳体封头", "椭圆形封头直边高度")
        if None not in (dh,d, t, h):
            return f"{dh}{d}×{t}({t_plus});h={h}"


    elif component_name == "固定鞍座":
        conn = get_connection(**db_config1)
        cursor = conn.cursor()
        # 获取鞍座型式代号（dh）
        cursor.execute("""
            SELECT 参数值 
            FROM 产品设计活动表_元件附加参数表 
            WHERE 产品ID = %s AND 元件名称 = '固定鞍座' AND 参数名称 = '鞍座型式代号'
            LIMIT 1
        """, (product_id,))
        row_dh = cursor.fetchone()
        dh = row_dh["参数值"] if row_dh and row_dh.get("参数值") not in (None, "", "None") else None
        # 获取鞍座高度h
        cursor.execute("""
            SELECT 参数值 
            FROM 产品设计活动表_元件附加参数表 
            WHERE 产品ID = %s AND 元件名称 = '固定鞍座' AND 参数名称 = '鞍座高度h'
            LIMIT 1
        """, (product_id,))
        row_h = cursor.fetchone()
        h = row_h["参数值"] if row_h and row_h.get("参数值") not in (None, "", "None") else None
        if dh is not None and h is not None:
            return f"{dh},h={h}"
        elif dh is not None:
            return f"{dh}"
        elif h is not None:
            return f"h={h}"
        else:
            return ""

    elif component_name == "滑动鞍座":
        conn = get_connection(**db_config1)
        cursor = conn.cursor()
        # 获取鞍座型式代号（dh）
        cursor.execute("""
            SELECT 参数值 
            FROM 产品设计活动表_元件附加参数表 
            WHERE 产品ID = %s AND 元件名称 = '滑动鞍座' AND 参数名称 = '鞍座型式代号'
            LIMIT 1
        """, (product_id,))
        row_dh = cursor.fetchone()
        dh = row_dh["参数值"] if row_dh and row_dh.get("参数值") not in (None, "", "None") else None
        # 获取鞍座高度h
        cursor.execute("""
            SELECT 参数值 
            FROM 产品设计活动表_元件附加参数表 
            WHERE 产品ID = %s AND 元件名称 = '滑动鞍座' AND 参数名称 = '鞍座高度h'
            LIMIT 1
        """, (product_id,))
        row_h = cursor.fetchone()
        h = row_h["参数值"] if row_h and row_h.get("参数值") not in (None, "", "None") else None
        if dh is not None and h is not None:
            return f"{dh},h={h}"
        elif dh is not None:
            return f"{dh}"
        elif h is not None:
            return f"h={h}"
        else:
            return ""
    elif component_name == "螺柱（管箱法兰）":
        dh = get_value(data, "管箱法兰", "螺栓公称直径")

        if dh is None:
            return None

        try:
            dh_val = int(re.search(r'\d+', str(dh)).group())
        except:
            dh_val = 0

        flange_thk_1 = get_value(data, "管箱法兰", "法兰名义厚度") or 0
        gasket_thk_1 = get_value(data, "管箱法兰", "垫片厚度") or 0
        flange_thk_2 = get_value(data, "壳体法兰", "法兰名义厚度") or 0
        gasket_thk_2 = get_value(data, "壳体法兰", "垫片厚度") or 0
        ttgd = get_ttgd_from_db(product_id) or 0

        l = 20 + 2 * dh_val + flange_thk_1 + gasket_thk_1 + flange_thk_2 + gasket_thk_2 - 2 * ttgd

        return f"{dh}x{l}"
    elif component_name == "螺母（管箱法兰）":
        dh = get_value(data, "管箱法兰", "螺栓公称直径")
        if dh is not None:
            return f"{dh}"
    elif component_name == "螺柱（管箱平盖）":
        dh = get_value(data, "管箱平盖", "螺栓公称直径")

        if dh is None:
            return None

        try:
            dh_val = int(re.search(r'\d+', str(dh)).group())
        except:
            dh_val = 0

        flange_thk_1 = get_value(data, "管箱平盖", "法兰名义厚度") or 0
        gasket_thk_1 = get_value(data, "管箱平盖", "垫片厚度") or 0
        flange_thk_2 = get_value(data, "头盖法兰", "法兰名义厚度") or 0
        gasket_thk_2 = get_value(data, "头盖法兰", "垫片厚度") or 0
        ttgd = get_ttgd_from_db(product_id) or 0

        l = 20 + 2 * dh_val + flange_thk_1 + gasket_thk_1 + flange_thk_2 + gasket_thk_2 - 2 * ttgd

        return f"{dh}x{l}"
    elif component_name == "螺母（管箱平盖）":
        dh = get_value(data, "管箱平盖", "螺栓公称直径")
        if dh is not None:
            return f"{dh}"
    # elif component_name == "接管(钢管)":
    #     dh = get_value(data, "管程入口接管", "接管外径")
    #     bh = get_value(data, "管程入口接管", "接管外径")
    #     l = get_value(data, "管程入口接管", "接管实际外伸长度")+get_value(data, "管程入口接管", "接管实际内伸长度")
    #     if None not in (dh, bh):
    #         return f"OD{dh}×{bh};L={l}"
    # elif component_name == "接管(钢管)":
    #     dh = get_value(data, "管程入口接管", "接管外径")
    #     bh = get_value(data, "管程入口接管", "接管名义厚度")
    #     l = get_value(data, "管程入口接管", "接管实际外伸长度")+get_value(data, "管程入口接管", "接管实际内伸长度")
    #     if None not in (dh, bh):
    #         return f"OD{dh}×{bh};L={l}"
    # elif component_name == "接管(钢板)":
    #     dh = get_value(data, "管程入口接管", "接管外径")
    #     bh = get_value(data, "管程入口接管", "接管名义厚度")
    #     l = get_value(data, "管程入口接管", "接管实际外伸长度")+get_value(data, "管程入口接管", "接管实际内伸长度")
    #     if None not in (dh, bh):
    #         return f"OD{dh}×{bh};L={l}"
    # elif component_name == "接管(钢锻件)":
    #     dh = get_value(data, "管程入口接管", "接管外径")
    #     bh = get_value(data, "管程入口接管", "接管内径")
    #     l = get_value(data, "管程入口接管", "接管实际外伸长度")+get_value(data, "管程入口接管", "接管实际内伸长度")
    #     if None not in (dh, bh):
    #         return f"Ø{dh}/Ø{bh}；L={l}"
    # 你可以在此添加更多规则：
    # elif component_name == "其他元件名称":
    #     return "你定义的规格格式"
    elif component_name == "铭牌支架":
        return "δ=5"
    elif component_name == "铭牌板":
        return "δ=2"
    elif component_name == "铆钉":
        return "Ø3×14"
    elif component_name in {"管程入口接管", "管程出口接管", "壳程入口接管", "壳程出口接管"}:
        print(component_name)
        od = get_value(data, component_name, "接管大端外径")
        thick = get_value(data, component_name, "接管大端壁厚")
        l1 = get_value(data, component_name, "接管实际外伸长度") or 0
        l2 = get_value(data, component_name, "接管实际内伸长度") or 0
        if None not in (od, thick):
            return f"OD{od}×{thick};L={l1 + l2}"

    return None  # 未匹配或数据缺失


# === 写入规格到 Excel ===
def write_spec_to_excel(data, excel_path, sheet_name, product_id):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]

    for row in sheet.iter_rows(min_row=8):  # 从第8行开始
        if len(row) < 5:
            continue
        d_cell = row[3]  # D列
        e_cell = row[4]  # E列

        if d_cell.value:
            name = str(d_cell.value).strip()
            print(name)
            spec = generate_spec(name, data, product_id)
            if spec is not None:
                e_cell.value = spec
            else:
                print(f"⚠️ 无法生成规格：{name}")

    wb.save(excel_path)
    print(f"✅ 已填写规格列至 Excel：{excel_path}")

def get_pipe_param_value(field_name):
    """
    从固定路径的“布管输入参数.json”中获取指定 paramName 对应的 paramValue。

    参数:
        field_name: str - 要查找的参数名称（paramName）

    返回:
        paramValue (str) 或 None
    """
    try:
        # === 1. 读取 config.ini 获取 product_directory 路径 ===
        config_path = os.path.expandvars(r"%APPDATA%\UDS\蓝滨数字化合作\data\config.ini")
        if not os.path.exists(config_path):
            print(f"❌ 配置文件未找到: {config_path}")
            return None

        with open(config_path, 'rb') as f:
            raw = f.read()
            encoding = chardet.detect(raw)['encoding'] or 'utf-8'

        config = configparser.ConfigParser()
        config.read_string(raw.decode(encoding))
        product_dir = os.path.normpath(config.get('ProjectInfo', 'product_directory', fallback=''))

        # === 2. 拼接布管输入参数 JSON 路径 ===
        pipe_json_path = os.path.join(product_dir, "中间数据", "布管输入参数.json")
        if not os.path.exists(pipe_json_path):
            print(f"❌ 未找到布管输入参数文件: {pipe_json_path}")
            return None

        # === 3. 加载 JSON 并查找字段值 ===
        with open(pipe_json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        for item in data:
            if item.get("paramName") == field_name:
                return item.get("paramValue")
    except Exception as e:
        print(f"❌ 读取参数 `{field_name}` 失败: {e}")

    return None
def get_ttgd_from_db(product_id):
    try:
        conn = get_connection(**db_config1)
        cursor = conn.cursor()
        sql = """
            SELECT 参数值
            FROM 产品设计活动表_元件附加参数表
            WHERE 产品ID = %s AND 元件名称 = '固定管板' AND 参数名称 = '管板凸台高度'
        """
        cursor.execute(sql, (product_id,))
        row = cursor.fetchone()
        conn.close()
        if row and "参数值" in row:
            return float(row["参数值"])
    except Exception as e:
        print(f"❌ 获取管板凸台高度失败: {e}")
    return 0  # 默认值为0，避免None参与计算出错

def insert_jiaguan_falan_rows(sheet, product_id, json_data):
    """
    在“管口”行后插入接管法兰行。
    - C列：法兰标准
    - D列：管口功能 + 接管法兰
    - E列：规格
    - H列：材料牌号（从 产品设计活动表_管口零件材料表）
    - L列：供货状态
    - M列：材料类型
    """

    # NPS → DN 映射（字符串形式）
    nps_to_dn = {
        "1/2": "15", "3/4": "20", "1": "25", "1-1/4": "32", "1-1/2": "40", "2": "50",
        "2-1/2": "65", "3": "80", "4": "100", "5": "125", "6": "150", "8": "200",
        "10": "250", "12": "300", "14": "350", "16": "400", "18": "450",
        "20": "500", "24": "600"
    }

    try:
        conn = get_connection(**db_config1)
        cursor = conn.cursor()

        # 1️⃣ 查询接管法兰主参数（法兰表）
        sql_main = """
            SELECT 法兰标准, 管口功能, 公称尺寸, 压力等级, 法兰型式, 密封面型式, 焊端规格
            FROM 产品设计活动表_管口表
            WHERE 产品ID = %s
        """
        cursor.execute(sql_main, (product_id,))
        rows = cursor.fetchall()

        if not rows:
            print("⚠️ 数据库中未找到接管法兰数据")
            conn.close()
            return

        # 2️⃣ 查询接管法兰 材料信息（零件材料表）
        sql_mat = """
            SELECT 材料牌号, 供货状态, 材料类型
            FROM 产品设计活动表_管口零件材料表
            WHERE 产品ID = %s AND 零件名称 = '接管法兰'
        """
        cursor.execute(sql_mat, (product_id,))
        mat_row = cursor.fetchone()
        conn.close()

        # 如果没有查到，也允许空值
        mat_grade = mat_row.get("材料牌号", "") if mat_row else ""
        supply_status = mat_row.get("供货状态", "") if mat_row else ""
        mat_type = mat_row.get("材料类型", "") if mat_row else ""

        # 3️⃣ 定位“管口”行
        insert_index = None
        for idx, row in enumerate(sheet.iter_rows(min_row=8), start=8):
            d_val = str(row[3].value).strip()
            if d_val == "管口":
                insert_index = idx + 1
                break

        if insert_index is None:
            print("❌ 未找到“管口”行，无法插入接管法兰")
            return

        # 4️⃣ 倒序插入并填写
        for data in reversed(rows):
            sheet.insert_rows(insert_index)

            standard = str(data.get("法兰标准", "")).strip()
            function = str(data.get("管口功能", "")).strip()
            dn = str(data.get("公称尺寸", "")).strip()
            pn = str(data.get("压力等级", "")).strip()
            flange_type = str(data.get("法兰型式", "")).strip()
            face_type = str(data.get("密封面型式", "")).strip()
            # 🔍 从 JSON 中提取焊端规格
            handuan_type = ""
            jiaguan_key = function + "接管"
            try:
                datas = json_data.get("DictOutDatas", {}).get(jiaguan_key, {}).get("Datas", [])
                for item in datas:
                    if item.get("Name") == "接管与管法兰或外部连接端壁厚（焊端规格）":
                        handuan_type = str(item.get("Value", "")).strip()
                        break
                else:
                    for item in datas:
                        if item.get("Name") == "接管与管法兰或外部连接端壁厚（焊端规格）":
                            handuan_type = str(item.get("Value", "")).strip()
                            break
            except Exception as e:
                print(f"⚠️ 获取 {jiaguan_key} 焊端规格失败: {e}")

            # 替换公称尺寸为 DN（若符合）
            dn = nps_to_dn.get(dn, dn)

            # C列
            sheet.cell(row=insert_index, column=3).value = standard
            # D列
            sheet.cell(row=insert_index, column=4).value = f"{function}接管法兰"

            # E列：规格
            if standard == "HG/T 20615-2009":
                spec = f"{flange_type} {dn}-{pn} {face_type} s={handuan_type}mm"
                print(flange_type)
            elif standard == "HG/T 20592-2009":
                spec = f"{flange_type} {dn}-{pn} {face_type} s={handuan_type}mm"
                print(flange_type)

            else:
                spec = f"{dn}-{pn} {flange_type} {face_type}"
            # G列：数量
            sheet.cell(row=insert_index, column=7).value = 1

            sheet.cell(row=insert_index, column=5).value = spec
            # H列（第8列）：材料牌号
            sheet.cell(row=insert_index, column=6).value = mat_grade
            # L列（第12列）：供货状态
            sheet.cell(row=insert_index, column=10).value = supply_status
            # M列（第13列）：材料类型
            sheet.cell(row=insert_index, column=11).value = mat_type

        print(f"✅ 已插入接管法兰 {len(rows)} 条，含材料信息")
        # === 四个接管法兰质量写入 ===
        try:
            conn1 = pymysql.connect(
                host="localhost", user="root", password="123456",
                database="产品设计活动库", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
            )
            conn2 = pymysql.connect(
                host="localhost", user="root", password="123456",
                database="材料库", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
            )

            # Step 1: 获取 公称尺寸类型、公称压力类型
            with conn1.cursor() as cursor:
                cursor.execute("""
                    SELECT 公称尺寸类型, 公称压力类型 
                    FROM 产品设计活动表_管口类型选择表 
                    WHERE 产品ID = %s LIMIT 1
                """, (product_id,))
                config = cursor.fetchone()
                size_type = config.get("公称尺寸类型", "DN").strip()
                press_type = config.get("公称压力类型", "PN").strip()

            # Step 2: 获取 N1~N4 管口信息
            with conn1.cursor() as cursor:
                cursor.execute("""
                    SELECT 管口代号, 公称尺寸, 压力等级, 法兰型式 
                    FROM 产品设计活动表_管口表 
                    WHERE 产品ID = %s AND 管口代号 IN ('N1', 'N2', 'N3', 'N4')
                """, (product_id,))
                kou_rows = cursor.fetchall()

            # Step 3: 查询 材料库.管法兰质量表
            flange_mass_map = {}
            with conn2.cursor() as cursor2:
                for row in kou_rows:
                    kou_id = row["管口代号"]
                    size = str(row["公称尺寸"]).strip()
                    pressure = str(row["压力等级"]).strip()
                    flange_type = row["法兰型式"].strip()

                    standard = "20592" if press_type == "PN" else "20615"
                    size_col = "DN" if size_type == "DN" else "NPS"
                    press_col = "PN" if press_type == "PN" else "Class"

                    cursor2.execute(f"""
                        SELECT 质量 FROM 管法兰质量表
                        WHERE 标准 = %s AND 法兰型式代号 = %s AND `{size_col}` = %s AND `{press_col}` = %s
                        LIMIT 1
                    """, (standard, flange_type, size, pressure))
                    res = cursor2.fetchone()
                    flange_mass_map[kou_id] = float(res["质量"]) if res and res.get("质量") else 0.0
                    print("✅ flange_mass_map =", flange_mass_map)

            conn1.close()
            conn2.close()

            # Step 4: 写入到 Excel 对应行
            for row in sheet.iter_rows(min_row=2):
                part_name = str(row[3].value).strip()
                print(f"【检查行名】第{row[0].row}行: '{part_name}'")

                if part_name == "管程入口接管法兰":
                    row[7].value = flange_mass_map.get("N1", 0)
                elif part_name == "管程出口接管法兰":
                    row[7].value = flange_mass_map.get("N2", 0)
                elif part_name == "壳程入口接管法兰":
                    row[7].value = flange_mass_map.get("N3", 0)
                elif part_name == "壳程出口接管法兰":
                    row[7].value = flange_mass_map.get("N4", 0)

        except Exception as e:
            print(f"❌ 获取接管法兰质量失败: {e}")
    except Exception as e:
        print(f"❌ 插入接管法兰时出错: {e}")

import json

def insert_jiaguan_rows(sheet, product_id, data, jisuan_json_path):
    """
    在“管口”行后插入接管行。
    每行包括：
    - D列：管口功能接管
    - E列：规格（依据材料类型判断格式）
    - G列：数量（默认为 1）
    - H列：材料牌号
    - L列：供货状态
    - M列：材料类型
    """

    # === 读取计算结果 JSON 文件 ===
    try:
        with open(jisuan_json_path, "r", encoding="utf-8") as f:
            jisuan_data = json.load(f)
            dict_out = jisuan_data.get("DictOutDatas", {})
    except Exception as e:
        print(f"❌ 无法读取计算结果 JSON: {e}")
        dict_out = {}

    # === 获取材料信息 ===
    conn = get_connection(**db_config1)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 材料牌号, 供货状态, 材料类型
        FROM 产品设计活动表_管口零件材料表
        WHERE 产品ID = %s AND 零件名称 = '接管'
    """, (product_id,))
    mat_row = cursor.fetchone()
    conn.close()

    mat_grade = mat_row.get("材料牌号", "") if mat_row else ""
    supply_status = mat_row.get("供货状态", "") if mat_row else ""
    mat_type = mat_row.get("材料类型", "") if mat_row else ""

    # === 找到“管口”行 ===
    insert_index = None
    for idx, row in enumerate(sheet.iter_rows(min_row=8), start=8):
        if str(row[3].value).strip() == "管口":
            insert_index = idx + 1
            break
    if insert_index is None:
        print("❌ 未找到“管口”行，无法插入接管")
        return

    # === 固定四个接管名称 ===
    jieguan_names = ["管程入口接管", "管程出口接管", "壳程入口接管", "壳程出口接管"]

    # === 倒序插入 ===
    for name in reversed(jieguan_names):
        spec = generate_spec(name, data) or ""

        # ⛳ 从计算 JSON 提取该接管的质量
        mass = ""
        module = dict_out.get(name, {})
        datas = module.get("Datas", [])
        for item in datas:
            if item.get("Name", "").strip() == "接管重量":
                mass = item.get("Value", "")
                break

        sheet.insert_rows(insert_index)
        sheet.cell(row=insert_index, column=4).value = name  # D列
        sheet.cell(row=insert_index, column=5).value = spec  # E列
        sheet.cell(row=insert_index, column=6).value = mat_grade  # H列
        sheet.cell(row=insert_index, column=7).value = 1  # G列：数量（写死为1）
        sheet.cell(row=insert_index, column=8).value = mass  # H列：接管重量
        sheet.cell(row=insert_index, column=10).value = supply_status  # L列
        sheet.cell(row=insert_index, column=11).value = mat_type  # M列




from openpyxl.styles import Alignment, Border, Side, Font

def clean_and_renumber(sheet):
    """
    删除指定结构件行，重新编号 A列，并设置格式（居中、边框、字体）。
    只编号到 D列有值的最后一行。
    """
    names_to_remove = {
        "螺母（保温支撑）", "螺柱（保温支撑）",
        "底板（固定鞍座）", "腹板（固定鞍座）", "筋板（固定鞍座）", "垫板（固定鞍座）",
        "底板（滑动鞍座）", "腹板（滑动鞍座）", "筋板（滑动鞍座）", "垫板（滑动鞍座）",
        # 新增结构件名称 ↓↓↓
        "支撑板（保温支撑）", "支撑环（保温支撑）", "支撑条（保温支撑）",
        "环首螺钉", "接地板/接地端子", "管口",
        "顶丝", "顶板", "堵板", "破涡器",
        "尾部支撑", "管箱吊耳","防冲板"
    }

    # 设置样式：边框、居中、字体
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    font_10 = Font(size=10)

    # 1️⃣ 删除指定结构件行
    rows_to_delete = []
    for idx, row in enumerate(sheet.iter_rows(min_row=8), start=8):
        d_val = str(row[3].value).strip() if row[3].value else ""
        if d_val in names_to_remove:
            rows_to_delete.append(idx)
    for idx in reversed(rows_to_delete):
        sheet.delete_rows(idx)

    # 2️⃣ 重新编号和格式化（从第8行起，遇 D列为空则停止）
    serial = 1
    for row in sheet.iter_rows(min_row=8):
        d_val = row[3].value
        if d_val is None or str(d_val).strip() == "":
            break
        row_idx = row[0].row
        sheet.cell(row=row_idx, column=1).value = serial  # A列编号
        serial += 1

        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = font_10



# === 主函数入口 ===
def main(json_file_path, excel_file_path, sheet_name, product_id):
    import openpyxl
    from openpyxl.cell.cell import MergedCell

    data = load_json_data(json_file_path)
    write_spec_to_excel(data, excel_file_path, sheet_name, product_id)

    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb[sheet_name]
    insert_jiaguan_falan_rows(sheet, product_id,data)
    insert_jiaguan_rows(sheet, product_id, data, "jisuan_output_new.json")
    clean_and_renumber(sheet)

    # ✅ 填充 I 列：G * H（即第7、8列），仅限 D 列有值的行
    for row in sheet.iter_rows(min_row=8):
        if isinstance(row[8], MergedCell):
            continue  # 跳过合并单元格

        d_val = row[3].value
        g_val = row[6].value
        h_val = row[7].value
        i_cell = row[8]

        if d_val and i_cell.value in (None, "", "None"):  # D 列有值且 I 列没填过
            try:
                g = float(g_val) if g_val not in (None, "", "None") else 0
                h = float(h_val) if h_val not in (None, "", "None") else 0
                i_cell.value = round(g * h, 3)
            except:
                i_cell.value = 0

    # ✅ 删除指定名称的无效零件行，并重新编号 A 列
    remove_names = {"旁路挡板", "中间挡板", "防冲板", "挡管"}
    rows_to_delete = []

    for i, row in enumerate(sheet.iter_rows(min_row=8), start=8):
        d_val = str(row[3].value).strip() if row[3].value else ""
        g_val = row[6].value
        if d_val in remove_names:
            if g_val in (None, "", "None", 0, 0.0, "0"):
                rows_to_delete.append(i)

    # 倒序删除以避免索引错乱
    for i in reversed(rows_to_delete):
        sheet.delete_rows(i)

    # ✅ 重排 A 列序号直到 D 列为空
    current_index = 1
    for row in sheet.iter_rows(min_row=8):
        d_val = row[3].value
        if d_val in (None, "", "None"):
            break
        row[0].value = current_index
        current_index += 1
        # ✅ 写入管箱法兰、固定管板、壳体法兰的质量（L-Q列）
        name_field_map = {
            "管箱法兰": ("管箱法兰", "法兰成型质量"),
            "固定管板": ("固定管板", "管板重量-成品"),
            "壳体法兰": ("壳体法兰", "法兰成型质量"),
            "头盖法兰": ("头盖法兰", "法兰成型质量"),
            "管箱平盖": ("管箱平盖", "法兰成型质量"),
        }

        for row in sheet.iter_rows(min_row=8):
            part_name = str(row[3].value).strip() if row[3].value else ""
            if part_name in name_field_map:
                module, key = name_field_map[part_name]
                try:
                    datas = data.get("DictOutDatas", {}).get(module, {}).get("Datas", [])
                    for item in datas:
                        if item.get("Name") == key:
                            val = item.get("Value", "")
                            # 写入 L 列（即 index 11），合并单元格区域 L-Q 只写 L 即可
                            row[11].value = "成型重量："+val
                            break
                except Exception as e:
                    print(f"⚠️ 处理 {part_name} 时出错：{e}")
    wb.save(excel_file_path)



# === 示例调用 ===
if __name__ == "__main__":
    main("jisuan_output_new.json", "材料清单_已填.xlsx", "Sheet1")

import json
import os

import chardet
import configparser
from openpyxl import load_workbook

import json
from openpyxl import load_workbook
import pymysql

from modules.chanpinguanli.chanpinguanli_main import product_manager
from modules.wenbenshengcheng.CalculateReport import generate_calReport

product_id = None


def on_product_id_changed(new_id):
    print(f"Received new PRODUCT_ID: {new_id}")
    global product_id
    product_id = new_id


# 测试用产品 ID（真实情况中由外部输入）
product_manager.product_id_changed.connect(on_product_id_changed)

def get_weld_area(product_id):
    db_config = {
        "host": "localhost",
        "user": "root",
        "password": "123456",
        "database": "产品设计活动库",
        "charset": "utf8mb4",
        "cursorclass": pymysql.cursors.DictCursor
    }
    conn = pymysql.connect(**db_config)
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 参数值 
                FROM 产品设计活动表_管口零件材料参数表 
                WHERE 产品ID = %s AND 参数名称 = '焊缝金属截面积'
            """, (product_id,))
            row = cursor.fetchone()
            return row["参数值"] if row and row["参数值"].strip() else "0"
    finally:
        conn.close()
def get_jietouxishu_data(product_id):
    db_config = {
        "host": "localhost",
        "user": "root",
        "password": "123456",
        "database": "产品设计活动库",
        "charset": "utf8mb4",
        "cursorclass": pymysql.cursors.DictCursor
    }
    conn = pymysql.connect(**db_config)
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 管程数值 
                FROM 产品设计活动表_设计数据表
                WHERE 产品ID = %s AND 参数名称 = "焊接接头系数*"
            """, (product_id,))
            row = cursor.fetchone()
            return row["管程数值"] if row and row["管程数值"].strip() else "0"
    finally:
        conn.close()
def get_pinggai_data(product_id):
    db_config = {
        "host": "localhost",
        "user": "root",
        "password": "123456",
        "database": "产品设计活动库",
        "charset": "utf8mb4",
        "cursorclass": pymysql.cursors.DictCursor
    }
    conn = pymysql.connect(**db_config)
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 参数值 
                FROM 产品设计活动表_元件附加参数表
                WHERE 产品ID = %s AND 元件名称 = "管箱平盖" AND 参数名称 = '覆层材料牌号'
            """, (product_id,))
            row = cursor.fetchone()
            if row and row.get("参数值") not in (None, "", "None"):
                return row["参数值"].strip()
            else:
                return "0"
    finally:
        conn.close()
def fill_calculation_report(json_path, excel_path, output_path):
    """
    综合写入逻辑：
    - 将每个模块中的 Id、Name、Value 分别写入 A/B/C 列（第1/2/3列）；
    - 同时根据 Id+Name 匹配，将 Value 写入 C列；
    - 特殊处理：固定管板合并壳体法兰字段，法兰/平盖做归属区分。
    """
    with open(json_path, "r", encoding="utf-8") as f:
        json_data = json.load(f)

    dict_out_data = json_data.get("DictOutDatas", {})
    wb = load_workbook(excel_path)

    def extract_flange_fields(flange_module_name, expected_value, include_suffix):
        result_map = {}
        flange_module = dict_out_data.get(flange_module_name, {})
        flange_datas = flange_module.get("Datas", [])

        id_to_position = {
            item["Id"]: item["Value"]
            for item in flange_datas
            if item.get("Name") == "法兰位置" and item.get("Id", "").startswith("m_NameFl")
        }

        for item in flange_datas:
            id_ = item.get("Id", "").strip()
            name = item.get("Name", "").strip()
            val = item.get("Value", "")
            if not id_ or not name:
                continue

            has_suffix = id_.endswith("2")
            if has_suffix != include_suffix:
                continue

            ref_id = "m_NameFl2" if has_suffix else "m_NameFl"
            if id_to_position.get(ref_id, "") != expected_value:
                continue

            result_map[(id_, name)] = val

        return result_map

    for sheet_name in wb.sheetnames:
        if sheet_name not in dict_out_data:
            print(f"⚠️ JSON 中未找到模块：{sheet_name}，跳过该表")
            continue

        module_data = dict_out_data[sheet_name]
        datas = module_data.get("Datas", [])
        if not isinstance(datas, list):
            print(f"⚠️ 模块 `{sheet_name}` 数据无效，跳过")
            continue

        sheet = wb[sheet_name]
        print(f"✅ 正在写入模块：{sheet_name}")

        # 清空 A/B/C 列内容
        for row in sheet.iter_rows(min_row=2):
            for cell in row[:3]:
                cell.value = None

        # Step 1: 写入 A/B/C 列
        for idx, item in enumerate(datas, start=2):
            sheet.cell(row=idx, column=1, value=item.get("Id", ""))
            sheet.cell(row=idx, column=2, value=item.get("Name", ""))
            sheet.cell(row=idx, column=3, value=item.get("Value", ""))

        # Step 2: 构建 (Id, Name) → Value 映射
        id_name_to_value_map = {}
        for item in datas:
            id_ = item.get("Id", "").strip()
            name = item.get("Name", "").strip()
            value = item.get("Value", "")
            if id_ and name:
                id_name_to_value_map[(id_, name)] = value


        # ✅ 固定管板：合并部分壳体法兰字段
        if sheet_name == "固定管板":
            extra_fields = {
                "垫片系数", "比压力", "垫片有效外径", "垫片有效内径", "垫片压紧力作用中心圆直径DG",
                "垫片名义内径", "垫片名义外径", "垫片"
            }
            flange_data = dict_out_data.get("壳体法兰", {}).get("Datas", [])
            for item in flange_data:
                id_ = item.get("Id", "").strip()
                name = item.get("Name", "").strip()
                value = item.get("Value", "")
                if name in extra_fields and id_:
                    id_name_to_value_map[(id_, name)] = value

        # ✅ 法兰归属处理
        if sheet_name in ("壳体法兰", "管箱法兰"):
            is_pipe_flange = (sheet_name == "管箱法兰")
            extracted = extract_flange_fields(
                "壳体法兰",
                "壳体法兰" if is_pipe_flange else "管箱法兰",
                include_suffix=is_pipe_flange
            )
            id_name_to_value_map.update(extracted)

        if sheet_name == "管箱平盖":
            extracted = extract_flange_fields("管箱平盖", "管箱平盖", include_suffix=True)
            id_name_to_value_map.update(extracted)

        # Step 3: 根据 A列和B列匹配 (Id, Name) → 填入 C列
        for row in sheet.iter_rows(min_row=2):
            id_val = str(row[0].value).strip() if row[0].value else ""
            name_val = str(row[1].value).strip() if row[1].value else ""
            key = (id_val, name_val)
            if key in id_name_to_value_map:
                row[2].value = id_name_to_value_map[key]
            else:
                print(f"❌ `{sheet_name}` 中未匹配字段名：{key}")

        # Step 4: 追加表中不存在的字段 (Id, Name) 对
        existing_pairs = {
            (str(row[0].value).strip(), str(row[1].value).strip())
            for row in sheet.iter_rows(min_row=2)
            if row[0].value and row[1].value
        }

        for (id_, name), value in id_name_to_value_map.items():
            if (id_, name) not in existing_pairs:
                next_row = sheet.max_row + 1
                sheet.cell(row=next_row, column=1, value=id_)
                sheet.cell(row=next_row, column=2, value=name)
                sheet.cell(row=next_row, column=3, value=value)

    wb.save(output_path)
    print(f"✅ 综合填充完成，保存为：{output_path}")
# def auto_map_bool(val):
#     if str(val).lower() == "true":
#         return "是"
#     elif str(val).lower() == "false":
#         return "否"
#     return val  # 其他值保持不变

MODULE_TO_SHEET_MAP = {
    "管箱法兰": "管箱法兰",
    "壳体法兰": "壳体法兰",
    "管箱平盖": "管箱平盖",
    "壳体平盖": "壳体平盖",
    "平盖": "管箱平盖",   # ✅
}



def write_flange_values(intermediate_excel_path, target_wb):
    from openpyxl import load_workbook
    from collections import defaultdict

    inter_wb = load_workbook(intermediate_excel_path, data_only=True)

    for sheet in inter_wb.worksheets:
        rows = list(sheet.iter_rows(min_row=2))
        id_name_value_list = [(str(r[0].value).strip(), str(r[1].value).strip(), r[2].value)
                              for r in rows if r[0].value and r[1].value]

        # 获取归属定义
        m_NameFl_raw = None
        m_NameFl2_raw = None
        for id_, name, value in id_name_value_list:
            if id_ == "m_NameFl":
                m_NameFl_raw = str(value).strip()
            elif id_ == "m_NameFl2":
                m_NameFl2_raw = str(value).strip()

        if not m_NameFl_raw or not m_NameFl2_raw:
            print(f"⚠️ `{sheet.title}` 缺少 m_NameFl 或 m_NameFl2，跳过")
            continue

        # 做映射
        m_NameFl_val = MODULE_TO_SHEET_MAP.get(m_NameFl_raw, m_NameFl_raw)
        m_NameFl2_val = MODULE_TO_SHEET_MAP.get(m_NameFl2_raw, m_NameFl2_raw)

        # 分组
        short_id_items = defaultdict(str)
        long_id_items = defaultdict(str)

        for id_, name, value in id_name_value_list:
            if id_ in ("m_NameFl", "m_NameFl2"):
                continue
            if id_.endswith("2"):
                long_id_items[name] = value
            else:
                short_id_items[name] = value

        # 写入短 ID（m_NameFl 对应模块）
        if m_NameFl_val in target_wb.sheetnames:
            sheet1 = target_wb[m_NameFl_val]
            for name, value in short_id_items.items():
                _write_to_sheet_by_name(sheet1, name, value)
        else:
            print(f"❌ 未找到目标 sheet：{m_NameFl_val}")

        # 写入长 ID（m_NameFl2 对应模块）
        if m_NameFl2_val in target_wb.sheetnames:
            sheet2 = target_wb[m_NameFl2_val]
            for name, value in long_id_items.items():
                _write_to_sheet_by_name(sheet2, name, value)
        else:
            print(f"❌ 未找到目标 sheet：{m_NameFl2_val}")

        print(f"✅ `{sheet.title}` 字段写入完成 → {m_NameFl_val} / {m_NameFl2_val}")


def _write_to_sheet_by_name(sheet, name, value):
    for row in sheet.iter_rows(min_row=2):
        if str(row[2].value).strip() == name:
            row[3].value = value
            return


# === 仅这些字段允许做 “是/否” 映射 ===
bool_field_names = {
    "是否以外径为基准",
    "Pt与Ps是否同时作用",
    "是否需要另加补强",
    "结论",
    "校核条件"
}

def fill_final_excel_from_intermediate(intermediate_excel_path, target_excel_path, output_excel_path, json_path):
    import json
    import os
    import configparser
    import chardet
    import pymysql
    from openpyxl import load_workbook

    # === 用户提供 product_id 外部变量 ===
    global product_id

    # === 工具函数：获取焊缝金属截面积 ===
    def get_weld_area(product_id):
        conn = pymysql.connect(
            host="localhost",
            user="root",
            password="123456",
            database="产品设计活动库",
            charset="utf8mb4",
            cursorclass=pymysql.cursors.DictCursor
        )
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT 参数值 
                    FROM 产品设计活动表_元件附加参数表
                    WHERE 产品ID = %s AND 元件名称 = '固定管板' AND 参数名称 = '焊缝金属截面积A3'
                """, (product_id,))
                row = cursor.fetchone()
                return str(row["参数值"]) if row and row["参数值"] not in (None, "", "None") else "0"
        finally:
            conn.close()

    # === 工具函数：bool 自动映射 ===
    def auto_map_bool(val):
        if str(val).strip() in ("True", "true", "1"):
            return "是"
        if str(val).strip() in ("False", "false", "0"):
            return "否"
        return val

    # === 工具函数：复制名义厚度 ===
    def copy_nominal_thickness(from_module, to_module, target_wb):
        from_sheet = target_wb.get_sheet_by_name(from_module)
        to_sheet = target_wb.get_sheet_by_name(to_module)
        if not from_sheet or not to_sheet:
            return

        value_map = {}
        for row in from_sheet.iter_rows(min_row=2):
            name_cell = row[0]
            val_cell = row[3]
            if name_cell.value and val_cell.value:
                value_map[str(name_cell.value).strip()] = val_cell.value

        for row in to_sheet.iter_rows(min_row=2):
            name_cell = row[0]
            val_cell = row[3]
            key = str(name_cell.value).strip() if name_cell.value else ""
            if key in value_map and not val_cell.value:
                val_cell.value = value_map[key]

    # === 工具函数：写入法兰字段（简化处理，只支持 Name→Value 写入） ===
    def write_flange_values(intermediate_excel_path, target_wb):
        inter_wb = load_workbook(intermediate_excel_path, data_only=True)
        for sheet in inter_wb.worksheets:
            rows = list(sheet.iter_rows(min_row=2))
            id_name_value_list = [(str(r[0].value).strip(), str(r[1].value).strip(), r[2].value)
                                  for r in rows if r[0].value and r[1].value]
            name_val_map = {name: val for _, name, val in id_name_value_list}
            if sheet.title in target_wb.sheetnames:
                target_sheet = target_wb[sheet.title]
                for row in target_sheet.iter_rows(min_row=2):
                    name_cell = row[0]
                    val_cell = row[3]
                    if name_cell.value and not val_cell.value:
                        name = str(name_cell.value).strip()
                        if name in name_val_map:
                            val_cell.value = name_val_map[name]

    # === 加载 JSON 判断结论 ===
    with open(json_path, "r", encoding="utf-8") as f:
        json_data = json.load(f)
    dict_out_data = json_data.get("DictOutDatas", {})
    module_success_map = {
        name: data.get("IsSuccess", False)
        for name, data in dict_out_data.items()
        if isinstance(data, dict)
    }

    # === 加载中间 Excel 数据 ===
    inter_wb = load_workbook(intermediate_excel_path, data_only=True)
    inter_data_map = {}
    for sheet in inter_wb.worksheets:
        name_value_map = {}
        for row in sheet.iter_rows(min_row=2):
            name = row[1].value  # B列
            value = row[2].value  # C列
            if name and name not in name_value_map:
                name_value_map[name] = value
        inter_data_map[sheet.title] = name_value_map

    # === 字段映射 ===
    field_reverse_maps = {
        "换热管排列方式(0:30°;1:60°;2:90°;3:45°)": {"0": "正三角形", "1": "转角正三角形", "2": "正方形", "3": "转角正方形"},
        "开孔所属位置": {"1": "圆筒", "2": "椭圆形封头", "3": "锥形封头或锥壳", "4": "平封头(平板）", "5": "碟形封头", "6": "球壳"},
        "接管类型": {"1": "圆形", "2": "椭圆形或长圆孔"},
        "开孔方位": {"1": "径向", "2": "斜向", "3": "切向或偏心"},
        "补强类型": {
            "1": "增加筒体厚度", "2": "增加接管厚度", "3": "补强圈补强", "4": "嵌入式接管补强",
            "5": "筒体和接管联合补强", "6": "接管和补强圈联合补强", "7": "筒体和补强圈联合补强", "8": "筒体和接管和补强圈联合补强"
        },
        "接管与壳体连接结构型式": {"1": "插入式", "2": "安放式"},
        "嵌入式接管补强类型": {"1": "a型", "2": "b型", "3": "c型"}
    }
    field_reverse_maps2 = {"压力试验类型": {"1": "液压", "2": "气压", "3": "气液"}}

    target_wb = load_workbook(target_excel_path)
    for sheet in target_wb.worksheets:
        sheet_name = sheet.title
        if sheet_name == "管箱平盖":
            pinggai_paihao = get_pinggai_data(product_id)
            jietouxishu = get_jietouxishu_data(product_id)
            # === 将换热管长度Lt 和 管程数 写入 D列 ===
            for idx in range(2, sheet.max_row + 1):
                c_cell = sheet.cell(row=idx, column=3)
                d_cell = sheet.cell(row=idx, column=4)
                c_val = str(c_cell.value).strip() if c_cell.value else ""

                if c_val == "平盖覆层材料牌号" and not d_cell.value:
                    if pinggai_paihao == "0":
                        pinggai_paihao = '-'
                    d_cell.value = pinggai_paihao

                    print(f"📌 写入 平盖覆层材料牌号 → {pinggai_paihao}")
                if c_val == "焊接接头系数ф" and not d_cell.value:
                    d_cell.value = jietouxishu
                    print(f"📌 写入 焊接接头系数ф → {jietouxishu}")
        # ✅ 固定管板额外字段写入（来自布管输入参数 JSON）
        if sheet_name == "固定管板":
            # === 从布管输入参数中读取值 ===
            config_path = os.path.expandvars(r"%APPDATA%\UDS\蓝滨数字化合作\data\config.ini")
            with open(config_path, 'rb') as f:
                raw = f.read()
                encoding = chardet.detect(raw)['encoding'] or 'utf-8'
            config = configparser.ConfigParser()
            config.read_string(raw.decode(encoding))
            product_dir = os.path.normpath(config.get('ProjectInfo', 'product_directory', fallback=''))
            tube_json_path = os.path.join(product_dir, "中间数据", "布管输入参数.json")

            tube_length = ""
            tube_pass_count = ""

            if os.path.exists(tube_json_path):
                with open(tube_json_path, 'r', encoding='utf-8') as f:
                    tube_json_data = json.load(f)
                params_list = tube_json_data if isinstance(tube_json_data, list) else tube_json_data.get("params", [])
                for param in params_list:
                    if param.get("paramName") == "换热管公称长度LN":
                        tube_length = str(param.get("paramValue", ""))
                    elif param.get("paramName") == "管程数":
                        tube_pass_count = str(param.get("paramValue", ""))

            # === 将换热管长度Lt 和 管程数 写入 D列 ===
            for idx in range(2, sheet.max_row + 1):
                c_cell = sheet.cell(row=idx, column=3)
                d_cell = sheet.cell(row=idx, column=4)
                c_val = str(c_cell.value).strip() if c_cell.value else ""

                if c_val == "换热管长度Lt" and not d_cell.value:
                    d_cell.value = tube_length
                    print(f"📌 写入 换热管长度Lt → {tube_length}")
                elif c_val == "管程数" and not d_cell.value:
                    d_cell.value = tube_pass_count
                    print(f"📌 写入 管程数 → {tube_pass_count}")

        # ✅ 特殊处理：换热管内压/外压/水压
        special_sheet_map = {
            "换热管内压": "内压",
            "换热管外压计算报告1": "外压",
            "换热管外压计算报告2": "水压"
        }

        if sheet_name in special_sheet_map:
            keyword = special_sheet_map[sheet_name]
            fixed_data = inter_data_map.get("固定管板", {})
            inter_sheet = inter_wb["固定管板"]
            filtered_map = {}
            for row in inter_sheet.iter_rows(min_row=2):
                id_val = str(row[0].value).strip() if row[0].value else ""
                name = str(row[1].value).strip() if row[1].value else ""
                value = row[2].value
                if keyword in id_val and name:
                    filtered_map[name] = value
            name_value_map = filtered_map.copy()
            for k, v in fixed_data.items():
                if k not in name_value_map:
                    name_value_map[k] = v
        elif sheet_name in inter_data_map:
            name_value_map = inter_data_map[sheet_name]
        else:
            print(f"⚠️ 跳过目标表 `{sheet_name}`，因中间结果中无对应 sheet")
            continue

        print(f"📄 正在处理目标表：{sheet_name}")
        for row in sheet.iter_rows(min_row=2):
            keyword_cell = row[0]
            output_cell = row[3]
            keyword = keyword_cell.value

            if keyword == "焊缝金属截面积A3":
                output_cell.value = get_weld_area(product_id)
                continue

            if keyword in name_value_map and (output_cell.value is None or str(output_cell.value).strip() == ""):
                val = name_value_map[keyword]
                if keyword in field_reverse_maps and val in field_reverse_maps[keyword]:
                    val = field_reverse_maps[keyword][val]
                if keyword in field_reverse_maps2 and val in field_reverse_maps2[keyword]:
                    val = field_reverse_maps2[keyword][val]

                # 仅在布尔字段中做“是/否”映射，其他字段保持原值
                if keyword in bool_field_names:
                    print(keyword)
                    output_cell.value = auto_map_bool(val)
                else:
                    output_cell.value = val

            elif keyword in name_value_map:
                print(f"⚠️ `{sheet_name}` 字段 `{keyword}` 已有值，跳过写入")

        # ✅ 填写结论
        if sheet_name in module_success_map:
            result = "合格" if module_success_map[sheet_name] else "不合格"
            for row in sheet.iter_rows(min_row=2):
                if "结论" in str(row[2].value):
                    row[3].value = result
                    print(f"📌 写入结论：{sheet_name} → {result}")

        sheet.column_dimensions['A'].hidden = True

    copy_nominal_thickness("壳体圆筒", "壳体法兰", target_wb)
    copy_nominal_thickness("管箱圆筒", "管箱法兰", target_wb)
    write_flange_values(intermediate_excel_path, target_wb)
    target_wb.save(output_excel_path)
    print(f"✅ 最终Excel已生成：{output_excel_path}")
def copy_nominal_thickness(sheet_from, sheet_to, wb):
    """
    在 sheet_from 中找出 C列为“名义厚度δn”的 D 列值，写入 sheet_to 中 C列为相同内容的那行的 D列。
    """
    try:
        source_value = None

        # 从 sheet_from 中找到名义厚度δn 对应的 D 列值
        for row in wb[sheet_from].iter_rows(min_row=1):
            if len(row) >= 4 and str(row[2].value).strip() == "名义厚度δn":
                source_value = row[3].value  # D列
                break

        if source_value is None:
            print(f"⚠️ 未在工作表 '{sheet_from}' 中找到“名义厚度δn”")
            return

        # 写入 sheet_to 的相同 C列项
        matched = False
        for row in wb[sheet_to].iter_rows(min_row=1):
            if len(row) >= 4 and str(row[2].value).strip() == "名义厚度δn":
                row[3].value = source_value
                matched = True
                print(f"🔁 已将 '{sheet_from}' 中“名义厚度δn”={source_value} 写入 '{sheet_to}'")
                break

        if not matched:
            print(f"⚠️ 未在工作表 '{sheet_to}' 中找到“名义厚度δn”，未能写入")

    except Exception as e:
        print(f"❌ 处理过程中出现错误：{e}")


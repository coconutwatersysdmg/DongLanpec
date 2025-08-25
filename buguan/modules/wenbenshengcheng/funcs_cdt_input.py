from modules.condition_input.funcs.db_cnt import get_connection
from PyQt5.QtWidgets import (QTableWidgetItem, QTableWidget, QHeaderView, QWidget,
                             QMessageBox, QUndoStack, QFileDialog, QComboBox, QStyledItemDelegate)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QColor, QStandardItemModel, QStandardItem, QBrush
import re
import ast
import os
import pandas as pd
from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from modules.condition_input.funcs.undo_command import CellEditCommand
from modules.condition_input.funcs.funcs_def_check import check_dn, check_work_pressure, check_work_temp_in, \
    check_work_temp_out, check_work_pressure_max, check_tubeplate_design_pressure_gap, check_design_pressure2, \
    check_design_pressure, check_design_temp_max, check_design_temp_max2, check_design_temp_min, \
    check_in_out_pressure_gap, check_trail_stand_pressure_medium_density, check_insulation_layer_thickness, \
    check_insulation_material_density, check_def_trail_stand_pressure_lying, check_def_trail_stand_pressure_stand, \
    check_trail_stand_pressure_type

#数据库连接
db_config_1 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': '产品条件库'
}

db_config_2 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': '产品设计活动库'
}

db_config_3 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': '产品需求库'
}

db_config_4 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': '项目需求库'
}

"""导入数据库数据表相关函数"""
def make_header_item(text):
    """
    创建一个“仿真表头”项：
    - 居中对齐
    - 加粗字体
    - 可选中（点击高亮列）
    - 不设置背景颜色（保留原始白色）
    """
    item = QTableWidgetItem(text)
    item.setTextAlignment(Qt.AlignCenter)

    # ✅ 可选中 + 不可编辑（用户可以点击高亮，但不能修改内容）
    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

    # ✅ 设置加粗字体
    font = item.font()
    font.setBold(True)
    item.setFont(font)

    return item

def load_design_data_if_exists(product_id):
    """
    给定产品ID，从设计活动库优先加载5张数据表，如不存在则退回产品条件库模板表。
    """
    design_tables = {
        "产品标准": "产品设计活动表_产品标准数据表",
        "设计数据": "产品设计活动表_设计数据表",
        "通用数据": "产品设计活动表_通用数据表",
        "检测数据": "产品设计活动表_无损检测数据表",
        "涂漆数据": "产品设计活动表_涂漆数据表"
    }

    template_tables = {
        "产品标准": "产品标准数据模板表",
        "设计数据": "设计数据模板表",
        "通用数据": "通用数据模板表",
        "检测数据": "无损检测数据模板表",
        "涂漆数据": "涂漆数据模板表"
    }

    result = {"数据": {}}  # 确保返回数据时有 "数据" 键

    # 判断设计库中是否有记录
    design_data_exists = False
    connection = get_connection(**db_config_2)
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT 1 FROM {design_tables['产品标准']} WHERE 产品ID = %s LIMIT 1", (product_id,))
            design_data_exists = bool(cursor.fetchone())
    finally:
        connection.close()

    # 逐表加载（优先设计库，后退模板库）
    for key in design_tables:
        db_used = db_config_2 if design_data_exists else db_config_1
        table_name = design_tables[key] if design_data_exists else template_tables[key]

        connection = get_connection(**db_used)
        try:
            with connection.cursor() as cursor:
                # 获取字段名，按表类型决定是否保留 参数ID 字段
                cursor.execute(f"DESCRIBE {table_name}")
                columns = cursor.fetchall()

                preserve_param_id = key in ["产品标准", "设计数据", "通用数据"]

                column_names = [
                    col['Field'] for col in columns
                    if (
                        (preserve_param_id or not col['Field'].endswith('参数ID')) and
                        col['Field'] not in ['所属类型', '所属型式'] and
                        '产品ID' not in col['Field'] and
                        '更改状态' not in col['Field']
                    )
                ]

                field_str = ', '.join([f"`{col}`" for col in column_names])
                if design_data_exists:
                    cursor.execute(f"SELECT {field_str} FROM {table_name} WHERE 产品ID = %s", (product_id,))
                else:
                    cursor.execute(f"SELECT {field_str} FROM {table_name}")

                rows = cursor.fetchall()

                # 清洗空值
                for row in rows:
                    for k in row:
                        if row[k] is None:
                            row[k] = ""

                data = {
                    "headers": column_names,
                    "rows": rows,
                    "count": len(rows)
                }

                # 设置界面用的“序号列”字段名（实际用于表格第0列）
                if preserve_param_id:
                    data["prepend_index_header"] = column_names[0]

                if key == "检测数据":
                    data["格式化"] = format_trail_table(column_names, rows)
                if key == "涂漆数据":
                    data["格式化"] = format_coating_table(column_names, rows)

                result["数据"][key] = data  # 存表格数据

            # 设置数据来源状态和导入状态
            result["data_source_status"] = "设计活动库" if design_data_exists else "条件模板"
            result["import_status"] = True if design_data_exists or len(rows) > 0 else False

        finally:
            connection.close()

    return result

def format_trail_table(headers, rows):
    # 将检测数据表按“接头种类”字段进行分组（用于合并同类行显示）
    grouped = {}
    for row in rows:
        接头种类 = row['接头种类']
        if 接头种类 not in grouped:
            grouped[接头种类] = []
        grouped[接头种类].append(row)
    return grouped

def format_coating_table(headers, rows):
    """
    将涂漆数据按“用途”字段进行分组
    并将“用途”字段中的复合值进行拆分（提取出 细类：底漆、中间漆、面漆）
    如：'内涂漆（壳程）_底漆' -> 用途='内涂漆（壳程）', 细类='底漆'
    """
    grouped = {}
    for row in rows:
        用途字段 = row['用途']
        if '）_' in 用途字段:
            左, 右 = 用途字段.split('）_')
            用途 = 左 + '）'     # 例：'内涂漆（壳程）'
            涂层 = 右           # 例：'底漆'
        else:
            用途 = 用途字段
            涂层 = ""

        row['_细类'] = 涂层  # ✅ 注意是临时字段
        if 用途 not in grouped:
            grouped[用途] = []
        grouped[用途].append(row)
    return grouped

def render_grouped_table(table_widget, grouped_data, headers, group_key_column=0):
    header_rows = 2
    total_rows = sum(len(v) for v in grouped_data.values())
    table_widget.setRowCount(total_rows + header_rows)
    table_widget.setColumnCount(len(headers))
    table_widget.setHorizontalHeaderLabels(headers)

    current_row = header_rows
    for group_key, row_list in grouped_data.items():
        span_start = current_row
        for row in row_list:
            for col_idx, key in enumerate(headers):
                if col_idx == group_key_column:
                    continue
                val = str(row.get(key, ""))
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(Qt.ItemIsEditable | Qt.ItemIsSelectable | Qt.ItemIsEnabled)

                detect_method = row.get("检测方法", "").strip()
                # 技术等级为 '/' → 不可编辑
                if detect_method in ["M.T.", "P.T.", "M.T.[FB]"] and key in ["壳程_技术等级", "管程_技术等级"]:
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

                # 合格级别为 M.T./P.T. → 不可编辑
                if detect_method in ["M.T.", "P.T.", "M.T.[FB]"] and key in ["壳程_合格级别", "管程_合格级别"]:
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

                # ✅ 写入默认值作为校验基准（基于 检测方法+比例 反推出来）
                if key.endswith("技术等级") or key.endswith("合格级别"):
                    side = "壳程" if "壳程" in key else "管程"
                    ratio = str(row.get(f"{side}_检测比例", "")).strip()
                    field_type = "技术等级" if "技术等级" in key else "合格级别"
                    from .funcs_cdt_input import compute_trail_default_grade  # 如果在本文件可省略
                    default_val = compute_trail_default_grade(detect_method, ratio, field_type)
                    if default_val:
                        item.setData(Qt.UserRole + 2, default_val)

                table_widget.setItem(current_row, col_idx, item)


            current_row += 1

        # ✅ 设置“接头种类”列为不可编辑
        group_item = QTableWidgetItem(group_key)
        group_item.setTextAlignment(Qt.AlignCenter)
        group_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
        table_widget.setSpan(span_start, group_key_column, len(row_list), 1)
        table_widget.setItem(span_start, group_key_column, group_item)
    table_widget.resizeColumnsToContents()

def set_multilevel_headers(table_widget: QTableWidget, top_headers: list, sub_headers: list, span_map: list):
    """
    设置 QTableWidget 的两级表头结构（不破坏数据内容）。
    - top_headers：一级标题（支持横向合并、纵向合并）
    - sub_headers：二级字段名
    - span_map：格式如 [(start, span)]，表示从哪列开始、合并几列
    """

    col_count = sum(span for _, span in span_map)
    header_rows = 2

    # 创建表头：先扩展一张空表，仅用于设置头部结构（内容之后渲染）
    table_widget.setColumnCount(col_count)
    table_widget.setRowCount(header_rows)  # 只设置前2行用于表头

    # 设置一级标题（带纵向合并）
    for i, (start, span) in enumerate(span_map):
        header_text = top_headers[i] if top_headers[i].strip() else " "
        item = make_header_item(header_text)

        if span == 1:
            table_widget.setSpan(0, start, 2, 1)  # 垂直合并2行
            table_widget.setItem(0, start, item)
        else:
            table_widget.setSpan(0, start, 1, span)  # 水平合并
            table_widget.setItem(0, start, item)

    # 设置子标题
    sub_col = 0
    for i, (start, span) in enumerate(span_map):
        if span > 1:
            for offset in range(span):
                item = make_header_item(sub_headers[sub_col])
                table_widget.setItem(1, start + offset, item)
                sub_col += 1
        else:
            sub_col += 1  # 跳过

    # 不设置内容行，让调用者单独设置数据内容行（从第2行开始）
    table_widget.verticalHeader().setVisible(False)
    table_widget.horizontalHeader().setVisible(False)

def render_coating_table(table_widget: QTableWidget, grouped_data: dict, exec_std_value: str = ""):
    headers = ["用途", "细类", "油漆类别", "颜色", "干膜厚度（μm）", "涂漆面积", "备注"]
    total_data_rows = sum(len(rows) for rows in grouped_data.values())
    table_widget.setRowCount(2 + total_data_rows)
    table_widget.setColumnCount(len(headers))

    all_rows = [row for group in grouped_data.values() for row in group]
    std_value = exec_std_value

    table_widget.verticalHeader().setVisible(False)
    table_widget.horizontalHeader().setVisible(False)

    # ✅ 第一行：执行标准/规范
    table_widget.setSpan(0, 0, 1, 2)
    table_widget.setItem(0, 0, make_header_item("执行标准/规范"))
    std_item = QTableWidgetItem(std_value)
    std_item.setTextAlignment(Qt.AlignCenter)
    std_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
    table_widget.setSpan(0, 2, 1, len(headers) - 2)
    table_widget.setItem(0, 2, std_item)

    # ✅ 第二行：表头
    table_widget.setSpan(1, 0, 1, 2)
    table_widget.setItem(1, 0, make_header_item("用途"))
    for col, header in enumerate(headers[2:], start=2):
        table_widget.setItem(1, col, make_header_item(header))

    current_row = 2
    for group_key, row_list in grouped_data.items():
        span_start = current_row
        merge_data = {"涂漆面积": "", "备注": ""}

        for idx, row in enumerate(row_list):
            values = [
                group_key,
                row.get("_细类", ""),
                row.get("油漆类别", ""),
                row.get("颜色", ""),
                row.get("干膜厚度（μm）", ""),
                row.get("涂漆面积", ""),
                row.get("备注", "")
            ]
            for col, val in enumerate(values):
                val = "" if val is None else str(val)
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)

                # ✅ 设置可编辑性（只用途/细类列是只读）
                if col in (0, 1):
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                else:
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)

                table_widget.setItem(current_row, col, item)

            if idx == 0:
                merge_data["涂漆面积"] = str(row.get("涂漆面积", "") or "")
                merge_data["备注"] = str(row.get("备注", "") or "")

            current_row += 1

        row_count = len(row_list)

        # ✅ 合并用途列
        item = QTableWidgetItem(group_key)
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
        table_widget.setSpan(span_start, 0, row_count, 1)
        table_widget.setItem(span_start, 0, item)

        # ✅ 合并涂漆面积
        area_item = QTableWidgetItem(merge_data["涂漆面积"])
        area_item.setTextAlignment(Qt.AlignCenter)
        area_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
        table_widget.setSpan(span_start, 5, row_count, 1)
        table_widget.setItem(span_start, 5, area_item)

        # ✅ 合并备注
        comment_item = QTableWidgetItem(merge_data["备注"])
        comment_item.setTextAlignment(Qt.AlignCenter)
        comment_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
        table_widget.setSpan(span_start, 6, row_count, 1)
        table_widget.setItem(span_start, 6, comment_item)

    table_widget.resizeColumnsToContents()

    # ✅ 设置 logical_headers：确保校验函数能获取正确列名
    table_widget.logical_headers = [
        "用途", "细类", "油漆类别", "颜色", "干膜厚度（μm）", "涂漆面积", "备注"
    ]

"""表格显示样式"""
def get_merged_cell_start(table_widget, row, col):
    """返回 (row, col) 所属合并单元格的起始行"""
    for r in range(table_widget.rowCount()):
        rowspan = table_widget.rowSpan(r, col)
        if rowspan > 1 and r <= row < r + rowspan:
            return r
    return row

def highlight_entire_row(table_widget):
    selected_indexes = table_widget.selectedIndexes()
    if not selected_indexes:
        return

    selected_rows = {i.row() for i in selected_indexes}
    selected_cols = {i.column() for i in selected_indexes}

    # ✅ 只在真正点击了表头时跳过整行高亮
    row_count = table_widget.rowCount()
    is_full_column_selected = (
        len(selected_cols) == 1 and
        len(selected_rows) >= row_count and
        all(table_widget.model().index(r, list(selected_cols)[0]) in selected_indexes for r in range(row_count))
    )
    if is_full_column_selected:
        return

    # ✅ 清除旧高亮（保持缺失项不动）
    for row in range(table_widget.rowCount()):
        for col in range(table_widget.columnCount()):
            item = table_widget.item(row, col)
            if item:
                if item.data(Qt.UserRole + 1) == "missing":
                    continue
                if row % 2 == 0:
                    item.setBackground(QColor("#ffffff"))
                else:
                    item.setBackground(QColor("#f0f0f0"))
                item.setForeground(QBrush())

    # ✅ 单独处理：合并单元格块（只高亮合并区域）
    for index in selected_indexes:
        row, col = index.row(), index.column()
        rowspan = table_widget.rowSpan(row, col)
        colspan = table_widget.columnSpan(row, col)

        if rowspan > 1 or colspan > 1:
            for r in range(row, row + rowspan):
                for c in range(col, col + colspan):
                    item = table_widget.item(r, c)
                    if item and item.data(Qt.UserRole + 1) != "missing":
                        item.setBackground(QColor("#d0e7ff"))
                        item.setForeground(QBrush(Qt.black))

    # ✅ 收集所有普通格所在的行（跳过合并起始格）
    rows_to_highlight = set()
    for index in selected_indexes:
        row, col = index.row(), index.column()
        rowspan = table_widget.rowSpan(row, col)
        colspan = table_widget.columnSpan(row, col)
        if rowspan == 1 and colspan == 1:
            rows_to_highlight.add(row)

    # ✅ 普通整行高亮（非合并格）
    for row in rows_to_highlight:
        for col in range(table_widget.columnCount()):
            if table_widget.rowSpan(row, col) > 1 or table_widget.columnSpan(row, col) > 1:
                continue  # 跳过合并格
            item = table_widget.item(row, col)
            if item and item.data(Qt.UserRole + 1) != "missing":
                item.setBackground(QColor("#d0e7ff"))
                item.setForeground(QBrush(Qt.black))

def apply_table_style(table_widget):
    table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    table_widget.verticalHeader().setVisible(False)
    table_widget.setAlternatingRowColors(True)
    table_widget.setSelectionBehavior(table_widget.SelectItems)

    # ✅ 为表头加上四边边框线 已修改
    table_widget.horizontalHeader().setStyleSheet("""
        QHeaderView::section {
            border: 1px solid #D8D8D8;        /* 更细更柔和的边框 */
            background-color: white;         /* 白色背景 */
            color: black;                    /* 黑色字体 */
            padding: 4px;                    /* 内边距让文字不挤 */
            font-weight: bold;               /* 加粗字体 */
        }
    """)


#新增
def shrink_index_column(table_widget, width: int = 100):
    """
    将第 0 列（默认是“序号”列）设为较小宽度
    """
    header = table_widget.horizontalHeader()
    header.setSectionResizeMode(0, QHeaderView.Fixed)
    table_widget.setColumnWidth(0, width)
#新增
def shrink_unit_column(table_widget, width: int = 300):
    """
    将第 2 列（默认是“参数单位”列）设为较小宽度
    """
    header = table_widget.horizontalHeader()
    header.setSectionResizeMode(2, QHeaderView.Fixed)
    table_widget.setColumnWidth(2, width)


"""存入数据库相关函数"""

def get_table_header_columns(table_widget):
    headers = []
    for col in range(table_widget.columnCount()):
        item = table_widget.horizontalHeaderItem(col)
        if item:
            true_field = item.data(Qt.UserRole)
            headers.append(true_field if true_field else item.text())
    return headers

def get_table_data(table_widget):
    """
    提取表格所有行数据为结构化列表，每行是一个 dict（包含第0列）
    """
    headers = get_table_header_columns(table_widget)
    data = []

    for row in range(table_widget.rowCount()):
        row_data = {}
        for col_index, header in enumerate(headers):
            item = table_widget.item(row, col_index)
            value = item.text() if item else ""
            row_data[header] = value
        data.append(row_data)

    return data

def save_data_to_database(data, product_id, table_name, table_widget, is_from_design_lib=True):
    """
    将表格数据保存至数据库：
    - 无论是 INSERT 还是 UPDATE，统一先对比模板表字段值，判断更改状态；
    - 更改状态字段统一标记；
    """
    connection = get_connection(**db_config_2)

    try:
        with connection.cursor() as cursor:
            header_columns = get_table_header_columns(table_widget)

            # 获取数据库字段结构
            cursor.execute(f"DESCRIBE {table_name}")
            table_columns = cursor.fetchall()

            # 获取“更改状态”字段名
            change_status_column = None
            for col in table_columns:
                if re.search(r'更改状态$', col['Field']):
                    change_status_column = col['Field']
                    break
            if not change_status_column:
                raise ValueError("未找到更改状态字段")

            # 确定“参数名称”字段
            name_column = "规范/标准名称" if "产品标准" in table_name else "参数名称"

            # 匹配模板表名
            template_table_mapping = {
                "产品设计活动表_产品标准数据表": "产品标准数据模板表",
                "产品设计活动表_设计数据表": "设计数据模板表",
                "产品设计活动表_通用数据表": "通用数据模板表"
            }
            template_table_name = template_table_mapping.get(table_name.replace("", ""), "")

            # 获取模板字段列表（用于对比）
            template_compare_fields = []
            if template_table_name:
                cursor.execute(f"DESCRIBE 产品条件库.{template_table_name}")
                template_compare_fields = [col['Field'] for col in cursor.fetchall()]

            # 数据库中参数ID字段名
            param_id_field = table_columns[0]['Field']
            param_id_column = header_columns[0]

            for row_idx, row in enumerate(data):
                param_name = row.get(name_column)
                if not param_name:
                    continue

                # 获取模板数据行
                template = None
                if template_table_name:
                    cursor.execute(
                        f"SELECT * FROM 产品条件库.{template_table_name} WHERE `{name_column}` = %s",
                        (param_name,)
                    )
                    template = cursor.fetchone()

                # 判断是否与模板数据有差异（更改状态）
                def is_changed(template_row, current_row):
                    if not template_row:
                        return True
                    for key in header_columns:
                        if key not in template_compare_fields:
                            continue  # 忽略“参数ID”等非模板字段
                        cur_val = str(current_row.get(key, "")).strip()
                        tpl_val = str(template_row.get(key, "")).strip()
                        if cur_val != tpl_val:
                            return True
                    return False

                change_detected = is_changed(template, row)

                if is_from_design_lib:
                    # UPDATE 操作
                    cursor.execute(
                        f"SELECT * FROM {table_name} WHERE 产品ID = %s AND `{name_column}` = %s",
                        (product_id, param_name)
                    )
                    existing = cursor.fetchone()
                    if existing:
                        update_values = {}
                        for key in header_columns:
                            new_val = row.get(key, "")
                            old_val = existing.get(key, "")
                            if str(new_val) != str(old_val):
                                update_values[key] = new_val
                        if update_values:
                            update_values[change_status_column] = change_detected
                            update_set = ', '.join([f"`{k}` = %s" for k in update_values])
                            cursor.execute(
                                f"UPDATE {table_name} SET {update_set} WHERE 产品ID = %s AND `{name_column}` = %s",
                                tuple(update_values.values()) + (product_id, param_name)
                            )
                else:
                    # INSERT 操作
                    insert_row = {}
                    for field in [col['Field'] for col in table_columns if col['Extra'] != "auto_increment"]:
                        if field == "产品ID":
                            insert_row[field] = product_id
                        elif field == param_id_field:
                            insert_row[field] = row.get(param_id_column, "")
                        elif field == change_status_column:
                            insert_row[field] = change_detected
                        else:
                            insert_row[field] = row.get(field, "")

                    columns = ', '.join(f"`{k}`" for k in insert_row)
                    placeholders = ', '.join(['%s'] * len(insert_row))
                    cursor.execute(
                        f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})",
                        tuple(insert_row.values())
                    )

        connection.commit()

    finally:
        connection.close()

def save_coating_table_to_database(table_widget: QTableWidget, table_name, product_id: int, source_status: str):
    """
    保存涂漆数据至【产品设计活动表_涂漆数据表】
    - 如果数据来源为条件模板，则执行 INSERT
    - 如果来源为设计活动库，则执行 UPDATE（根据 产品ID + 参数ID 匹配）
    """
    connection = get_connection(**db_config_2)

    try:
        with connection.cursor() as cursor:
            # ✅ 获取执行标准/规范（表格第0行第2列）
            exec_std_item = table_widget.item(0, 2)
            exec_std = exec_std_item.text().strip() if exec_std_item else ""

            id_counter = 1  # 参数ID，从1开始

            row_count = table_widget.rowCount()
            current_row = 2

            while current_row < row_count:
                # ✅ 当前组用途
                usage_item = table_widget.item(current_row, 0)
                current_usage = usage_item.text().strip() if usage_item else ""

                # ✅ 合并列提取：面积、备注
                paint_area_item = table_widget.item(current_row, 5)
                comment_item = table_widget.item(current_row, 6)
                group_paint_area = paint_area_item.text().strip() if paint_area_item else ""
                group_comment = comment_item.text().strip() if comment_item else ""

                sub_row = current_row
                while sub_row < row_count:
                    usage_item_sub = table_widget.item(sub_row, 0)
                    sub_usage = usage_item_sub.text().strip() if usage_item_sub else ""
                    if sub_row != current_row and sub_usage != current_usage:
                        break  # 下一组开始

                    # ✅ 各字段
                    subtype = table_widget.item(sub_row, 1).text().strip() if table_widget.item(sub_row, 1) else ""
                    category = table_widget.item(sub_row, 2).text().strip() if table_widget.item(sub_row, 2) else ""
                    color = table_widget.item(sub_row, 3).text().strip() if table_widget.item(sub_row, 3) else ""
                    thickness = table_widget.item(sub_row, 4).text().strip() if table_widget.item(sub_row, 4) else ""
                    full_usage = f"{current_usage}_{subtype}" if subtype else current_usage

                    if source_status == "条件模板":
                        # ✅ 插入
                        cursor.execute(f"""
                            INSERT INTO {table_name} (
                                `涂漆数据参数ID`, `产品ID`, `用途`, `油漆类别`, `颜色`,
                                `干膜厚度（μm）`, `涂漆面积`, `备注`
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            id_counter,
                            product_id,
                            full_usage,
                            category,
                            color,
                            thickness,
                            group_paint_area,
                            group_comment
                        ))

                    else:  # 来源为“设计活动库” → UPDATE
                        cursor.execute(f"""
                            UPDATE {table_name}
                            SET `用途` = %s,
                                `油漆类别` = %s,
                                `颜色` = %s,
                                `干膜厚度（μm）` = %s,
                                `涂漆面积` = %s,
                                `备注` = %s
                            WHERE `涂漆数据参数ID` = %s AND `产品ID` = %s
                        """, (
                            full_usage,
                            category,
                            color,
                            thickness,
                            group_paint_area,
                            group_comment,
                            id_counter,
                            product_id
                        ))

                    id_counter += 1
                    sub_row += 1

                current_row = sub_row

        connection.commit()

    finally:
        connection.close()

def save_trail_table_to_database(table_widget: QTableWidget, table_name: str, product_id: int, source_status: str):
    """
    保存无损检测数据至【产品设计活动表_无损检测数据表】
    - 支持条件模板插入 or 设计活动库更新
    - 接头种类为合并分组列（需展开）
    - 表格格式为：检测方法、壳程（3列）、管程（3列）
    """
    connection = get_connection(**db_config_2)

    try:
        with connection.cursor() as cursor:
            # ✅ 递增参数id
            id_counter = 1

            row_count = table_widget.rowCount()
            current_row = 2  # 数据从第2行开始（前2行为表头）

            while current_row < row_count:
                # ✅ 获取分组字段：接头种类（合并项）
                joint_type_item = table_widget.item(current_row, 0)
                current_joint_type = joint_type_item.text().strip() if joint_type_item else ""

                sub_row = current_row
                while sub_row < row_count:
                    # 判断是否是新组
                    if sub_row != current_row:
                        joint_type_check = table_widget.item(sub_row, 0)
                        if joint_type_check and joint_type_check.text().strip():
                            break

                    # ✅ 提取每一行字段
                    detect_method = table_widget.item(sub_row, 1).text().strip() if table_widget.item(sub_row, 1) else ""

                    shell_tech = table_widget.item(sub_row, 2).text().strip() if table_widget.item(sub_row, 2) else ""
                    shell_ratio = table_widget.item(sub_row, 3).text().strip() if table_widget.item(sub_row, 3) else ""
                    shell_level = table_widget.item(sub_row, 4).text().strip() if table_widget.item(sub_row, 4) else ""

                    tube_tech = table_widget.item(sub_row, 5).text().strip() if table_widget.item(sub_row, 5) else ""
                    tube_ratio = table_widget.item(sub_row, 6).text().strip() if table_widget.item(sub_row, 6) else ""
                    tube_level = table_widget.item(sub_row, 7).text().strip() if table_widget.item(sub_row, 7) else ""

                    if source_status == "条件模板":
                        # ✅ INSERT 插入
                        cursor.execute(f"""
                            INSERT INTO {table_name} (
                                `无损检测数据参数ID`, `产品ID`, `接头种类`, `检测方法`,
                                `壳程_技术等级`, `壳程_检测比例`, `壳程_合格级别`,
                                `管程_技术等级`, `管程_检测比例`, `管程_合格级别`
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            id_counter,
                            product_id,
                            current_joint_type,
                            detect_method,
                            shell_tech, shell_ratio, shell_level,
                            tube_tech, tube_ratio, tube_level
                        ))
                    else:
                        # ✅ UPDATE 更新
                        cursor.execute(f"""
                            UPDATE {table_name}
                            SET `接头种类` = %s,
                                `检测方法` = %s,
                                `壳程_技术等级` = %s,
                                `壳程_检测比例` = %s,
                                `壳程_合格级别` = %s,
                                `管程_技术等级` = %s,
                                `管程_检测比例` = %s,
                                `管程_合格级别` = %s
                            WHERE `无损检测数据参数ID` = %s AND `产品ID` = %s
                        """, (
                            current_joint_type,
                            detect_method,
                            shell_tech, shell_ratio, shell_level,
                            tube_tech, tube_ratio, tube_level,
                            id_counter,
                            product_id
                        ))

                    id_counter += 1
                    sub_row += 1

                current_row = sub_row

        connection.commit()

    finally:
        connection.close()

def save_all_tables(viewer, product_id):
    """
    保存所有表格数据（标准、设计、通用、涂漆、无损检测）至数据库
    """
    try:
        if not product_id:
            QMessageBox.warning(viewer, "产品ID无效", "产品ID不能为空")
            return

        is_from_design_lib = viewer.design_data_source == "设计活动库"

        # 提取数据并保存到各自表
        save_data_to_database(
            get_table_data(viewer.tableWidget_product_std),
            product_id,
            "产品设计活动表_产品标准数据表",
            viewer.tableWidget_product_std,
            is_from_design_lib
        )

        save_data_to_database(
            get_table_data(viewer.tableWidget_design_data),
            product_id,
            "产品设计活动表_设计数据表",
            viewer.tableWidget_design_data,
            is_from_design_lib
        )

        save_data_to_database(
            get_table_data(viewer.tableWidget_general_data),
            product_id,
            "产品设计活动表_通用数据表",
            viewer.tableWidget_general_data,
            is_from_design_lib
        )

        save_coating_table_to_database(
            viewer.tableWidget_coating_data,
            "产品设计活动表_涂漆数据表",
            product_id,
            viewer.design_data_source
        )

        save_trail_table_to_database(
            viewer.tableWidget_trail_data,
            "产品设计活动表_无损检测数据表",
            product_id,
            viewer.design_data_source
        )
        viewer.design_data_source = "设计活动库"
    except Exception as e:
        QMessageBox.critical(viewer, "保存失败", f"保存数据时发生错误：{str(e)}")

"""保存前检查必填项"""
def validate_required_fields(table_widget, mode="设计数据"):
    """
    检查带星号的“参数名称”对应的必填字段是否为空
    - mode="设计数据"：要求壳程数值、管程数值必须填写
    - mode="通用数据"：要求参数值必须填写
    - 特殊强制：进、出口压力差 的管程数值为必填
    """
    required_col_name = {
        "设计数据": ["壳程数值", "管程数值"],
        "通用数据": ["数值"]
    }

    header_map = {}
    for col in range(table_widget.columnCount()):
        item = table_widget.horizontalHeaderItem(col)
        if item:
            header_map[item.text()] = col

    name_col = header_map.get("参数名称")
    if name_col is None:
        return False, []

    required_cols = [header_map.get(cn) for cn in required_col_name[mode] if cn in header_map]

    missing_rows = []

    for row in range(table_widget.rowCount()):
        name_item = table_widget.item(row, name_col)
        if not name_item:
            continue
        name_text = name_item.text().strip()

        # ✅ 常规：带 * 的参数检查
        if "*" in name_text:
            for col in required_cols:
                val_item = table_widget.item(row, col)
                if not val_item or not val_item.text().strip():
                    missing_rows.append((row, name_text))
                    break  # 当前行已有缺失字段

        # ✅ 强制补充项：进、出口压力差 的“管程数值”必须填写
        if mode == "设计数据" and name_text == "进、出口压力差":
            col = header_map.get("管程数值")
            if col is not None:
                val_item = table_widget.item(row, col)
                if not val_item or not val_item.text().strip():
                    missing_rows.append((row, name_text + "（管程）"))

    return len(missing_rows) > 0, missing_rows


"""高亮未填项"""
def highlight_missing_required_rows(table_widget: QTableWidget, missing_info: list):
    """
    高亮缺失值的行（浅蓝色），并恢复非缺失行为交替背景色。
    使用 Qt.UserRole+1 标记缺失行。
    """
    for row in range(table_widget.rowCount()):
        for col in range(table_widget.columnCount()):
            item = table_widget.item(row, col)
            if item:
                # 清除旧标记
                item.setData(Qt.UserRole + 1, None)

                # 恢复交替颜色
                if row % 2 == 0:
                    item.setBackground(QColor("#ffffff"))
                else:
                    item.setBackground(QColor("#f0f0f0"))

    # 设置缺失行背景并添加标记
    for row_idx, _ in missing_info:
        for col in range(table_widget.columnCount()):
            item = table_widget.item(row_idx, col)
            if item:
                item.setBackground(QColor("#90d7ec"))  # 浅蓝色
                item.setData(Qt.UserRole + 1, "missing")  # ✅ 标记为缺失

"""参数值类型限制，关联限制"""
def safe_set_text_and_color(widget, text, color=None):
    if hasattr(widget, "setText"):
        widget.setText(text)
        if hasattr(widget, "setToolTip"):
            widget.setToolTip(text)  # ✅ 加这一行
    if isinstance(widget, QWidget) and color:
        widget.setStyleSheet(f"color: {color};")

def validate_design_table_cell(param_name: str, column_name: str, value: str, line_edit_widget, table_widget=None, col_index=None) -> bool:
    """
    主入口函数，负责分派规则函数
    - 返回值：校验结果等级 "ok" / "warn" / "error"
    """

    param_name = param_name.strip()
    column_name = column_name.strip()
    key = (param_name, column_name)

    # ✅ 用户主动清空时，允许为空（后续由“是否必填”统一校验）
    if value.strip() == "":
        safe_set_text_and_color(line_edit_widget, "", "black")
        return "ok"

    try:
        # ✅ 自定义规则表（check_xxx）
        custom_rules = {
            ("公称直径*", "壳程数值"): check_dn,
            ("公称直径*", "管程数值"): check_dn,
            ("工作压力", "壳程数值"): check_work_pressure,
            ("工作压力", "管程数值"): check_work_pressure,
            ("工作温度（入口）", "壳程数值"): check_work_temp_in,
            ("工作温度（入口）", "管程数值"): check_work_temp_in,
            ("工作温度（出口）", "壳程数值"): check_work_temp_out,
            ("工作温度（出口）", "管程数值"): check_work_temp_out,
            ("最高允许工作压力", "壳程数值"): check_work_pressure_max,
            ("最高允许工作压力", "管程数值"): check_work_pressure_max,
            ("管板设计压差", "壳程数值"): check_tubeplate_design_pressure_gap,
            ("管板设计压差", "管程数值"): check_tubeplate_design_pressure_gap,
            ("设计压力*", "壳程数值"): check_design_pressure,
            ("设计压力*", "管程数值"): check_design_pressure,
            ("设计压力2（设计工况2）", "壳程数值"): check_design_pressure2,
            ("设计压力2（设计工况2）", "管程数值"): check_design_pressure2,
            ("设计温度（最高）*", "壳程数值"): check_design_temp_max,
            ("设计温度（最高）*", "管程数值"): check_design_temp_max,
            ("设计温度2（设计工况2）", "壳程数值"): check_design_temp_max2,
            ("设计温度2（设计工况2）", "管程数值"): check_design_temp_max2,
            ("最低设计温度", "壳程数值"): check_design_temp_min,
            ("最低设计温度", "管程数值"): check_design_temp_min,
            ("进、出口压力差", "壳程数值"): check_in_out_pressure_gap,
            ("进、出口压力差", "管程数值"): check_in_out_pressure_gap,
            ("自定义耐压试验压力（卧）", "壳程数值"): check_def_trail_stand_pressure_lying,
            ("自定义耐压试验压力（卧）", "管程数值"): check_def_trail_stand_pressure_lying,
            ("自定义耐压试验压力（立）", "壳程数值"): check_def_trail_stand_pressure_stand,
            ("自定义耐压试验压力（立）", "管程数值"): check_def_trail_stand_pressure_stand,
            ("耐压试验介质密度", "壳程数值"): check_trail_stand_pressure_medium_density,
            ("耐压试验介质密度", "管程数值"): check_trail_stand_pressure_medium_density,
            ("绝热层厚度", "壳程数值"): check_insulation_layer_thickness,
            ("绝热层厚度", "管程数值"): check_insulation_layer_thickness,
            ("绝热材料密度", "壳程数值"): check_insulation_material_density,
            ("绝热材料密度", "管程数值"): check_insulation_material_density,
            ("耐压试验类型*", "壳程数值"): check_trail_stand_pressure_type,
            ("耐压试验类型*", "管程数值"): check_trail_stand_pressure_type
        }

        # ✅ 通用规则（基础类型/范围检查）
        base_rules = {
            ("介质密度", "壳程数值"): ("float", None, None),
            ("介质密度", "管程数值"): ("float", None, None),
            ("介质入口流速", "壳程数值"): ("float", None, None),
            ("介质入口流速", "管程数值"): ("float", None, None),
            ("液柱静压力", "壳程数值"): ("float", (0, 1e10), "液柱静压力的参数值不能为负，请核对后输入"),
            ("液柱静压力", "管程数值"): ("float", None, "液柱静压力的参数值不能为负，请核对后输入"),
            ("腐蚀裕量*", "壳程数值"): ("float", (0, 1e10), "腐蚀裕量的参数值不能为负，请核对后输入"),
            ("腐蚀裕量*", "管程数值"): ("float", None, "腐蚀裕量的参数值不能为负，请核对后输入")
        }

        print(f"[校验函数] param={param_name}, col={column_name}, value='{value}'")

        if key in custom_rules:
            result, msg = custom_rules[key](value, line_edit_widget, param_name, column_name, table_widget, col_index)
            if result == "ok":
                safe_set_text_and_color(line_edit_widget, "", "black")
            elif result == "warn":
                safe_set_text_and_color(line_edit_widget, msg, "orange")
            elif result == "error":
                safe_set_text_and_color(line_edit_widget, msg, "red")
            return result

        if key in base_rules:
            try:
                dtype, limits, msg = base_rules[key]
                if dtype == "int":
                    num = int(value)
                elif dtype == "float":
                    num = float(value)
                else:
                    safe_set_text_and_color(line_edit_widget, "输入数据类型有误，请确认后输入", "red")
                    return "error"
                if limits:
                    min_v, max_v = limits
                    if not (min_v <= num <= max_v):
                        safe_set_text_and_color(line_edit_widget, msg, "red")
                        return "error"
                safe_set_text_and_color(line_edit_widget, "", "black")
                return "ok"
            except Exception:
                safe_set_text_and_color(line_edit_widget, "校验异常，请确认输入", "red")
                return "error"

        return "ok"

    except Exception:
        safe_set_text_and_color(line_edit_widget, "校验异常，请确认输入", "red")
        return "error"

def validate_general_table_cell(param_name: str, value: str, line_edit_widget, table_widget=None) -> str:
    """
    通用数据表 校验入口函数
    - param_name: 参数名称
    - value: 用户输入的参数值（字符串）
    - line_edit_widget: QLineEdit 显示提示
    - 返回值: 校验等级 "ok" / "warn" / "error"
    """

    param_name = param_name.strip()

    # ✅ 主动清空，允许通过
    if value.strip() == "":
        safe_set_text_and_color(line_edit_widget, "", "black")  # ✅ 正确
        return "ok"

    try:
        # ✅ 自定义规则（check_xxx 通常联动或复杂校验）
        custom_rules = {
            # ("参数名称",): check_xxx,
        }

        # ✅ 通用规则（类型 + 范围判断）
        base_rules = {
            ("设计使用年限*",): ("int", (0, 1e10), "设计使用年限不能为负，请核对后输入"),
            ("基本风压",): ("float", (0, 1e10), "基本风压值不能为负，请核对后输入"),
            ("雪压值",): ("float", (0, None), "雪压值不能为负，请核对后输入"),
            # ... 继续补充更多通用项
        }

        key = (param_name,)

        # ✅ 优先匹配自定义规则
        if key in custom_rules:
            result, msg = custom_rules[key](value, line_edit_widget, param_name, table_widget)
            if result == "ok":
                safe_set_text_and_color(line_edit_widget, "", "black")
            elif result == "warn":
                safe_set_text_and_color(line_edit_widget, msg, "orange")
            elif result == "error":
                safe_set_text_and_color(line_edit_widget, msg, "red")
            return result  # "ok" / "warn" / "error"

        # ✅ 通用处理
        if key in base_rules:
            dtype, limits, msg = base_rules[key]

            # 🧠 第一层：手动类型转换错误提示
            try:
                if dtype == "int":
                    num = int(value)
                elif dtype == "float":
                    num = float(value)
                else:
                    safe_set_text_and_color(line_edit_widget, "输入数据类型有误，请确认后输入", "red")
                    return "error"
            except ValueError:
                safe_set_text_and_color(line_edit_widget, "输入数据类型有误，请确认后输入", "red")
                return "error"

            # 🧠 第二层：其他逻辑错误
            try:
                if limits:
                    min_v, max_v = limits
                    if (min_v is not None and num < min_v) or (max_v is not None and num > max_v):
                        safe_set_text_and_color(line_edit_widget, msg, "red")
                        return "error"

                safe_set_text_and_color(line_edit_widget, "", "black")
                return "ok"

            except Exception:
                safe_set_text_and_color(line_edit_widget, "校验异常，请确认输入", "red")
                return "error"

        return "ok"  # 无匹配项默认通过

    except Exception as e:
        safe_set_text_and_color(line_edit_widget, "校验异常，请确认输入", "red")
        return "error"

def validate_trail_table_cell(column_name: str, value: str, tip_widget, table_widget=None) -> str:
    """
    检测数据表 - 通用列校验器（仅对“检测比例”做范围检查）
    支持格式：50、≥30、>20，范围限制为 [0, 100]
    """
    if value.strip() == "":
        safe_set_text_and_color(tip_widget, "", "black")
        return "ok"

    if not re.search(r"检测比例[%]?$", column_name):
        return "ok"  # 仅校验“检测比例”列

    val = value.strip()

    # ✅ 合法匹配正则： 纯数字 或 ≥数字 或 >数字（不带 %）
    pattern = r"^(≥|>)?\d{1,3}$"
    if not re.match(pattern, val):
        safe_set_text_and_color(tip_widget, "请输入合法格式，如 50，≥30 或 >20", "red")
        return "error"

    # ✅ 数值范围判断（提取数字部分）
    try:
        num_part = int(re.sub(r"[^\d]", "", val))
        if not (0 <= num_part <= 100):
            safe_set_text_and_color(tip_widget, "检测比例应在 0 ~ 100 之间，请核对后输入", "red")
            return "error"
    except Exception:
        safe_set_text_and_color(tip_widget, "检测比例格式异常", "red")
        return "error"

    safe_set_text_and_color(tip_widget, "", "black")
    return "ok"

def validate_coating_table_cell(column_name: str, value: str, tip_widget, table_widget=None) -> str:
    """
    涂漆数据表 校验器
    - 针对：干膜厚度（μm）、涂漆面积 两列进行校验
    """
    if value.strip() == "":
        safe_set_text_and_color(tip_widget, "", "black")
        return "ok"

    val = value.strip()

    # ✅ 如果列名像“列5”，说明未传入真实逻辑列头 → 尝试自己查
    if column_name.startswith("列") and table_widget and hasattr(table_widget, "logical_headers"):
        try:
            col_index = int(column_name.replace("列", ""))
            column_name = table_widget.logical_headers[col_index]
        except Exception:
            # 万一列号非法，直接跳过
            return "ok"

    col = column_name.strip()

    if col not in ["干膜厚度（μm）", "涂漆面积"]:
        return "ok"  # 其他列无需校验

    try:
        num = float(val)
    except ValueError:
        safe_set_text_and_color(tip_widget, "输入数据类型有误，请确认后输入", "red")
        return "error"

    if num <= 0:
        safe_set_text_and_color(tip_widget, f"{col}必须为正数，请核对后输入", "red")
        return "error"

    safe_set_text_and_color(tip_widget, "", "black")
    return "ok"

def dispatch_cell_validation(viewer, table, row, col, param_name, column_name, value, *args, **kwargs):
    print(f"[调试] dispatch_cell_validation: col={column_name}, value={value}")

    mode = getattr(table, "validation_mode", "design")

    if value.strip() == "":
        safe_set_text_and_color(viewer.line_tip, "", "black")
        return "ok"

    if mode == "design":
        return validate_design_table_cell(param_name, column_name, value, viewer.line_tip, table, col)

    elif mode == "general":
        if column_name != "数值":
            safe_set_text_and_color(viewer.line_tip, "", "black")
            return "ok"
        return validate_general_table_cell(param_name, value, viewer.line_tip, table)

    elif mode == "trail":
        result = validate_trail_table_cell(column_name, value, viewer.line_tip, table)
        if result == "error":
            return result

        item = table.item(row, col)
        if item:
            default_val = item.data(Qt.UserRole + 2)
            if default_val:
                if column_name.endswith("技术等级") and is_grade_lower(value, default_val):
                    msg = "技术等级不能低于默认值，请核对后输入"
                    safe_set_text_and_color(viewer.line_tip, msg, "red")
                    if hasattr(viewer, "import_tip_list"):
                        viewer.import_tip_list.append(f"[检测数据] 第{row - 1}行 - {column_name}: ❌ {msg}")

                    QTimer.singleShot(0, lambda: table.item(row, col).setText(""))
                    return "error"
                elif column_name.endswith("合格级别") and is_qualify_lower(value, default_val):
                    msg = "合格级别不能低于默认值，请核对后输入"
                    safe_set_text_and_color(viewer.line_tip, msg, "red")
                    if hasattr(viewer, "import_tip_list"):
                        viewer.import_tip_list.append(f"[检测数据] 第{row - 1}行 - {column_name}: ❌ {msg}")
                    QTimer.singleShot(0, lambda: table.item(row, col).setText(""))
                    return "error"

        safe_set_text_and_color(viewer.line_tip, "", "black")
        return result

    elif mode == "coating":
        return validate_coating_table_cell(column_name, value, viewer.line_tip, table)

    return "ok"


"""参考数据导入相关函数"""

def get_ref_data_excel_path(product_id: int) -> str:
    """
    给定产品ID，查询并返回对应的 条件输入数据表.xlsx 完整路径
    """
    try:
        # 第一步：连接产品需求库，查产品需求表
        connection = get_connection(**db_config_3)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT `项目ID`, `产品编号`, `产品名称`, `设备位号`
                FROM `产品需求表`
                WHERE `产品ID` = %s
                LIMIT 1
            """, (product_id,))
            product_row = cursor.fetchone()
        connection.close()

        if not product_row:
            raise ValueError(f"未找到产品ID {product_id} 的产品需求信息。")

        project_id = product_row['项目ID']
        product_code = product_row['产品编号']
        product_name = product_row['产品名称']
        device_loc_id = product_row['设备位号']

        # 第二步：连接项目需求库，查项目需求表
        connection = get_connection(**db_config_4)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT `项目保存路径`,`项目名称`,`业主名称`
                FROM `项目需求表`
                WHERE `项目ID` = %s
                LIMIT 1
            """, (project_id,))
            project_row = cursor.fetchone()
        connection.close()

        if not project_row:
            raise ValueError(f"未找到项目ID {project_id} 的项目信息。")

        project_save_path = project_row['项目保存路径']
        project_path = project_row['项目名称']
        yezhu_path = project_row['业主名称']
        pinjie_path = f"{yezhu_path}_{project_path}"
        # 第三步：拼接路径
        folder_name = f"{product_code}_{product_name}_{device_loc_id}"
        full_path = os.path.join(project_save_path, pinjie_path, folder_name, "条件输入数据表.xlsx")
        # ✅ 检查文件是否存在
        if not os.path.exists(full_path):
            raise FileNotFoundError(f"未找到文件：{full_path}")

        return full_path

    except Exception as e:
        # 可以根据需要在这里统一处理异常（比如打印日志，或者继续往上抛）
        raise e

def get_user_selected_excel_path(parent_widget=None) -> str:
    """
    弹出文件选择框，获取用户选择的Excel路径
    """
    file_path, _ = QFileDialog.getOpenFileName(
        parent_widget,
        "选择条件输入数据表",
        "",
        "Excel Files (*.xlsx);;All Files (*)"
    )
    if not file_path:
        raise FileNotFoundError("用户未选择文件")
    return file_path

def update_product_standard_table_from_excel(excel_path: str, table_widget):
    """
    从Excel中读取‘产品标准’Sheet，按规范/标准名称匹配，更新界面表格中的‘规范/标准代号’列
    """
    try:
        df = pd.read_excel(excel_path, sheet_name="产品标准", dtype=str)
        df.fillna("", inplace=True)

        # ✅ 构建映射表：规范/标准名称 -> 规范/标准代号
        std_map = {str(k).strip(): str(v).strip() for k, v in zip(df.iloc[:, 1], df.iloc[:, 2])}
        # 注意这里用的是第1列（B列，“规范/标准名称”），不是序号列了！

        for row in range(table_widget.rowCount()):
            name_item = table_widget.item(row, 1)  # 第1列是规范/标准名称
            target_item = table_widget.item(row, 2)  # 第2列是规范/标准代号

            if not name_item or not target_item:
                continue

            name = str(name_item.text()).strip()
            if name in std_map:
                target_item.setText(std_map[name])

    except Exception as e:
        raise RuntimeError(f"导入产品标准失败：{str(e)}")

def update_design_data_table_from_excel(excel_path: str, table_widget):
    """
    从Excel中读取‘设计数据’Sheet，按参数名称匹配，更新‘壳程数值’和‘管程数值’
    如果本地界面中“绝热层类型”是“无”，则跳过对应侧的绝热材料、厚度、密度的导入
    """
    try:
        import pandas as pd
        df = pd.read_excel(excel_path, sheet_name="设计数据", dtype=str)
        df.fillna("", inplace=True)

        # Excel 中构建映射表
        data_map = {
            str(row[1]).strip(): (str(row[3]).strip(), str(row[4]).strip())
            for _, row in df.iterrows()
        }

        # ✅ 获取界面当前的“绝热层类型”值
        insulation_type_shell = ""
        insulation_type_tube = ""
        for row in range(table_widget.rowCount()):
            name_item = table_widget.item(row, 1)
            if name_item and name_item.text().strip() == "绝热层类型":
                shell_item = table_widget.item(row, 3)
                tube_item = table_widget.item(row, 4)
                insulation_type_shell = shell_item.text().strip() if shell_item else ""
                insulation_type_tube = tube_item.text().strip() if tube_item else ""
                break

        skip_shell = insulation_type_shell == "无"
        skip_tube = insulation_type_tube == "无"

        print(f"[导入判定] 绝热层类型: 壳程={insulation_type_shell}, 管程={insulation_type_tube} | skip_shell={skip_shell}, skip_tube={skip_tube}")

        for row in range(table_widget.rowCount()):
            name_item = table_widget.item(row, 1)
            if not name_item:
                continue

            name = name_item.text().strip()
            if name not in data_map:
                continue

            shell_val, tube_val = data_map[name]

            # 判断是否为绝热项且需要跳过
            if name in {"绝热材料", "绝热层厚度", "绝热材料密度"}:
                if skip_shell:
                    shell_val = ""  # 不导入壳程
                if skip_tube:
                    tube_val = ""  # 不导入管程

            # 更新壳程
            shell_item = table_widget.item(row, 3)
            if shell_item:
                shell_item.setText(shell_val)

            # 更新管程
            tube_item = table_widget.item(row, 4)
            if tube_item:
                tube_item.setText(tube_val)

    except Exception as e:
        raise RuntimeError(f"导入设计数据失败：{str(e)}")

def update_general_data_table_from_excel(excel_path: str, table_widget):
    """
    从Excel中读取‘通用数据’Sheet，按参数名称匹配，更新‘参数值’。
    多选项字段将自动识别并标准化为“；”分隔格式。
    """
    try:
        df = pd.read_excel(excel_path, sheet_name="通用数据", dtype=str)
        df.fillna("", inplace=True)

        # 构建映射表：参数名称 -> 参数值
        data_map = {
            str(row[1]).strip(): str(row[3]).strip()
            for _, row in df.iterrows()
        }

        for row in range(table_widget.rowCount()):
            name_item = table_widget.item(row, 1)  # 第1列是参数名称
            value_item = table_widget.item(row, 3)  # 第3列是参数值

            if not name_item or not value_item:
                continue

            name = name_item.text().strip()
            if name not in data_map:
                continue

            raw_val = data_map[name]
            config = GENERAL_PARAM_CONFIG.get(name)

            # 不做修改，保留原始值，等待后续 validate_all_tables_after_import() 中统一处理
            value_item.setText(raw_val)


    except Exception as e:
        raise RuntimeError(f"导入通用数据失败：{str(e)}")

def update_trail_data_table_from_excel(excel_path: str, table_widget):
    """
    从Excel中读取‘检测数据’Sheet，只更新壳程/管程字段，
    行对齐从界面row=2开始，Excel从第3行开始（跳过两级表头）
    """
    try:
        df = pd.read_excel(excel_path, sheet_name="检测数据", header=None, skiprows=2, dtype=str)
        df.fillna("", inplace=True)

        field_to_col = {
            "壳程_技术等级": 2,
            "壳程_检测比例": 3,
            "壳程_合格级别": 4,
            "管程_技术等级": 5,
            "管程_检测比例": 6,
            "管程_合格级别": 7
        }

        current_row = 2  # ✅ 第2行是界面第一个数据行
        for _, row in df.iterrows():
            if current_row >= table_widget.rowCount():
                break

            values = {
                "壳程_技术等级": str(row[2]).strip(),
                "壳程_检测比例": str(row[3]).strip(),
                "壳程_合格级别": str(row[4]).strip(),
                "管程_技术等级": str(row[5]).strip(),
                "管程_检测比例": str(row[6]).strip(),
                "管程_合格级别": str(row[7]).strip()
            }

            # ✅ 获取当前行检测方法
            method_item = table_widget.item(current_row, 1)
            method = method_item.text().strip() if method_item else ""

            for field, col in field_to_col.items():
                val = values.get(field, "")
                item = table_widget.item(current_row, col)
                if not item:
                    item = QTableWidgetItem()
                    table_widget.setItem(current_row, col, item)

                item.setText(val)
                # 手动触发校验
                from modules.condition_input.funcs.funcs_cdt_input import dispatch_cell_validation
                viewer = getattr(table_widget, "viewer", None)
                if viewer:
                    header_item = table_widget.horizontalHeaderItem(col)
                    column_name = header_item.text().strip() if header_item else ""
                    dispatch_cell_validation(viewer, table_widget, current_row, col, "", column_name, val)

            if method:
                for side in ["壳程", "管程"]:
                    tech_col = field_to_col.get(f"{side}_技术等级")
                    qualify_col = field_to_col.get(f"{side}_合格级别")

                    tech_val = table_widget.item(current_row, tech_col).text().strip() if table_widget.item(current_row,
                                                                                                            tech_col) else ""
                    qualify_val = table_widget.item(current_row, qualify_col).text().strip() if table_widget.item(
                        current_row, qualify_col) else ""

                    if not tech_val and not qualify_val:
                        from .funcs_cdt_input import autofill_trail_test_grade
                        autofill_trail_test_grade(table_widget, current_row, side,
                                                  getattr(table_widget, "undo_stack", None))

            current_row += 1

    except Exception as e:
        raise RuntimeError(f"导入检测数据失败：{str(e)}")

def update_coating_data_table_from_excel(excel_path: str, coating_table_widget, product_std_table_widget):
    """
    从Excel中读取‘涂漆数据’Sheet，更新执行标准和每组涂层数据，
    执行标准统一从产品标准表中的“涂漆标准”获取。
    """
    try:
        df = pd.read_excel(excel_path, sheet_name="涂漆数据", dtype=str, header=None)
        df.fillna("", inplace=True)

        # ✅ 从产品标准表中获取“涂漆标准”的规范代号
        coating_std_value = ""
        for row in range(product_std_table_widget.rowCount()):
            name_item = product_std_table_widget.item(row, 1)
            value_item = product_std_table_widget.item(row, 2)
            if name_item and name_item.text().strip() == "涂漆标准" and value_item:
                coating_std_value = value_item.text().strip()
                break

        # ✅ 设置到涂漆数据表第0行第2列（执行标准/规范）
        std_item = coating_table_widget.item(0, 2)
        if std_item:
            std_item.setText(coating_std_value)

        # ✅ 涂层数据从第3行开始（即df的第2行索引）
        excel_rows = []
        current_usage = ""

        for idx in range(2, len(df)):
            row = df.iloc[idx]
            usage = str(row[0]).strip()
            if usage:
                current_usage = usage

            excel_rows.append({
                "用途": current_usage,
                "细类": str(row[1]).strip(),
                "油漆类别": str(row[2]).strip(),
                "颜色": str(row[3]).strip(),
                "干膜厚度（μm）": str(row[4]).strip(),
                "涂漆面积": str(row[5]).strip(),
                "备注": str(row[6]).strip()
            })

        # ✅ 写入界面表格
        current_row = 2
        last_usage = None

        while current_row < coating_table_widget.rowCount() and excel_rows:
            excel_row = excel_rows.pop(0)
            usage = excel_row["用途"]

            for col_idx, field in enumerate([
                "用途", "细类", "油漆类别", "颜色", "干膜厚度（μm）", "涂漆面积", "备注"
            ]):
                if col_idx in (0, 1):
                    continue  # 用途、细类列不更新

                item = coating_table_widget.item(current_row, col_idx)
                if not item:
                    continue

                val = excel_row.get(field, "")
                if field in ("涂漆面积", "备注"):
                    if usage != last_usage:
                        item.setText(val)
                else:
                    item.setText(val)

            last_usage = usage
            current_row += 1

    except Exception as e:
        raise RuntimeError(f"导入涂漆数据失败：{str(e)}")

def import_all_reference_data(excel_path: str, viewer: QWidget):
    """
    给定Excel路径和界面viewer对象，一次性导入所有参考数据并更新到界面
    """
    viewer.import_tip_list = []  # ✅ 存储 dispatch 校验中捕获的错误提示

    update_product_standard_table_from_excel(excel_path, viewer.tableWidget_product_std)
    update_design_data_table_from_excel(excel_path, viewer.tableWidget_design_data)
    update_general_data_table_from_excel(excel_path, viewer.tableWidget_general_data)
    update_trail_data_table_from_excel(excel_path, viewer.tableWidget_trail_data)
    update_coating_data_table_from_excel(
        excel_path,
        viewer.tableWidget_coating_data,
        viewer.tableWidget_product_std
    )

    trigger_all_cross_table_relations(viewer)
    validate_all_tables_after_import(viewer)

"""导入参考数据对应的检查"""
def validate_all_tables_after_import(viewer: QWidget):
    tip_list = []

    # ✅ 设计数据表（新增：校验下拉值）
    product_id = getattr(viewer, "product_id", "")
    design_dropdown_config = apply_design_data_dropdowns(viewer=viewer, product_id=product_id)

    table = viewer.tableWidget_design_data
    for row in range(table.rowCount()):
        param_item = table.item(row, 1)
        if not param_item or not param_item.text():
            continue
        param_name = param_item.text().strip()

        for col_index, col_name in [(3, "壳程数值"), (4, "管程数值")]:
            cell_item = table.item(row, col_index)
            if not cell_item or not cell_item.text():
                continue
            val = cell_item.text().strip()

            conf = design_dropdown_config.get(param_name)
            if conf and not conf.get("editable", False):
                allowed = conf.get("options", [])
                if val not in allowed:
                    cell_item.setText("")
                    tip_list.append(f"[设计数据] {param_name} - {col_name}: ❌ 非法下拉值“{val}”，已清空")
                    continue

            result = validate_design_table_cell(param_name, col_name, val, QTableWidgetItem(), table, col_index)
            if result == "error":
                cell_item.setText("")
                tip_list.append(f"[设计数据] {param_name} - {col_name}: ❌ 非法值，已清空")
            elif result == "warn":
                tip_list.append(f"[设计数据] {param_name} - {col_name}: ⚠️ 可疑值")

    # ✅ 通用数据表
    table = viewer.tableWidget_general_data
    for row in range(table.rowCount()):
        param_item = table.item(row, 1)
        value_item = table.item(row, 3)
        if not param_item or not value_item or not param_item.text() or not value_item.text():
            continue
        param_name = param_item.text().strip()
        val = value_item.text().strip()

        conf = GENERAL_PARAM_CONFIG.get(param_name)
        if conf and not conf.get("editable", False):  # ✅ 仅校验不可编辑字段
            corrected_val, msg = validate_dropdown_value(param_name, val, GENERAL_PARAM_CONFIG)
            value_item.setText(corrected_val)
            if msg:
                tip_list.append(f"[通用数据] {param_name}: {msg}")
            continue

        # ✅ 再做常规校验
        result = validate_general_table_cell(param_name, val, QTableWidgetItem(), table)
        if result == "error":
            value_item.setText("")
            tip_list.append(f"[通用数据] {param_name}: ❌ 非法值，已清空")
        elif result == "warn":
            tip_list.append(f"[通用数据] {param_name}: ⚠️ 可疑值")

    # ✅ 检测数据表：检测比例列已有校验，这里扩展对委托配置列校验（技术等级/合格级别）
    trail_config = apply_trail_data_dropdowns()
    table = viewer.tableWidget_trail_data
    for row in range(2, table.rowCount()):
        method_item = table.item(row, 1)
        method = method_item.text().strip() if method_item else ""
        conf = trail_config.get(method)

        for col_index in [2, 4, 5, 7]:
            item = table.item(row, col_index)
            if not item or not item.text() or not conf:
                continue
            val = item.text().strip()
            valid_options = []
            for cols, opts in conf.items():
                if col_index in cols:
                    valid_options = opts
                    break
            if valid_options and val not in valid_options:
                item.setText("")
                tip_list.append(f"[检测数据] 第{row + 1}行 - 列{col_index + 1}: ❌ 非法下拉值“{val}”，已清空")

        # 检测比例列校验（保持原有逻辑）
        for col_index in [3, 6]:
            item = table.item(row, col_index)
            if not item or not item.text():
                continue
            val = item.text().strip()
            header_item = table.horizontalHeaderItem(col_index)
            header = header_item.text() if header_item else f"列{col_index}"
            result = validate_trail_table_cell(header, val, QTableWidgetItem(), table)
            if result == "error":
                item.setText("")
                tip_list.append(f"[检测数据] 第{row + 1}行 - {header}: ❌ 非法值，已清空")
            elif result == "warn":
                tip_list.append(f"[检测数据] 第{row + 1}行 - {header}: ⚠️ 可疑值")

    # ✅ 涂漆数据表
    table = viewer.tableWidget_coating_data
    for row in range(2, table.rowCount()):
        for col_index in [4, 5]:
            item = table.item(row, col_index)
            if not item or not item.text():
                continue
            val = item.text().strip()
            # ✅ 优先从 logical_headers 获取列名
            if hasattr(table, "logical_headers") and col_index < len(table.logical_headers):
                header = table.logical_headers[col_index]
            else:
                header_item = table.horizontalHeaderItem(col_index)
                header = header_item.text().strip() if header_item and header_item.text() else f"列{col_index}"
            result = validate_coating_table_cell(header, val, QTableWidgetItem(), table)
            print(f"Validating column: {header}, value: {val}, result: {result}")
            if result == "error":
                item.setText("")
                tip_list.append(f"[涂漆数据] 第{row+1}行 - {header}: ❌ 非法值，已清空")
            elif result == "warn":
                tip_list.append(f"[涂漆数据] 第{row+1}行 - {header}: ⚠️ 可疑值")

    # ✅ 合并导入校验过程中记录的提示
    if hasattr(viewer, "import_tip_list"):
        tip_list.extend(viewer.import_tip_list)

    # ✅ 显示提示：主显示 + tooltip 显示完整内容
    tip_message = "\n".join(tip_list) if tip_list else "✅ 所有导入数据校验通过。"
    viewer.line_tip.setText(tip_message[:80].replace("\n", " | "))
    viewer.line_tip.setToolTip(tip_message)
    viewer.line_tip.setStyleSheet("color: black;")  # ✅ 强制黑色字体

def trigger_all_cross_table_relations(viewer: QWidget):
    """
    仅触发“绝热层类型”联动，避免影响焊接接头等其他联动逻辑。
    用于导入参考数据时确保绝热项锁定状态正确。
    """
    table = viewer.tableWidget_design_data
    for row in range(table.rowCount()):
        param_item = table.item(row, 1)
        if not param_item:
            continue
        param_name = param_item.text().strip()

        if "绝热层类型" == param_name:
            for col in [3, 4]:  # 壳程和管程列
                item = table.item(row, col)
                if item and item.text().strip():
                    handle_cross_table_triggers(viewer, table, row, col)

def validate_dropdown_value(param_name: str, value: str, config: dict) -> (str, str):
    """
    检查并返回合法的下拉框值，非法则返回 ("", msg)。
    - param_name: 参数名称
    - value: 原始值
    - config: 对应的下拉配置（如 GENERAL_PARAM_CONFIG）
    """
    val = value.strip()
    conf = config.get(param_name)
    if not conf:
        return val, ""

    allowed = conf.get("options", [])
    typ = conf.get("type", "single")

    if typ == "single":
        if val not in allowed:
            return "", f"❌ 非法下拉值“{val}”，已清空"

    elif typ == "multi":
        clean_text = re.sub(r"[;；,，\s]+", "", val)

        matched = [opt for opt in allowed if opt in clean_text]

        if not matched:
            return "", f"❌ 非法选项“{value}”，已清空"

        corrected = "；".join(matched)
        return corrected, ""

    return val, ""

"""保存至本地条件输入数据表"""
def is_file_locked(filepath: str) -> bool:
    """
    判断文件是否被占用（即是否可写）
    """
    import tempfile
    import os

    if not os.path.exists(filepath):
        return False

    try:
        # 尝试以追加方式打开，如果失败说明文件被占用
        with open(filepath, 'a'):
            return False
    except IOError:
        return True

def save_local_condition_file(product_id: int, viewer: QWidget) -> bool:
    """
    保存界面数据到本地 Excel，如果文件被占用则提示并返回 False。
    """
    local_path = get_ref_data_excel_path(product_id)

    if is_file_locked(local_path):
        QMessageBox.warning(viewer, "文件占用", f"请先关闭本地文件：\n{local_path}\n然后重试保存。")
        return False  # 阻止继续

    try:
        wb = load_workbook(local_path)
    except FileNotFoundError:
        print(f"未找到本地条件数据文件：{local_path}")
        return False

    update_sheet_from_table(wb["产品标准"], viewer.tableWidget_product_std, col_start=1, col_end=3, excel_col_offset=2, excel_row_offset=2)
    update_sheet_from_table(wb["设计数据"], viewer.tableWidget_design_data, col_start=1, col_end=5, excel_col_offset=2, excel_row_offset=2)
    update_sheet_from_table(wb["通用数据"], viewer.tableWidget_general_data, col_start=1, col_end=4, excel_col_offset=2, excel_row_offset=2)
    update_sheet_from_table(wb["检测数据"], viewer.tableWidget_trail_data, col_start=2, col_end=8, excel_col_offset=3, excel_row_offset=1)
    update_sheet_from_table(wb["涂漆数据"], viewer.tableWidget_coating_data, col_start=2, col_end=7, excel_col_offset=3, excel_row_offset=1)

    wb.save(local_path)
    print(f"✅ 本地条件数据表已成功保存到: {local_path}")
    return True

def update_sheet_from_table(sheet, table_widget, col_start=0, col_end=None, excel_col_offset=1, excel_row_offset=2):
    """
    将 table_widget 的指定列范围写入到 sheet 中，跳过 MergedCell，支持 Excel 起始列和起始行偏移
    - col_start / col_end：界面表格读取列范围
    - excel_col_offset：写入到Excel起始列（比如B列就是2）
    - excel_row_offset：写入到Excel起始行（比如第2行/第3行）
    """
    rows = table_widget.rowCount()
    total_cols = table_widget.columnCount()
    col_end = col_end if col_end is not None else total_cols

    for row in range(rows):
        for col in range(col_start, col_end):
            item = table_widget.item(row, col)
            value = item.text() if item else ""

            excel_row = row + excel_row_offset
            excel_col = excel_col_offset + (col - col_start)

            cell = sheet.cell(row=excel_row, column=excel_col)

            if isinstance(cell, MergedCell):
                continue  # ⚡ 是合并单元格的从属格，不能写
            cell.value = value
"""跨表联动逻辑函数"""
def show_info_tip(viewer: QWidget, message: str):
    viewer.line_tip.setText(message)
    viewer.line_tip.setToolTip(message)

def handle_cross_table_triggers(viewer: QWidget, changed_table: QTableWidget, row: int, col: int):
    undo_stack = getattr(viewer, "undo_stack", None)

    # ✅ 涂漆标准 → 执行标准/规范联动
    if changed_table == viewer.tableWidget_product_std:

        name_item = changed_table.item(row, 1)
        value_item = changed_table.item(row, 2)
        if name_item and value_item and name_item.text().strip() == "涂漆标准":
            std_value = value_item.text().strip()
            target_table = viewer.tableWidget_coating_data
            std_cell = target_table.item(0, 2)

            if std_cell is None:
                std_cell = QTableWidgetItem()
                std_cell.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                std_cell.setTextAlignment(Qt.AlignCenter)
                target_table.setItem(0, 2, std_cell)

            old_val = std_cell.text()
            if std_value != old_val and undo_stack:
                cmd = CellEditCommand(target_table, 0, 2, old_val, std_value)
                undo_stack.push(cmd)
                cmd.redo()

            show_info_tip(viewer, "[涂漆数据]执行标准/规范已自动刷新。")

    # ✅ 焊接接头系数* → 检测数据（仅壳程或管程）
    elif changed_table == viewer.tableWidget_design_data:
        name_item = changed_table.item(row, 1)
        if not name_item:
            return

        param_name = name_item.text().strip()

        # ✅ 焊接接头系数联动检测数据
        if "焊接接头系数*" in param_name:
            if col == 3:
                shell_val = changed_table.item(row, 3).text().strip()
                update_trail_table_side_only(viewer.tableWidget_trail_data, "壳程", shell_val, undo_stack)
                show_info_tip(viewer, "[检测数据]壳程检测比例及合格级别已自动刷新。")
            elif col == 4:
                tube_val = changed_table.item(row, 4).text().strip()
                update_trail_table_side_only(viewer.tableWidget_trail_data, "管程", tube_val, undo_stack)
                show_info_tip(viewer, "[检测数据]管程检测比例及合格级别已自动刷新。")

        # ✅ 绝热层类型联动
        elif param_name == "绝热层类型":
            side = "壳程" if col == 3 else "管程" if col == 4 else None
            if not side:
                return

            cell = changed_table.item(row, col)
            val_text = cell.text().strip() if cell else ""
            prev_val = getattr(cell, "_prev_val", "") if cell else ""
            cell._prev_val = val_text  # 记录当前为下次使用

            is_none_now = val_text == "无"
            is_none_prev = prev_val == "无"

            # ✅ 仅当从“无”↔其他值之间变化时联动
            if is_none_now == is_none_prev:
                print("跳过绝热层类型联动（状态未变化）")
                return

            make_fields_editable = not is_none_now
            param_names = {"绝热材料", "绝热层厚度", "绝热材料密度"}

            for r in range(changed_table.rowCount()):
                sub_item = changed_table.item(r, 1)
                if not sub_item or sub_item.text().strip() not in param_names:
                    continue

                target_col = 3 if side == "壳程" else 4
                target_item = changed_table.item(r, target_col)
                if target_item is None:
                    target_item = QTableWidgetItem()
                    changed_table.setItem(r, target_col, target_item)

                if not make_fields_editable:
                    target_item.setText("")
                    target_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                else:
                    target_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)

            show_info_tip(viewer, f"[设计数据]{side}绝热项状态已更新")

    # ✅ 检测比例 → 联动补齐 技术等级 和 合格级别（仅当为空）
    # ✅ 新增：清空其中任一字段 → 自动清空其余两个字段
    elif changed_table == viewer.tableWidget_trail_data:
        header_item = changed_table.horizontalHeaderItem(col)
        col_name = header_item.text().strip() if header_item else ""
        side = None

        if "壳程" in col_name:
            side = "壳程"
        elif "管程" in col_name:
            side = "管程"

        # 自动补齐技术等级与合格级别
        if col_name in [f"{side}_检测比例"] and side:
            did_fill = autofill_trail_test_grade(changed_table, row, side, undo_stack)
            if did_fill:
                show_info_tip(viewer, f"[检测数据]{side}检测比例已自动联动更新技术等级与合格级别")

        # 清空联动逻辑
        if side and col_name in [f"{side}_技术等级", f"{side}_检测比例", f"{side}_合格级别"]:
            item = changed_table.item(row, col)
            if item and item.text().strip() == "":
                related_cols = {
                    f"{side}_技术等级": [f"{side}_检测比例", f"{side}_合格级别"],
                    f"{side}_检测比例": [f"{side}_技术等级", f"{side}_合格级别"],
                    f"{side}_合格级别": [f"{side}_技术等级", f"{side}_检测比例"]
                }
                for other_col_name in related_cols.get(col_name, []):
                    col_idx = next((i for i in range(changed_table.columnCount())
                                    if changed_table.horizontalHeaderItem(i).text().strip() == other_col_name), None)
                    if col_idx is not None:
                        target_item = changed_table.item(row, col_idx)
                        if target_item and target_item.text().strip():
                            old_val = target_item.text()
                            target_item.setText("")
                            if undo_stack:
                                from .undo_command import CellEditCommand
                                undo_stack.push(CellEditCommand(changed_table, row, col_idx, old_val, ""))

def update_trail_table_side_only(table: QTableWidget, side: str, factor_val: str, undo_stack=None):
    """
    根据焊接接头系数，联动更新检测数据表指定侧（壳程或管程）的：
    - 技术等级
    - 检测比例
    - 合格级别
    ✅ 同时设置默认值（UserRole+2）用于后续校验。
    """
    factor_map = {
        "1":    ("AB", "100", "Ⅱ"),
        "1.0":  ("AB", "100", "Ⅱ"),
        "0.9":  ("AB", "100", "Ⅱ"),
        "0.85": ("AB", "≥20", "Ⅲ"),
        "0.8":  ("AB", "≥20", "Ⅲ")
    }

    if factor_val not in factor_map:
        print(f"❎ 跳过无效系数: {factor_val}")
        return

    row = 2  # 固定行（第一行数据）
    col_map = {
        "壳程": {"等级": 2, "比例": 3, "合格": 4},
        "管程": {"等级": 5, "比例": 6, "合格": 7}
    }

    if side not in col_map:
        return

    grade_val, ratio_val, qualify_val = factor_map[factor_val]
    values_to_set = {
        "等级": grade_val,
        "比例": ratio_val,
        "合格": qualify_val
    }

    for field, new_val in values_to_set.items():
        col = col_map[side][field]
        item = table.item(row, col)
        if not item:
            item = QTableWidgetItem()
            table.setItem(row, col, item)

        old_val = item.text()
        item.setText(new_val)
        item.setData(Qt.UserRole + 2, new_val)  # ✅ 设置默认值以供后续校验使用

        if undo_stack and old_val != new_val:
            from modules.condition_input.funcs.undo_command import CellEditCommand
            undo_stack.push(CellEditCommand(table, row, col, old_val, new_val))

    print(f"✅ {side}联动成功: 系数={factor_val} → 等级={grade_val}, 比例={ratio_val}, 合格={qualify_val}")

def autofill_trail_test_grade(trail_table: QTableWidget, row: int, side: str, undo_stack: QUndoStack) -> bool:
    """
    自动推导 技术等级 / 合格级别（无论是否为空，强制写入）：
    - side: "壳程" / "管程"
    - 返回值：是否发生写入
    """
    headers = {trail_table.horizontalHeaderItem(c).text().strip(): c
               for c in range(trail_table.columnCount()) if trail_table.horizontalHeaderItem(c)}

    method_item = trail_table.item(row, headers.get("检测方法"))
    ratio_item = trail_table.item(row, headers.get(f"{side}_检测比例"))
    if not method_item or not ratio_item:
        return False

    method = method_item.text().strip()
    ratio = ratio_item.text().strip()
    if not method or not ratio:
        return False

    if validate_trail_table_cell(f"{side}_检测比例", ratio, None, trail_table) != "ok":
        return False

    import re
    try:
        ratio_num = float(re.sub(r"[^\d.]", "", ratio))
    except ValueError:
        return False

    match_table = {
        "R.T.":  [("100", "AB", "Ⅱ"), ("≥20", "AB", "Ⅲ")],
        "D.R.":  [("100", "AB", "Ⅱ"), ("≥20", "AB", "Ⅲ")],
        "C.R.":  [("100", "AB", "Ⅱ"), ("≥20", "AB", "Ⅲ")],
        "U.T.":  [("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "U.I.T.": [("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "TOFD": [("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "PAUT": [("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "M.T.": [("100", "/",  "Ⅰ")],
        "P.T.": [("100", "/",  "Ⅰ")],
        "M.T.[FB]": [("100", "/", "Ⅰ")]
    }

    candidates = match_table.get(method)
    if not candidates:
        return False

    selected_grade = ""
    selected_qualify = ""
    for limit_str, grade, qualify in candidates:
        if ratio_num >= float(re.sub(r"[^\d.]", "", limit_str)):
            selected_grade = grade
            selected_qualify = qualify
            break

    def force_update_cell(col_name: str, new_val: str) -> bool:
        col = headers.get(col_name)
        if col is None:
            return False
        old_item = trail_table.item(row, col)
        old_val = old_item.text().strip() if old_item else ""
        if not old_item:
            old_item = QTableWidgetItem()
            trail_table.setItem(row, col, old_item)

        old_item.setText(new_val)
        old_item.setData(Qt.UserRole + 2, new_val)
        if undo_stack:
            undo_stack.push(CellEditCommand(trail_table, row, col, old_val, new_val))
        return old_val != new_val

    did_fill1 = force_update_cell(f"{side}_技术等级", selected_grade)
    did_fill2 = force_update_cell(f"{side}_合格级别", selected_qualify)
    return did_fill1 or did_fill2

def compute_trail_default_grade(method: str, ratio_str: str, field_type: str) -> str:
    """
    根据检测方法和检测比例，返回默认 技术等级 或 合格级别。
    - method: 检测方法，如 "R.T."
    - ratio_str: 比例字段，如 "100" 或 "≥20"
    - field_type: "技术等级" 或 "合格级别"
    """
    match_table = {
        "R.T.":  [("100", "AB", "Ⅱ"), ("≥20", "AB", "Ⅲ")],
        "D.R.":  [("100", "AB", "Ⅱ"), ("≥20", "AB", "Ⅲ")],
        "C.R.":  [("100", "AB", "Ⅱ"), ("≥20", "AB", "Ⅲ")],
        "U.T.":  [("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "U.I.T.":[("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "TOFD": [("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "PAUT": [("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "M.T.": [("100", "/",  "Ⅰ")],
        "P.T.": [("100", "/",  "Ⅰ")],
        "M.T.[FB]": [("100", "/", "Ⅰ")]
    }

    import re
    def extract_num(s):
        try:
            return float(re.sub(r"[^\d.]", "", s))
        except:
            return -1

    ratio_num = extract_num(ratio_str)
    candidates = match_table.get(method, [])

    for limit_str, tech, qualify in candidates:
        if ratio_num >= extract_num(limit_str):
            return tech if field_type == "技术等级" else qualify
    return ""

"""技术等级和合格级别不能低于默认值"""
GRADE_ORDER = {"AB": 1, "B": 2, "C": 3}
QUALIFY_ORDER = {"Ⅲ": 1, "Ⅱ": 2, "Ⅰ": 3}

def is_grade_lower(user_val: str, default_val: str) -> bool:
    return GRADE_ORDER.get(user_val, 0) < GRADE_ORDER.get(default_val, 0)

def is_qualify_lower(user_val: str, default_val: str) -> bool:
    return QUALIFY_ORDER.get(user_val, 0) < QUALIFY_ORDER.get(default_val, 0)


"""下拉框定义"""
class MultiParamComboDelegate(QStyledItemDelegate):
    def __init__(self, config: dict, parent=None, viewer=None, undo_stack=None):
        super().__init__(parent)
        self.config = config  # {参数名: {"type": "single"|"multi", "options": [...], "editable": bool}}
        self.viewer = viewer
        self.undo_stack = undo_stack

    def _get_config(self, index):
        row, col = index.row(), index.column()
        param_item = self.parent().item(row, 1)
        if not param_item:
            return None, None
        param_name = param_item.text().strip()
        return self.config.get(param_name), param_name

    def createEditor(self, parent, option, index):
        conf, _ = self._get_config(index)
        if not conf:
            return super().createEditor(parent, option, index)

        if conf["type"] == "multi":
            editor = CheckableComboBox(conf["options"], parent)
            return editor
        else:
            combo = QComboBox(parent)
            combo.addItems(conf["options"])
            combo.setEditable(conf.get("editable", False))
            return combo

    def setEditorData(self, editor, index):
        conf, _ = self._get_config(index)
        if not conf:
            return super().setEditorData(editor, index)
        val = index.data()

        if conf["type"] == "multi":
            values = [v.strip() for v in val.split("；") if v.strip()]
            editor.setCheckedItems(values)
        else:
            i = editor.findText(val)
            editor.setCurrentIndex(i if i >= 0 else 0)

    def setModelData(self, editor, model, index):
        conf, param_name = self._get_config(index)
        if not conf:
            return super().setModelData(editor, model, index)

        old_val = index.data()

        if conf["type"] == "multi":
            new_val = "；".join(editor.checkedItems())
        else:
            new_val = editor.currentText()

        model.setData(index, new_val)

        if old_val != new_val and self.undo_stack:
            cmd = CellEditCommand(self.parent(), index.row(), index.column(), old_val, new_val)
            self.undo_stack.push(cmd)

        # 校验 & 联动
        if self.viewer:
            row, col = index.row(), index.column()
            table = self.parent()
            param_item = table.item(row, 1)
            param_name = param_item.text().strip() if param_item else ""

            if hasattr(table, "logical_headers"):
                column_name = table.logical_headers[col]
            else:
                header_item = table.horizontalHeaderItem(col)
                column_name = header_item.text().strip() if header_item else ""

            # ✅ 调用统一校验分发
            dispatch_cell_validation(self.viewer, table, row, col, param_name, column_name, new_val)

            handle_cross_table_triggers(self.viewer, table, row, col)

#创建自定义 QComboBox 带 checkbox
class CheckableComboBox(QComboBox):
    def __init__(self, options, parent=None):
        super().__init__(parent)
        self.setEditable(True)
        self.setInsertPolicy(QComboBox.NoInsert)
        self.view().setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.setModel(QStandardItemModel(self))
        self._options = options
        self._init_items(options)
        self.lineEdit().setReadOnly(True)
        self.lineEdit().setText("")

    def _init_items(self, options):
        for text in options:
            item = QStandardItem(text)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
            item.setData(Qt.Unchecked, Qt.CheckStateRole)
            self.model().appendRow(item)

    def hidePopup(self):
        selected = []
        for i in range(self.model().rowCount()):
            item = self.model().item(i)
            if item.checkState() == Qt.Checked:
                selected.append(item.text())
        self.lineEdit().setText("；".join(selected))
        super().hidePopup()

    def setCheckedItems(self, values: list):
        # 合并为一个原始字符串，用于乱序/无分隔判断
        raw_text = "".join(values)

        selected = []
        for i in range(self.model().rowCount()):
            item = self.model().item(i)
            option_text = item.text()
            # 若 option_text 在任何原始片段中出现（哪怕没分号），也视为勾选
            if any(option_text in v for v in values) or option_text in raw_text:
                item.setCheckState(Qt.Checked)
                selected.append(option_text)
            else:
                item.setCheckState(Qt.Unchecked)

        self.lineEdit().setText("；".join(selected))

    def checkedItems(self) -> list:
        return [self.model().item(i).text()
                for i in range(self.model().rowCount())
                if self.model().item(i).checkState() == Qt.Checked]

"""添加各表格下拉框"""

#勿删有用！！！
def _get_config(self, index):
    try:
        row, col = index.row(), index.column()
        param_item = self.parent().item(row, 1)
        if not param_item:
            return None, None
        param_name = param_item.text().strip()
        return self.config.get(param_name), param_name
    except Exception as e:
        print(f"[下拉框配置错误] 无法获取参数名: {e}")
        return None, None

#设计数据下拉框
def fetch_design_dropdown_config(product_id):
    """
    从数据库读取所有下拉字段配置，返回 config 字典
    """
    config = {}
    conn = get_connection(**db_config_1)
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 参数名称, type, editable, options
                FROM 设计数据选项模板
            """)
            rows = cursor.fetchall()
            for row in rows:
                param = row["参数名称"]
                typ = row["type"]
                editable = str(row["editable"]).lower() in ("true", "1", "是")
                try:
                    options = ast.literal_eval(row["options"])
                except Exception as e:
                    print(f"⚠️ 参数 {param} 的选项解析失败：{e}")
                    options = []

                config[param] = {
                    "type": typ,
                    "editable": editable,
                    "options": options
                }
    finally:
        conn.close()

    return config
def apply_design_data_dropdowns(table_widget=None, product_id=None, viewer=None, undo_stack=None):
    config = fetch_design_dropdown_config(product_id)

    # ⚠️ 特殊逻辑：耐压试验类型，根据产品类型删减选项
    if product_id:
        prod_type = get_product_type_from_db(product_id)
        if prod_type == "管壳式热交换器":
            if "耐压试验类型*" in config:
                config["耐压试验类型*"]["options"] = ["", "液压试验", "气压试验"]

    return config
def get_product_type_from_db(product_id):
    from modules.condition_input.funcs.db_cnt import get_connection
    conn = get_connection(**db_config_3)
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT 产品类型 FROM 产品需求表 WHERE 产品ID = %s", (product_id,))
            result = cursor.fetchone()
            return result.get("产品类型") if result else ""
    finally:
        conn.close()

#通用数据下拉框
def fetch_general_dropdown_config():
    """
    从数据库读取通用数据表的下拉字段配置
    """
    config = {}
    conn = get_connection(**db_config_1)
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 参数名称, type, editable, options
                FROM 通用数据选项模板
            """)
            rows = cursor.fetchall()
            for row in rows:
                name = row["参数名称"]
                typ = row["type"]
                editable = str(row["editable"]).lower() in ("true", "1", "是")
                try:
                    options = ast.literal_eval(row["options"])
                except Exception as e:
                    print(f"⚠️ 参数 {name} 的 options 无法解析：{e}")
                    options = []

                config[name] = {
                    "type": typ.strip(),
                    "editable": editable,
                    "options": options
                }
    finally:
        conn.close()
    return config
def apply_general_data_dropdowns():
    return fetch_general_dropdown_config()
#勿删
GENERAL_PARAM_CONFIG = fetch_general_dropdown_config()

def fetch_trail_dropdown_config():
    """
    从数据库读取“检测数据”下拉选项配置，返回结构如：
    {
        "R.T.": {
            (2,5): [...],
            (4,7): [...]
        },
        ...
    }
    """
    config = {}
    conn = get_connection(**db_config_1)
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT `接头种类`, `column`, `options` FROM 无损检测数据选项模板")
            for row in cursor.fetchall():
                method = row["接头种类"]
                column_str = row["column"]  # 例如 "2,5"
                try:
                    columns = tuple(int(c.strip()) for c in column_str.split(","))
                    options = ast.literal_eval(row["options"])
                except Exception as e:
                    print(f"❌ 检测数据选项解析失败: {method}-{column_str}: {e}")
                    continue

                if method not in config:
                    config[method] = {}
                config[method][columns] = options
    finally:
        conn.close()
    return config
class TrailTableComboDelegate(QStyledItemDelegate):
    def __init__(self, config=None, parent=None):
        super().__init__(parent)
        self.config = config or {}

    def createEditor(self, parent, option, index):
        method_item = index.sibling(index.row(), 1)
        method_name = method_item.data().strip() if method_item and method_item.data() else ""

        col = index.column()
        options = []
        method_conf = self.config.get(method_name)
        if method_conf:
            for key_cols, vals in method_conf.items():
                if col in key_cols:
                    options = vals
                    break

        if not options:
            return super().createEditor(parent, option, index)

        combo = QComboBox(parent)
        combo.addItems(options)
        combo.setEditable(False)
        # QTimer.singleShot(0, combo.showPopup)  # ✅ 自动弹出

        return combo

    def setModelData(self, editor, model, index):
        method_item = index.sibling(index.row(), 1)
        method_name = method_item.data().strip() if method_item and method_item.data() else ""

        col = index.column()
        new_val = editor.currentText()

        old_val = index.data()
        model.setData(index, new_val)

        # ✅ 撤销记录
        table = self.parent()
        undo_stack = getattr(table, "undo_stack", None)
        if undo_stack and old_val != new_val:
            from modules.condition_input.funcs.undo_command import CellEditCommand
            undo_stack.push(CellEditCommand(table, index.row(), index.column(), old_val, new_val))

        # ✅ 调用校验 & 联动
        viewer = getattr(table, "viewer", None)
        if viewer:
            row = index.row()
            header_item = table.horizontalHeaderItem(col)
            column_name = header_item.text().strip() if header_item else ""
            from modules.condition_input.funcs.funcs_cdt_input import dispatch_cell_validation, handle_cross_table_triggers

            dispatch_cell_validation(viewer, table, row, col, "", column_name, new_val)
            QTimer.singleShot(0, lambda: handle_cross_table_triggers(viewer, table, row, col))

        # 自动提示等级选项改变也能触发联动

    def is_dropdown_cell(self, index):
        col = index.column()
        row = index.row()

        # ✅ 跳过前2行（表头）或越界行
        if row < 2 or row >= self.parent().rowCount():
            return False

        method_item = index.sibling(row, 1)
        if not method_item:
            return False

        method_data = method_item.data()
        if not isinstance(method_data, str):
            return False

        method_name = method_data.strip()
        method_conf = self.config.get(method_name, {})

        for key_cols in method_conf.keys():
            if col in key_cols:
                return True

        return False


def apply_trail_data_dropdowns():
    return fetch_trail_dropdown_config()



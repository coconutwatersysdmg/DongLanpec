import os, re
import shutil
import openpyxl
from PyQt5.QtWidgets import QFileDialog
from collections import defaultdict

from openpyxl.reader.excel import load_workbook

from modules.wenbenshengcheng.db_cnt import get_connection

# ✅ 映射配置
sheet_config = {
    "壳体封头": {
        "sheet_name": "壳体封头",
        "元件名称": "壳体封头",
        "参数映射": {
            "材料类型": "D13",
            "材料牌号": "D14"
        }
    },
    "壳体圆筒": {
        "sheet_name": "壳体圆筒",
        "元件名称": "壳体圆筒",
        "参数映射": {
            "材料类型": "D13",
            "材料牌号": "D14"
        }
    },
    "管箱法兰": {
        "sheet_name": "管箱法兰",
        "参数组": [
            {
                "元件名称": "管箱圆筒",
                "参数映射": {
                    "材料类型": "D9",
                    "材料牌号": "D10"
                }
            },
            {
                "元件名称": "管箱法兰",
                "参数映射": {
                    "法兰类型": "D28",
                    "材料类型": "D29",
                    "材料牌号": "D30"
                }
            },
            {
                "元件名称": "管箱垫片",
                "参数映射": {
                    "垫片系数m": "D19",
                    "垫片比压力y": "D20"
                }
            },
            {
                "元件名称": "螺柱（管箱法兰）",
                "参数映射": {
                    "材料牌号": "D13"
                }
            }
        ]
    },
    "壳体法兰": {
        "sheet_name": "壳体法兰",
        "参数组": [
            {
                "元件名称": "壳体圆筒",
                "参数映射": {
                    "材料类型": "D9",
                    "材料牌号": "D10"
                }
            },
            {
                "元件名称": "壳体法兰",
                "参数映射": {
                    "法兰类型": "D28",
                    "材料类型": "D29",
                    "材料牌号": "D30"
                }
            },
            {
                "元件名称": "管箱侧垫片",
                "参数映射": {
                    "垫片系数m": "D20",
                    "垫片比压力y": "D21"
                }
            },
            {
                "元件名称": "螺柱（管箱法兰）",
                "参数映射": {
                    "材料牌号": "D13"
                }
            }
        ]
    },
    "管箱圆筒": {
        "sheet_name": "管箱圆筒",
        "元件名称": "管箱圆筒",
        "参数映射": {
            "材料类型": "D13",
            "材料牌号": "D14"
        }
    },
    "管箱封头": {
        "sheet_name": "管箱封头",
        "元件名称": "管箱封头",
        "参数映射": {
            "材料类型": "D13",
            "材料牌号": "D14"
        }
    },
    "换热管内压": {
        "sheet_name": "换热管内压",
        "元件名称": "U形换热管",
        "参数映射": {
            "材料类型": "D10",
            "材料牌号": "D9"
        }
    },
    "换热管外压计算报告1": {
        "sheet_name": "换热管外压计算报告1",
        "元件名称": "U形换热管",
        "参数映射": {
            "材料类型": "D9",
            "材料牌号": "D8"
        }
    },
    "换热管外压计算报告2": {
        "sheet_name": "换热管外压计算报告2",
        "元件名称": "U形换热管",
        "参数映射": {
            "材料类型": "D8",
            "材料牌号": "D9"
        }
    },
    "固定管板": {
        "sheet_name": "固定管板",
        "参数组": [
            {
                "元件名称": "固定管板",
                "参数映射": {
                    "材料类型": "D16",
                    "材料牌号": "D15"
                }
            },
            {
                "元件名称": "U形换热管",
                "参数映射": {
                    "材料类型": "D26",
                    "材料牌号": "D25"
                }
            },
            {
                "元件名称": "管箱垫片",
                "参数映射": {
                    "垫片系数m": "D38",
                    "垫片比压力y": "D39"
                }
            }
        ]
    },
    "分程隔板": {
        "sheet_name": "分程隔板",
        "元件名称": "分程隔板",
        "参数映射": {
            "材料类型": "D7",
            "材料牌号": "D8"
        }
    }
}
pipe_param_source = {
    "接管腐蚀裕量": "产品设计活动表_管口零件材料参数表",
    "接管与壳体连接结构型式": "产品设计活动表_管口零件材料参数表",
    "接管材料类型": "产品设计活动表_管口零件材料表",
    "接管材料牌号": "产品设计活动表_管口零件材料表"
}

db_config = {
    'host': 'localhost',
    'port': 3306,
    'user': 'donghua704',
    'password': '123456',
    'database': '产品设计活动库'
}


def generate_calReport(product_id: str):
    template_path = os.path.join(os.path.dirname(__file__), "jisuanbaogao.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError("未找到模板文件: " + template_path)

    temp_path = copy_template_file(template_path, product_id)
    fill_template_with_data(temp_path, product_id)
    save_report_to_user_path(temp_path)


def copy_template_file(template_path: str, product_id: str) -> str:
    """
    复制模板文件，并根据数据库中的管口信息复制相应模板 sheet（壳程/管程）。
    最终删除模板 sheet，仅保留复制内容。
    """
    import openpyxl

    temp_path = os.path.join(os.path.dirname(template_path), "temp_calculate_report.xlsx")
    shutil.copy2(template_path, temp_path)
    wb = openpyxl.load_workbook(temp_path)

    # 检查模板是否存在
    if "壳程模板" not in wb.sheetnames or "管程模板" not in wb.sheetnames:
        print("⚠️ 缺少壳程模板或管程模板 sheet，跳过复制")
        return temp_path

    template_shell = wb["壳程模板"]
    template_tube = wb["管程模板"]

    # 连接数据库获取管口定义
    connection = get_connection(
        db_config['host'],
        db_config['port'],
        db_config['user'],
        db_config['password'],
        db_config['database']
    )

    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 管口代号, 管口用途, 管口所属元件
                FROM 产品设计活动表_管口表
                WHERE 产品ID = %s
            """, (product_id,))
            rows = cursor.fetchall()
            print(f"✅ 获取管口定义共 {len(rows)} 条")

            for i, row in enumerate(rows):
                try:
                    code = row["管口代号"]
                    use = row["管口用途"]
                    owner = row["管口所属元件"]
                    new_title = f"{code}_{use}_{owner}"[:31]

                    # 选择模板
                    if owner in ["壳体封头", "壳体圆筒"]:
                        source_template = template_shell
                        print(f"  🔧 使用壳程模板创建 sheet: {new_title}")
                    elif owner in ["管箱封头", "管箱圆筒"]:
                        source_template = template_tube
                        print(f"  🔧 使用管程模板创建 sheet: {new_title}")
                    else:
                        print(f"  ⚠️ 未识别的管口所属元件: {owner}，跳过")
                        continue

                    copied_sheet = wb.copy_worksheet(source_template)
                    copied_sheet.title = new_title
                except Exception as e:
                    print(f"  ⚠️ 第 {i + 1} 条复制失败: {e}")

        # 删除两个原始模板
        for sheet_name in ["壳程模板", "管程模板"]:
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
                print(f"🧹 已移除模板 sheet: {sheet_name}")

        wb.save(temp_path)
        print("✅ 所有管口 sheet 已复制并保存")

    finally:
        connection.close()

    return temp_path


def get_param_value(connection, product_id: str, component_name: str, param_name: str):
    """
    从 产品设计活动表_元件附加参数表 读取某元件的某参数
    """
    with connection.cursor() as cursor:
        sql = """
            SELECT 参数值, 参数单位
            FROM 产品设计活动表_元件附加参数表
            WHERE 产品ID = %s AND 元件名称 = %s AND 参数名称 = %s
        """
        cursor.execute(sql, (product_id, component_name, param_name))
        return cursor.fetchone()


def get_material_value_by_guankou(product_id: str, param_name: str):
    col_map = {
        "接管材料类型": "材料类型",
        "接管材料牌号": "材料牌号"
    }
    if param_name not in col_map:
        return None

    column = col_map[param_name]
    conn = get_connection(**db_config)
    try:
        with conn.cursor() as cursor:
            cursor.execute(f"""
                SELECT `{column}` FROM 产品设计活动表_管口零件材料表
                WHERE 产品ID = %s AND 零件名称 = '接管'
            """, (product_id,))
            row = cursor.fetchone()
            return row[column] if row else None
    finally:
        conn.close()


def get_guankou_param(product_id: str, param_name: str):
    """
    从 产品设计活动表_管口零件材料参数表 中查询参数（无需零件名）
    """
    conn = get_connection(**db_config)
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 参数值 FROM 产品设计活动表_管口零件材料参数表
                WHERE 产品ID = %s AND 参数名称 = %s
            """, (product_id, param_name))
            row = cursor.fetchone()
            return row["参数值"] if row else None
    finally:
        conn.close()


def fill_template_with_data(file_path: str, product_id: str):
    try:
        get_conditionInput_data(file_path, product_id)
        fill_excel_with_data(file_path, product_id)
        json_path = "jisuan_output_new.json"
        fill_excel_with_dictoutdatas(file_path, json_path)
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise e


def get_conditionInput_data(file_path: str, product_id: str):
    """
    从数据库中获取设计数据和通用数据，并将匹配值写入 Excel。
    - 跳过首个 sheet
    - 匹配 C 列参数名，对应写入 D 列
    """
    connection = get_connection(
        db_config['host'],
        db_config['port'],
        db_config['user'],
        db_config['password'],
        db_config['database']
    )

    try:
        # 获取设计数据
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 参数名称, 壳程数值, 管程数值
                FROM 产品设计活动表_设计数据表
                WHERE 产品ID = %s
            """, (product_id,))
            design_rows = cursor.fetchall()

            design_dict = {}
            for i, row in enumerate(design_rows):
                if not row:
                    continue
                try:
                    pname = row["参数名称"]
                    shell_val = row["壳程数值"]
                    tube_val = row["管程数值"]
                except Exception as e:
                    continue
                design_dict[pname] = {
                    "壳程数值": shell_val,
                    "管程数值": tube_val
                }

        # 获取通用数据
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 参数名称, 数值
                FROM 产品设计活动表_通用数据表
                WHERE 产品ID = %s AND 参数名称 = "是否以外径为基准*"
            """, (product_id,))
            general_rows = cursor.fetchall()

            general_dict = {}
            for i, row in enumerate(general_rows):
                if not row:
                    continue
                try:
                    pname = row["参数名称"]
                    value = row["数值"]
                except Exception as e:
                    continue
                general_dict[pname] = value

        # 字段映射：数据库字段 → 表格中可能的展示字段
        alias_dict = {
            "设计压力*": ["设计压力P", "壳程设计压力P", "管程设计压力P"],
            "设计温度（最高）*": ["设计温度T", "壳程设计温度T", "管程设计温度T", "管箱分程隔板设计温度",
                                "管/壳程设计温度"],
            "腐蚀裕量*": ["腐蚀裕量C", "壳程腐蚀裕量C", "管程腐蚀裕量C"],
            "公称直径*": ["公称直径", "管口公称直径", "公称直径DN"],
            "焊接接头系数*": ["焊接接头系数", "焊接接头系数ф", "纵向焊接接头系数ϕ"],
            "耐压试验类型*": ["压力试验类型", "耐压试验类型"],
            "是否以外径为基准*": ["是否以外径为基准"]
        }

        def normalize(text):
            import re
            return re.sub(r"[\s*（）：:°℃\[\]<>/]|mm|MPa", "", str(text)).lower()

        # 加载 Excel
        wb = openpyxl.load_workbook(file_path)

        for i, sheet in enumerate(wb.worksheets):
            sheet_name = sheet.title
            if i == 0 and "封面" in sheet_name:
                continue

            use_shell = any(x in sheet_name for x in ["壳体", "壳程"])
            use_tube = any(x in sheet_name for x in ["管箱", "管程"])
            source_type = "通用"
            if use_shell:
                source_type = "壳程数值"
            elif use_tube:
                source_type = "管程数值"

            for row in sheet.iter_rows(min_row=5, max_col=4):
                if len(row) < 4 or not row[2].value:
                    continue

                param_name = str(row[2].value).strip()
                d_cell = row[3]
                # ❌ 跳过字段规则：例如 公称直径dB
                if "公称直径dB" in param_name:
                    print(f"⏩ 跳过字段（排除规则命中）：{param_name}")
                    continue
                norm_param = normalize(param_name)
                matched_db_field = None

                for std_field, aliases in alias_dict.items():
                    for alias in aliases:
                        if normalize(alias) in norm_param or norm_param in normalize(alias):
                            matched_db_field = std_field
                            break
                    if matched_db_field:
                        break

                if not matched_db_field:
                    continue

                value = None
                if matched_db_field == "是否以外径为基准*":
                    value = general_dict.get(matched_db_field)
                else:
                    if matched_db_field in design_dict:
                        # 特殊逻辑：较大者的温度取 max(壳,管)
                        if "较大者" in param_name and "温度" in param_name and matched_db_field == "设计温度（最高）*":
                            try:
                                s_val = float(design_dict[matched_db_field].get("壳程数值") or 0)
                                t_val = float(design_dict[matched_db_field].get("管程数值") or 0)
                                value = max(s_val, t_val)
                            except:
                                print("    ⚠️ 较大者温度计算失败")
                        else:
                            if source_type == "通用":
                                if "壳程" or "壳体" in param_name:
                                    value = design_dict[matched_db_field].get("壳程数值")
                                elif "管程" or "管箱" in param_name:
                                    value = design_dict[matched_db_field].get("管程数值")
                                else:
                                    value = design_dict[matched_db_field].get("管程数值")
                            else:
                                value = design_dict[matched_db_field].get(source_type)
                    else:
                        print(f"    ⚠️ 设计字段 {matched_db_field} 不存在于 design_dict")
                if value is not None:
                    d_cell.value = value
                else:
                    print("    ⚠️ 最终未写入任何值")

        wb.save(file_path)
    finally:
        connection.close()


def fill_excel_with_data(file_path: str, product_id: str):
    connection = get_connection(**db_config)
    wb = load_workbook(file_path)

    try:
        # ✅ 先处理固定的 sheet_config（静态映射）
        for config_key, config in sheet_config.items():
            sheet_name = config.get("sheet_name")
            if sheet_name not in wb.sheetnames:
                print(f"⚠️ 未找到 sheet: {sheet_name}，跳过")
                continue

            sheet = wb[sheet_name]
            param_groups = config.get("参数组")

            if param_groups:
                for group in param_groups:
                    component_name = group.get("元件名称")
                    param_map = group.get("参数映射", {})
                    for param_name, cell in param_map.items():
                        result = get_param_value(connection, product_id, component_name, param_name)
                        if result:
                            sheet[cell] = result.get("参数值", "")
                        else:
                            print(f"⚠️ {component_name} 的参数 {param_name} 未找到")
            else:
                component_name = config.get("元件名称")
                param_map = config.get("参数映射", {})
                for param_name, cell in param_map.items():
                    result = get_param_value(connection, product_id, component_name, param_name)
                    if result:
                        sheet[cell] = result.get("参数值", "")
                    else:
                        print(f"⚠️ {component_name} 的参数 {param_name} 未找到")

        # ✅ 接下来处理动态生成的接管 sheet
        unified_guankou_type = get_material_value_by_guankou(product_id, "接管材料类型")
        unified_guankou_grade = get_material_value_by_guankou(product_id, "接管材料牌号")

        for sheet_name in wb.sheetnames:
            if sheet_name.count("_") != 2:
                continue  # 排除非接管子表

            guankou_code, usage, connected_part = sheet_name.split("_")
            sheet = wb[sheet_name]

            part_type = get_param_value(connection, product_id, connected_part, "材料类型")
            part_grade = get_param_value(connection, product_id, connected_part, "材料牌号")

            for row in sheet.iter_rows(min_row=5, max_col=4):
                if not row[2].value:
                    continue
                param_name = str(row[2].value).strip()
                d_cell = row[3]

                if param_name == "接管材料类型":
                    d_cell.value = unified_guankou_type
                elif param_name == "接管材料牌号":
                    d_cell.value = unified_guankou_grade
                elif param_name in ["接管腐蚀裕量", "接管与壳体连接结构型式"]:
                    value = get_guankou_param(product_id, param_name)
                    if value is not None:
                        d_cell.value = value
                elif param_name == f"{connected_part}材料类型":
                    if part_type:
                        d_cell.value = part_type.get("参数值", "")
                elif param_name == f"{connected_part}材料牌号":
                    if part_grade:
                        d_cell.value = part_grade.get("参数值", "")
                elif param_name in pipe_param_source:
                    table = pipe_param_source[param_name]
                    if table == "产品设计活动表_管口零件材料参数表":
                        d_cell.value = get_param_value(connection, product_id, guankou_code, param_name)
            # ✅ 最后补充从 JSON 中读取 DictOutDatas 中的“壳体圆筒”部分

    finally:
        wb.save(file_path)
        connection.close()
        print(f"✅ 数据已成功写入：{file_path}")


import json
from openpyxl import load_workbook



def fill_excel_with_dictoutdatas(file_path: str, json_path: str):
    wb = load_workbook(file_path)

    with open(json_path, 'r', encoding='utf-8') as f:
        json_data = json.load(f)
    print(json_data)
    outdatas = json_data.get("DictOutDatas", {})

    value_map = {}
    for sheet in wb.worksheets:
        # ✅ 壳体圆筒字段映射
        cyl_data = outdatas.get("壳体圆筒", {})
        if cyl_data and cyl_data.get("IsSuccess"):
            for item in cyl_data.get("Datas", []):
                id_ = item.get("Id", "")
                name = item.get("Name", "")
                value = item.get("Value", "")
                if id_ == "工况1：YT9" and name == "计算压力":
                    value_map["计算压力pc"] = value
                elif id_ == "工况1：YT3" and "圆筒内/外径" in name:
                    value_map["圆筒内直径Di"] = value
                    value_map["圆筒外直径Do"] = value
                elif id_ == "工况1：YT23" and name == "圆筒长度":
                    value_map["圆筒长度L"] = value
                elif id_ == "工况1：YT30" and name == "材料试验温度下许用应力":
                    value_map["耐压试验温度下材料许用应力[σ]"] = value
                elif id_ == "工况1：YT31" and name == "材料设计温度下许用应力":
                    value_map["设计温度下材料许用应力[σ]t"] = value
                    value_map["设计温度下计算应力σt"] = value
                elif id_ == "工况1：YT35" and name == "钢材厚度负偏差":
                    value_map["材料厚度负偏差C1"] = value
                elif id_ == "工况1：YT54" and name == "圆筒压力试验压力":
                    value_map["试验压力值PT"] = value
                elif id_ == "工况1：YT50" and name == "圆筒内压强度计算厚度δc":
                    value_map["计算厚度δ"] = value
                elif id_ == "工况1：YT53" and name == "圆筒有效厚度":
                    value_map["有效厚度δe"] = value
                    value_map["最大允许工作压力[Pw]"] = value
                elif id_ == "工况2：YT1" and name == "圆筒名义厚度":
                    value_map["名义厚度δn"] = value
                    value_map["校核条件"] = value

        # ✅ 壳体封头字段映射
        head_data = outdatas.get("壳体封头", {})
        if head_data and head_data.get("IsSuccess"):
            for item in head_data.get("Datas", []):
                id_ = item.get("Id", "")
                name = item.get("Name", "")
                value = item.get("Value", "")
                if id_ == "工况1：EHB11" and name == "计算压力":
                    value_map["计算压力Pc"] = value
                elif id_ == "工况1：EHB12" and name == "椭圆形封头计算内径":
                    value_map["封头内直径Di"] = value
                elif id_ == "工况1：EHB14" and name == "椭圆形封头内曲面深度":
                    value_map["封头内曲面深度hi"] = value
                elif id_ == "工况1：EHB52" and name == "椭圆形封头形状系数K":
                    value_map["椭圆形封头形状系数K"] = value
                elif id_ == "工况2：EHB27" and name == "材料厚度负偏差":
                    value_map["材料厚度负偏差C1"] = value
                elif id_ == "工况1：EHB23" and name == "材料耐压试验温度下许用应力":
                    value_map["室温下材料许用应力[σ]"] = value
                elif id_ == "工况1：EHB22" and name == "材料设计温度下许用应力":
                    value_map["设计温度下材料许用应力[σ]t"] = value
                elif id_ == "工况1：EHB48" and name == "椭圆形封头耐压试验压力":
                    value_map["试验压力值PT"] = value
                elif id_ == "工况1：EHB58" and name == "椭圆形封头最大允许工作压力":
                    value_map["最大允许工作压力[Pw]"] = value
                elif id_ == "工况1：EHB63" and name == "封头直边段(圆筒)最小名义厚度":
                    value_map["最小成形厚度"] = value
                elif id_ == "工况1：EHB57" and name == "椭圆形封头计算厚度":
                    value_map["计算厚度δh"] = value
                elif id_ == "工况1：EHB60" and name == "封头直边段(圆筒)有效厚度":
                    value_map["有效厚度δeh"] = value
                elif id_ == "工况1：EHB46" and name.strip() == "椭圆形封头名义厚度":
                    value_map["名义厚度δnh"] = value


            # ✅ 管箱法兰字段映射
        flange_data = outdatas.get("管箱法兰", {})
        if flange_data and flange_data.get("IsSuccess"):
            for item in flange_data.get("Datas", []):
                id_ = item.get("Id", "")
                name = item.get("Name", "").strip()
                value = item.get("Value", "")
                if not name:
                    continue

                mapping = {
                    "法兰计算压力": "计算压力Pc",
                    "设计温度下法兰材料许用应力": "设计温度下材料许用应力[σ]nt",
                    "法兰名义厚度": "名义厚度δn",
                    "螺栓公称直径": "公称直径dB",
                    "室温下螺栓材料许用应力": "室温下材料许用应力[σ]b",
                    "设计温度下螺栓材料许用应力": "设计温度下材料许用应力[σ]bt",
                    "螺栓根径": "螺栓小径",
                    "螺栓数量": "数量n",
                    "垫片名义内径": "垫片内径D2G",
                    "垫片名义外径": "垫片外径D3G",
                    "垫片有效内径": "垫片接触外径D3G'",
                    "垫片有效外径": "垫片接触外径D3G'",
                    "有效密封宽度b": "垫片有效密封宽度b",
                    "基本密封宽度b0": "垫片基本密封宽度bo",
                    "垫片压紧力作用中心圆直径DG": "垫片压紧力作用中心圆直径DG",
                    "分程隔板处垫片有效密封面积": "分程隔板垫片接触面积",
                    "室温下法兰材料许用应力": "室温度下材料许用应力[σ]f",
                    "法兰当量计算内径": "法兰内径(扣除腐蚀裕量)Di",
                    "法兰名义外径": "法兰外径Do",
                    "螺栓中心圆直径": "螺栓孔中心圆直径Db",
                    "法兰颈部小端有效厚度": "法兰颈部小端有效厚度δn0",
                    "法兰颈部大端有效厚度": "法兰颈部大端有效厚度δn1",
                    "预紧状态下，需要的最小螺栓载荷Wa": "预紧状态下需要的最小螺栓载荷Wa",
                    "操作状态下，需要的最小螺栓载荷Wp": "操作状态下需要的最小螺栓载荷Wp",
                    "预紧状态下，需要的螺栓总截面积Aa": "预紧状态下需要的螺栓总截面积Aa",
                    "实际使用的螺栓总截面积Ab": "实际使用的螺栓总截面积Ab",
                    "螺栓设计载荷W": "预紧状态下螺栓设计载荷W",
                    "螺栓间距": "螺栓间距S",
                    "最大螺栓间距": "螺栓允许最大间距Smax",
                    "最小螺栓间距": "螺栓允许最小间距Smin",
                    "FD": "内压引起的内径截面上的轴向力FD",
                    "预紧状态下法兰垫片压紧力FG": "预紧状态下需要的最小垫片压紧力FG",
                    "FT": "内压引起的总轴向力与内径截面上的轴向力的差值FT=F-FD",
                    "MD": "MD=FD·LD",
                    "MG": "MG=FG·LG",
                    "MT": "MT=FT·LT",
                    "法兰预紧力矩Ma": "预紧状态法兰力矩Ma",
                    "法兰操作力矩Mp": "操作状态法兰力矩Mp",
                    "法兰设计力矩Mo": "法兰设计力矩Mo",
                    "ho": "ho",
                    "K": "K",
                    "T": "T（表7-9）",
                    "Z": "Z（表7-9）",
                    "Y": "Y（表7-9）",
                    "U": "U（表7-9）",
                    "VI": "V1（查图7-4）",
                    "FI": "F1（查图7-3）",
                    "e": "e",
                    "d1": "d1",
                    "f": "f（查图7-7）",
                    "ψ": "Ψ",
                    "γ": "γ",
                    "η": "η",
                    "λ": "λ",
                    "β": "β",
                    "轴向应力": "法兰轴向应力σH",
                    "径向应力": "法兰径向应力σR",
                    "切向应力": "法兰切向应力σT",
                    "综合应力": "综合应力",
                    "预紧状态刚度": "刚度系数J",
                    "壳程设计压力": "试验压力值PT（卧式）",
                    "法兰有效厚度": "法兰有效厚度δf"
                }

                key = mapping.get(name, name)
                value_map[key] = value
        # ✅ 壳体圆筒字段映射
        if sheet.title == "管箱圆筒":
            cyl_data = outdatas.get("管箱圆筒", {})
            if cyl_data and cyl_data.get("IsSuccess"):
                for item in cyl_data.get("Datas", []):
                    id_ = item.get("Id", "")
                    name = item.get("Name", "")
                    value = item.get("Value", "")
                    if id_ == "工况1：YT9" and name == "计算压力":
                        value_map["计算压力pc"] = value
                    elif id_ == "工况1：YT3" and "圆筒内/外径" in name:
                        value_map["圆筒内直径Di"] = value
                        print(value)
                        value_map["圆筒外直径Do"] = value
                    elif id_ == "工况1：YT23" and name == "圆筒长度":
                        value_map["圆筒长度L"] = value
                    elif id_ == "工况1：YT30" and name == "材料试验温度下许用应力":
                        value_map["耐压试验温度下材料许用应力[σ]"] = value
                    elif id_ == "工况1：YT31" and name == "材料设计温度下许用应力":
                        value_map["设计温度下材料许用应力[σ]t"] = value

                        value_map["设计温度下计算应力σt"] = value
                    elif id_ == "工况1：YT35" and name == "钢材厚度负偏差":
                        value_map["材料厚度负偏差C1"] = value
                    elif id_ == "工况1：YT54" and name == "圆筒压力试验压力":
                        value_map["试验压力值PT"] = value
                    elif id_ == "工况1：YT50" and name == "圆筒内压强度计算厚度δc":
                        value_map["计算厚度δ"] = value
                    elif id_ == "工况1：YT53" and name == "圆筒有效厚度":
                        value_map["有效厚度δe"] = value
                        value_map["最大允许工作压力[Pw]"] = value
                    elif id_ == "工况2：YT1" and name == "圆筒名义厚度":
                        value_map["名义厚度δn"] = value
                        value_map["校核条件"] = value
                    key = mapping.get(name, name)
                    value_map[key] = value
             # ✅ 壳体封头字段映射
        if sheet.title == "管箱封头":
            head_data = outdatas.get("管箱封头", {})
            if head_data and head_data.get("IsSuccess"):
                for item in head_data.get("Datas", []):
                    id_ = item.get("Id", "")
                    name = item.get("Name", "")
                    value = item.get("Value", "")
                    if id_ == "工况1：EHB11" and name == "计算压力":
                        value_map["计算压力Pc"] = value
                    elif id_ == "工况1：EHB12" and name == "椭圆形封头计算内径":
                        value_map["封头内直径Di"] = value
                    elif id_ == "工况1：EHB14" and name == "椭圆形封头内曲面深度":
                        value_map["封头内曲面深度hi"] = value
                    elif id_ == "工况1：EHB52" and name == "椭圆形封头形状系数K":
                        value_map["椭圆形封头形状系数K"] = value
                    elif id_ == "工况2：EHB27" and name == "材料厚度负偏差":
                        value_map["材料厚度负偏差C1"] = value
                    elif id_ == "工况1：EHB23" and name == "材料耐压试验温度下许用应力":
                        value_map["室温下材料许用应力[σ]"] = value
                    elif id_ == "工况1：EHB22" and name == "材料设计温度下许用应力":
                        value_map["设计温度下材料许用应力[σ]t"] = value
                    elif id_ == "工况1：EHB48" and name == "椭圆形封头耐压试验压力":
                        value_map["试验压力值PT"] = value
                    elif id_ == "工况1：EHB58" and name == "椭圆形封头最大允许工作压力":
                        value_map["最大允许工作压力[Pw]"] = value
                    elif id_ == "工况1：EHB63" and name == "封头直边段(圆筒)最小名义厚度":
                        value_map["最小成形厚度"] = value
                    elif id_ == "工况1：EHB57" and name == "椭圆形封头计算厚度":
                        value_map["计算厚度δh"] = value
                    elif id_ == "工况1：EHB60" and name == "封头直边段(圆筒)有效厚度":
                        value_map["有效厚度δeh"] = value
                    elif id_ == "工况1：EHB46" and name.strip() == "椭圆形封头名义厚度":
                        value_map["名义厚度δnh"] = value
                    key = mapping.get(name, name)
                    value_map[key] = value

        if sheet.title == "管箱法兰":
            # ✅ 管箱法兰字段映射
            flange_data = outdatas.get("壳体法兰", {})
            if flange_data and flange_data.get("IsSuccess"):
                for item in flange_data.get("Datas", []):
                    id_ = item.get("Id", "")
                    name = item.get("Name", "").strip()
                    value = item.get("Value", "")
                    if not name:
                        continue

                    mapping = {
                        "法兰计算压力": "计算压力Pc",
                        "设计温度下法兰材料许用应力": "设计温度下材料许用应力[σ]nt",
                        "法兰名义厚度": "名义厚度δn",
                        "螺栓公称直径": "公称直径dB",
                        "室温下螺栓材料许用应力": "室温下材料许用应力[σ]b",
                        "设计温度下螺栓材料许用应力": "设计温度下材料许用应力[σ]bt",
                        "螺栓根径": "螺栓小径",
                        "螺栓数量": "数量n",
                        "垫片名义内径": "垫片内径D2G",
                        "垫片名义外径": "垫片外径D3G",
                        "垫片有效内径": "垫片接触外径D3G'",
                        "垫片有效外径": "垫片接触外径D3G'",
                        "有效密封宽度b": "垫片有效密封宽度b",
                        "基本密封宽度b0": "垫片基本密封宽度bo",
                        "垫片压紧力作用中心圆直径DG": "垫片压紧力作用中心圆直径DG",
                        "分程隔板处垫片有效密封面积": "分程隔板垫片接触面积",
                        "室温下法兰材料许用应力": "室温度下材料许用应力[σ]f",
                        "法兰当量计算内径": "法兰内径(扣除腐蚀裕量)Di",
                        "法兰名义外径": "法兰外径Do",
                        "螺栓中心圆直径": "螺栓孔中心圆直径Db",
                        "法兰颈部小端有效厚度": "法兰颈部小端有效厚度δn0",
                        "法兰颈部大端有效厚度": "法兰颈部大端有效厚度δn1",
                        "预紧状态下，需要的最小螺栓载荷Wa": "预紧状态下需要的最小螺栓载荷Wa",
                        "操作状态下，需要的最小螺栓载荷Wp": "操作状态下需要的最小螺栓载荷Wp",
                        "预紧状态下，需要的螺栓总截面积Aa": "预紧状态下需要的螺栓总截面积Aa",
                        "实际使用的螺栓总截面积Ab": "实际使用的螺栓总截面积Ab",
                        "螺栓设计载荷W": "预紧状态下螺栓设计载荷W",
                        "螺栓间距": "螺栓间距S",
                        "最大螺栓间距": "螺栓允许最大间距Smax",
                        "最小螺栓间距": "螺栓允许最小间距Smin",
                        "FD": "内压引起的内径截面上的轴向力FD",
                        "预紧状态下法兰垫片压紧力FG": "预紧状态下需要的最小垫片压紧力FG",
                        "FT": "内压引起的总轴向力与内径截面上的轴向力的差值FT=F-FD",
                        "MD": "MD=FD·LD",
                        "MG": "MG=FG·LG",
                        "MT": "MT=FT·LT",
                        "法兰预紧力矩Ma": "预紧状态法兰力矩Ma",
                        "法兰操作力矩Mp": "操作状态法兰力矩Mp",
                        "法兰设计力矩Mo": "法兰设计力矩Mo",
                        "ho": "ho",
                        "K": "K",
                        "T": "T（表7-9）",
                        "Z": "Z（表7-9）",
                        "Y": "Y（表7-9）",
                        "U": "U（表7-9）",
                        "VI": "V1（查图7-4）",
                        "FI": "F1（查图7-3）",
                        "e": "e",
                        "d1": "d1",
                        "f": "f（查图7-7）",
                        "ψ": "Ψ",
                        "γ": "γ",
                        "η": "η",
                        "λ": "λ",
                        "β": "β",
                        "轴向应力": "法兰轴向应力σH",
                        "径向应力": "法兰径向应力σR",
                        "切向应力": "法兰切向应力σT",
                        "综合应力": "综合应力",
                        "预紧状态刚度": "刚度系数J",
                        "壳程设计压力": "试验压力值PT（卧式）",
                        "法兰有效厚度": "法兰有效厚度δf"
                    }

                    key = mapping.get(name, name)
                    value_map[key] = value
        # ✅ 固定管板字段映射
        if sheet.title == "固定管板":

            tubedata = outdatas.get("固定管板", {})
            if tubedata and tubedata.get("IsSuccess"):
                for item in tubedata.get("Datas", []):
                    id_ = item.get("Id", "")
                    name = item.get("Name", "").strip()
                    value = item.get("Value", "")
                    if not name:
                        continue

                    mapping = {
                        "Pt与Ps是否同时作用": "Pt与Ps是否同时作用",
                        "换热管使用场合": "换热管使用场合",
                        "换热管与管板连接方式": "换热管与管板连接方式 ( 胀接或焊接)",
                        "设计温度下管板材料许用应力": "设计温度下管板材料许用应力[σ]rt",
                        "设计温度下管板材料弹性模量": "设计温度下管板材料弹性模量Ep",
                        "名义厚度 δn'": "管板名义厚度δn",
                        "管程侧隔板槽深": "管程侧隔板槽深 h2",
                        "管板外径": "管板外径Do",
                        "管板强度削弱系数": "管板强度削弱系数 μ",
                        "拉脱力许用值": "许用拉脱力[q]",
                        "设计温度下换热管材料弹性模量": "设计温度下换热管材料弹性模量Et",
                        "设计温度下换热管材料屈服强度": "设计温度下换热管材料屈服点ReLt",
                        "设计温度下换热管材料许用应力": "设计温度下换热管材料许用应力[σ]tt",
                        "换热管外径": "换热管外径d",
                        "换热管壁厚": "换热管壁厚δt",
                        "换热管根数": "换热管根数n",
                        "管程数": "管程数",
                        "换热管排列方式(0:30°;1:60°;2:90°;3:45°)": "换热管排列形式",
                        "换热管中心距": "换热管中心距 S",
                        "外压计算长度": "换热管长度Lt",
                        "沿水平隔板槽一侧的排管根数": "水平隔板一侧排管根数nx",
                        "垫片与密封面接触外径": "垫片外径 Do",
                        "垫片与密封面接触内径": "垫片内径 Di",
                        "垫片压紧力作用中心圆直径": "垫片压紧力作用中心圆直径DG",
                        "管板布管区面积At": "管板布管区面积At",
                        "一根换热管管壁金属横截面积": "一根换热管管壁金属横截面积 a",
                        "管板布管区当量直径Dt": "管板布管区当量直径  Dt",
                        "管程计算用设计压力 Pt'": "系数ρt",
                        "计算厚度δ": "管板最小厚度计算-计算厚度δ",
                        "换热管与管板连接的拉脱力": "拉脱力q"
                    }

                    key = mapping.get(name, name)
                    value_map[key] = value
        # ✅ 固定管板 + 换热管内压字段映射
            # 针对特定 sheet 加载对应模块数据
        if sheet.title == "换热管内压":
            fixed_tube_data = outdatas.get("固定管板", {})
            if fixed_tube_data and fixed_tube_data.get("IsSuccess"):
                for item in fixed_tube_data.get("Datas", []):
                    name = item.get("Name", "").strip()
                    value = item.get("Value", "")
                    if not name:
                        continue
                    mapping = {
                        "管程计算用设计压力 Pt'": "管程设计压力或壳程真空压力(较大者)",
                        "壳程计算用设计压力 Ps'": "管程设计压力或壳程真空压力(较大者)",
                        "管板计算压力Pd": "计算压力",
                        "内孔焊连接的换热管许用轴向应力": "试验温度下换热管材料许用应力[σ]",
                        "设计温度下换热管材料许用应力": "设计温度下换热管材料许用应力 [σ]t",
                        "换热管外径": "换热管外径Do",
                        "换热管最大允许工作压力": "最大允许工作压力[Pw]",
                        "换热管设计温度下的计算应力": "设计温度下计算应力σt",
                        "换热管内压强度计算厚度δc": "换热管计算厚度δc",
                        "换热管有效厚度": "换热管有效厚度δe",
                        "换热管名义厚度": "换热管名义厚度δn"
                    }
                    key = mapping.get(name)
                    if key:
                        value_map[key] = value













        for row in sheet.iter_rows(min_row=5, max_col=4):
            if len(row) < 4 or not row[2].value:
                continue
            param_name = str(row[2].value).strip()
            if param_name in value_map:
                row[3].value = value_map[param_name]
                print(f"✅ 写入 {sheet.title} → {param_name}: {value_map[param_name]}")

    for row in sheet.iter_rows(min_row=5, max_col=4):
        if len(row) < 4 or not row[2].value:
            continue
        param_name = str(row[2].value).strip()
        if param_name in value_map:
            row[3].value = value_map[param_name]
            print(f"✅ 写入 {sheet.title} → {param_name}: {value_map[param_name]}")


    for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=5, max_col=4):
                if not row[2].value:
                    continue
                param_name = str(row[2].value).strip()
                if param_name in value_map:
                    row[3].value = value_map[param_name]
                    # print(f"✅ 写入 {sheet.title} → {param_name}: {value_map[param_name]}")

    wb.save(file_path)
    print("✅ DictOutDatas 数据写入完成")


def save_report_to_user_path(temp_path: str):
    """
    让用户选择一个路径保存最终报告
    """
    save_path, _ = QFileDialog.getSaveFileName(None, "保存计算报告", "", "Excel 文件 (*.xlsx)")
    if save_path:
        shutil.copy2(temp_path, save_path)
        print(f"✅ 文件保存成功: {save_path}")
    else:
        print("⚠️ 用户取消保存")


def normalize_name(name: str) -> str:
    import re
    if not name:
        return ""
    return re.sub(r"[\s\(\)（）：:°℃\[\]<>/]|mm|MPa", "", name).lower()

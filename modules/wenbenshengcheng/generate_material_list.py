import json
import math
import os

import configparser

import chardet
import pymysql
import openpyxl
from openpyxl.reader.excel import load_workbook

from modules.chanpinguanli.chanpinguanli_main import product_manager
from modules.wenbenshengcheng.cunguige import get_ttgd_from_db

product_id = None


def on_product_id_changed(new_id):
    print(f"Received new PRODUCT_ID: {new_id}")
    global product_id
    product_id = new_id


# 测试用产品 ID（真实情况中由外部输入）
product_manager.product_id_changed.connect(on_product_id_changed)

# === 精准映射：元件名称 → List[(section, 字段名, 类型)]




# === 数量 & 单重填写逻辑 ===
import os
import json
import openpyxl
import chardet
import configparser
import pymysql

# === 精准映射：元件名称 → List[(section, 字段名, 类型)]
mapping_dict = {
    "管箱封头": [("管箱封头", "椭圆形封头质量 kg", "质量")],
    "管箱圆筒": [("管箱圆筒", "圆筒重量kg", "质量")],
    "管箱法兰": [("管箱法兰", "法兰毛坯质量", "质量")],
    "固定管板": [("固定管板", "管板重量-毛坯", "质量")],
    "U形换热管": [
        ("固定管板", "换热管根数", "数量")
    ],
    "壳体法兰": [("壳体法兰", "法兰毛坯质量", "质量")],
    "管箱平盖":[("管箱平盖", "法兰毛坯质量", "质量")],
    "头盖法兰":[("头盖法兰", "法兰毛坯质量", "质量")],
    "壳体圆筒": [("壳体圆筒", "圆筒重量kg", "质量")],
    "壳体封头": [("壳体封头", "椭圆形封头质量 kg", "质量")],
    "固定鞍座": [("固定鞍座", "鞍式支座质量", "质量")],
    "螺柱（管箱法兰）": [("管箱法兰", "螺栓数量", "数量")],
    "尾部支撑": [("管束", "尾部支撑数量", "数量")],
    "折流板": [("管束", "折流板数量", "数量")],
    "分程隔板": [("管箱分程隔板", "水平隔板数量", "数量"),
                 ("管箱分程隔板", "管箱分程隔板重量", "质量")],
    "螺柱（壳体法兰）": [("壳体法兰", "螺栓数量", "数量")]
}

def load_json_file(path):
    if not os.path.exists(path):
        print(f"⚠️ JSON 文件不存在: {path}")
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def fill_quantity_weight(json_data, sheet):
    updated = 0
    for row in sheet.iter_rows(min_row=8):
        name_cell = row[3]
        qty_cell = row[6]
        wt_cell = row[7]

        if not name_cell.value:
            continue

        item_name = str(name_cell.value).strip()
        if item_name not in mapping_dict:
            continue

        for section, field_name, data_type in mapping_dict[item_name]:
            datas = json_data.get("DictOutDatas", {}).get(section, {}).get("Datas", [])
            for item in datas:
                if item.get("Name") == field_name:
                    val = item.get("Value", "")
                    try:
                        val = float(val)
                    except:
                        pass

                    if data_type == "数量":
                        qty_cell.value = val
                    elif data_type == "质量":
                        wt_cell.value = val
                    updated += 1
                    break

        h_val = wt_cell.value
        g_val = qty_cell.value
        if (h_val is not None and h_val != "") and (g_val is None or g_val == ""):
            qty_cell.value = 1

    print(f"✅ 已写入数量/单重，共更新 {updated} 项（含自动补1）")
# ✅ 获取材料密度（依赖两个数据库）
def get_material_density(component_name, product_id):
    try:
        conn = pymysql.connect(
            host="localhost", user="root", password="123456",
            database="产品设计活动库", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
        )
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 参数值 FROM 产品设计活动表_元件附加参数表
                WHERE 产品ID = %s AND 元件名称 = %s AND 参数名称 = '材料牌号' LIMIT 1
            """, (product_id, component_name))
            row = cursor.fetchone()
            if row:
                material = row["参数值"]

                conn2 = pymysql.connect(
                    host="localhost", user="root", password="123456",
                    database="材料库", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
                )
                with conn2.cursor() as cursor2:
                    cursor2.execute("""
                        SELECT 材料密度 FROM 材料密度表 WHERE 材料牌号 = %s LIMIT 1
                    """, (material,))
                    row2 = cursor2.fetchone()
                    if row2:
                        return float(row2["材料密度"])
    except Exception as e:
        print(f"❌ 获取材料密度失败: {e}")
    return None
def fill_special_items(sheet, jisuan_data, pipe_data, pipe_input_data, product_id):
    import re


    def get_actual_diameter(dh):
        try:
            conn = pymysql.connect(
                host="localhost", user="root", password="123456",
                database="材料库", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
            )
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT 实际直径 FROM 螺栓直径对应表 WHERE 螺栓公称直径 = %s LIMIT 1
                """, (str(dh),))
                row = cursor.fetchone()
                if row:
                    return float(row["实际直径"])
        except Exception as e:
            print(f"❌ 获取实际直径失败: {e}")
        return None

    def get_luozhu_length(data, product_id):
        dh = get_value(data, "管箱法兰", "螺栓公称直径")
        if dh is None:
            return None
        try:
            dh_val = int(re.search(r'\d+', str(dh)).group())
        except:
            dh_val = 0
        flange_thk_1 = get_value(data, "管箱法兰", "法兰名义厚度") or 0
        gasket_thk_1 = get_value(data, "管箱法兰", "垫片厚度") or 0
        flange_thk_2 = get_value(data, "壳体法兰", "法兰名义厚度") or 0
        gasket_thk_2 = get_value(data, "管箱法兰", "垫片厚度") or 0
        ttgd = get_ttgd_from_db(product_id) or 0
        return 20 + 2 * dh_val + flange_thk_1 + gasket_thk_1 + flange_thk_2 + gasket_thk_2 - 2 * ttgd

    def get_value(data, section, name):
        for section_name, section_data in data.get("DictOutDatas", {}).items():
            if section_name == section:
                for item in section_data.get("Datas", []):
                    if item.get("Name") == name:
                        try:
                            return float(item["Value"])
                        except:
                            return item["Value"]
        return None

    def count_valid_items(data, key):
        return len(data.get(key, [])) if isinstance(data.get(key, []), list) else 0

    # === 滑道质量计算 ===
    def calc_slipway_mass(pipe_input_data, jisuan_output_data, density):
        # === 从布管输入参数.json 中获取高度和厚度 ===
        slipway_height = None
        slipway_thick = None
        for item in pipe_input_data:
            if item.get("paramId") == "LB_SlipWayHeight":
                slipway_height = float(item.get("paramValue", 0)) / 1000
            elif item.get("paramId") == "LB_SlipWayThick":
                slipway_thick = float(item.get("paramValue", 0)) / 1000

        if not slipway_height or not slipway_thick:
            print("❌ 未找到滑道高度或厚度")
            return None

        # === 从 jisuan_output_new.json 中获取滑道长度 ===
        slipway_length = None
        try:
            datas = jisuan_output_data.get("DictOutDatas", {}).get("管束", {}).get("Datas", [])
            for item in datas:
                if item.get("Name") == "滑道长度":
                    slipway_length = float(item.get("Value", 0)) / 1000
                    break
        except Exception as e:
            print(f"❌ 获取滑道长度失败: {e}")
            return None

        if not slipway_length:
            print("❌ 滑道长度无效")
            return None

        # === 计算质量 ===
        try:
            volume = slipway_length * slipway_height * slipway_thick  # m³
            mass = volume * density
            return round(mass, 2)
        except Exception as e:
            print(f"❌ 质量计算失败: {e}")
            return None

    denisty_huadao = get_material_density("滑道", product_id) *1000
    print("denisty_huadao",denisty_huadao)
    slipway_mass = calc_slipway_mass(pipe_input_data, jisuan_data, denisty_huadao)
    print("slipway_mass",slipway_mass)
    def calc_weight(R_mm, thickness_mm, density):
        try:
            R_m = float(R_mm) / 2000  # 直径/2并转米
            t_m = float(thickness_mm) / 1000
            return round(math.pi * R_m ** 2 * t_m * density*1000, 2)
        except Exception as e:
            print(f"❌ 计算质量失败: {e}")
            return None

    def get_param(datas, name):
        for item in datas:
            if item.get("Name") == name:
                return item.get("Value")
        return None

    # === 基础数据获取 ===
    datas = jisuan_data.get("DictOutDatas", {}).get("管箱法兰", {}).get("Datas", [])
    luozhu_qty = next((int(float(item.get("Value", "0"))) for item in datas if item.get("Name") == "螺栓数量"), None)
    datas2 = jisuan_data.get("DictOutDatas", {}).get("管箱平盖", {}).get("Datas", [])
    luozhu_qty2 = next((int(float(item.get("Value", "0"))) for item in datas if item.get("Name") == "螺栓数量"), None)
    guanshu_datas = jisuan_data.get("DictOutDatas", {}).get("管束", {}).get("Datas", [])
    baffle_R = get_param(guanshu_datas, "折流板/支持板外直径")
    baffle_t = get_param(guanshu_datas, "折流板厚度")
    support_R = get_param(guanshu_datas, "折流板/支持板外直径")
    support_t = get_param(guanshu_datas, "支持板厚度")

    saddle_data = jisuan_data.get("DictOutDatas", {}).get("鞍座", {}).get("Datas", [])
    saddle_mass = get_param(saddle_data, "鞍式支座质量")
    saddle_mass = float(saddle_mass) if saddle_mass not in (None, "", "None") else None

    uhx_data = jisuan_data.get("DictOutDatas", {}).get("固定管板", {}).get("Datas", [])
    uhx_mass = get_param(uhx_data, "单根换热管重量kg")
    uhx_mass = float(uhx_mass) if uhx_mass not in (None, "", "None") else None

    tie_rods = pipe_data.get("TieRodsParam", [])

    # === 公称直径 DN ===
    dn_value = None
    try:
        conn1 = pymysql.connect(
            host="localhost", user="root", password="123456",
            database="产品设计活动库", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
        )
        with conn1.cursor() as cursor:
            cursor.execute("""
                SELECT 管程数值 FROM 产品设计活动表_设计数据表
                WHERE 产品ID = %s AND 参数名称 = '公称直径*' LIMIT 1
            """, (product_id,))
            row = cursor.fetchone()
            if row and row.get("管程数值"):
                dn_value = float(row["管程数值"])
            print(dn_value)
        conn1.close()
    except:
        pass

    qty = None
    if dn_value:
        try:
            conn2 = pymysql.connect(
                host="localhost", user="root", password="123456",
                database="配置库", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
            )
            with conn2.cursor() as cursor:
                cursor.execute("SELECT value FROM user_config WHERE id = 2.16")
                row = cursor.fetchone()
                if row:
                    config = eval(row["value"])
                    values = config[1][1:]
                    if dn_value < 800:
                        qty = values[0]
                    elif 800 <= dn_value <= 2000:
                        qty = values[1]
                    else:
                        qty = values[2]
            conn2.close()
        except:
            pass

    # === 默认配置 ===
    fixed_info_map = {
        "铭牌支架": (1, 10),
        "铭牌板": (1, 0.5),
        "铆钉": (8, 0.02)
    }

    quantity_map = {
        # "旁路挡板": count_valid_items(pipe_data, "BPBs"),
        "拉杆": len(tie_rods),
        # "中间挡板": count_valid_items(pipe_data, "VerticalBaffle"),
        "滑道": count_valid_items(pipe_data, "SlipWays"),
        "防冲板": 1 if isinstance(pipe_data.get("ImpingementPlate"), dict) else 0,
        "定距管": len(tie_rods),
        "螺母（拉杆）": len(tie_rods),
        "管箱侧垫片": 1,
        "管箱垫片": 1,
    }

    # === 遍历写入 Excel sheet ===
    for row in sheet.iter_rows(min_row=2):
        name = str(row[3].value).strip()

        if name in quantity_map:
            row[6].value = quantity_map[name]
            if name == "滑道" and slipway_mass:
                row[7].value = slipway_mass
            if name == "拉杆":
                print(123123123123)
                row[6].value = len(tie_rods)

                dh_str = get_value(jisuan_data, "管箱法兰", "螺栓公称直径")
                try:
                    match = re.search(r"M(\d+)", str(dh_str))
                    dh_val = int(match.group(1)) if match else None
                except:
                    dh_val = None
                print(dh_val)
                R = dh_val / 2 if dh_val else None
                H1= get_param(jisuan_data.get("DictOutDatas", {}).get("管束", {}).get("Datas", []), "拉杆长度1")
                H2= get_param(jisuan_data.get("DictOutDatas", {}).get("管束", {}).get("Datas", []), "拉杆长度2")
                H = max(H1,H2)
                print(H)
                density = get_material_density("拉杆", product_id)
                print(density)
                row[6].value = len(tie_rods)

                if R and H and density:
                    try:
                        R_m = R / 1000
                        H_m = float(H) / 1000
                        mass = round((math.pi * R_m ** 2 / 4) * H_m * density * 1000, 2)
                        row[7].value = mass
                    except:
                        pass
            if name == "螺母（拉杆）":
                row[6].value = quantity_map.get("螺母（拉杆）", 0)

                # === 获取公称直径，查找质量 ===
                dia = get_value(jisuan_data, "管箱法兰", "螺栓公称直径")
                if dia:
                    try:
                        conn3 = pymysql.connect(
                            host="localhost", user="root", password="123456",
                            database="材料库", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
                        )
                        with conn3.cursor() as cursor:
                            cursor.execute("""
                                SELECT `管法兰专用螺母` 
                                FROM `螺母近似质量表`
                                WHERE 规格 = %s
                                LIMIT 1
                            """, (str(dia),))
                            row_m = cursor.fetchone()
                            if row_m and row_m.get("管法兰专用螺母"):
                                mass_per_unit = float(row_m["管法兰专用螺母"])
                                row[7].value = mass_per_unit
                        conn3.close()
                    except Exception as e:
                        print(f"❌ 查询螺母质量失败: {e}")
            elif name == "定距管":
                uhx_data = jisuan_data.get("DictOutDatas", {}).get("固定管板", {}).get("Datas", [])
                uhx_mass = get_param(uhx_data, "单根换热管重量kg")
                uhx_mass = float(uhx_mass) if uhx_mass not in (None, "", "None") else None
                row[7].value = uhx_mass
            # if name == "拉杆" and uhx_mass:
            #     row[7].value = round(len(tie_rods) * uhx_mass, 2)
        # 替换原来 mapping_dict 的 "U形换热管" 项：
        elif name == "U形换热管":
            # 获取管程数
            tube_pass = None
            for p in pipe_input_data:
                if p.get("paramId") == "LB_TubePassCount":
                    tube_pass = int(p.get("paramValue", 0))
                    break
            print("tube_pass",tube_pass)
            # 获取换热管长度
            tube_length = None
            for p in pipe_input_data:
                if p.get("paramId") == "LB_TubeLong":
                    tube_length = float(p.get("paramValue", 0))
                    break
            print("tube_length",tube_length)
            # 获取换热管外径、壁厚
            outer_dia = None
            wall_thick = None
            for d in jisuan_data.get("DictOutDatas", {}).get("固定管板", {}).get("Datas", []):
                print()
                if d.get("Name") == "换热管外径":
                    outer_dia = float(d.get("Value", 0))
                elif d.get("Name") == "换热管壁厚":
                    wall_thick = float(d.get("Value", 0))
            print("outer_dia",outer_dia)
            print("wall_thick",wall_thick)
            if tube_pass and tube_length and outer_dia and wall_thick:
                # 计算内径
                inner_dia = outer_dia - wall_thick
                area = math.pi * ((outer_dia / 1000) ** 2 - (inner_dia / 1000) ** 2) / 4  # m²
                print("area",area)
                # 获取TubesParam
                tubes_param = pipe_data.get("TubesParam", [])
                print("tubes_param",tubes_param)
                total_mass = 0.0
                for group in tubes_param:
                    script_items = group.get("ScriptItem", [])
                    print(script_items)
                    for item in script_items:
                        x = item.get("CenterPt", {}).get("X", 0.0)
                        y = item.get("CenterPt", {}).get("Y", 0.0)

                        if tube_pass == 2:
                            u_radius = abs(y) / 1000  # m
                            print(u_radius)
                        elif tube_pass in [4, 6]:
                            u_radius = abs(x) / 1000  # m
                        else:
                            continue

                        expand_len = u_radius * math.pi + (tube_length / 1000) * 2  # m
                        print("expand_len",expand_len)
                        single_mass = area * expand_len * 7850  # kg
                        print("single_mass", single_mass)
                        total_mass += single_mass
                row[7].value = '/'
                row[8].value = round(total_mass, 3)

        elif name == "旁路挡板":
            bpb_list = pipe_data.get("BPBs", [])
            heights = pipe_data.get("BPBHeights", [])
            width_mm = pipe_data.get("BPBThick", 0)

            row[6].value = len(bpb_list)

            try:
                # 获取长度
                H1 = get_param(jisuan_data.get("DictOutDatas", {}).get("管束", {}).get("Datas", []), "拉杆长度1")
                H2 = get_param(jisuan_data.get("DictOutDatas", {}).get("管束", {}).get("Datas", []), "拉杆长度2")
                length_m = max(float(H1 or 0), float(H2 or 0)) / 1000  # 转换成 float 后再除1000

                # 获取密度
                density = get_material_density("旁路挡板", product_id)  # kg/m³

                # 显示第一个挡板质量
                if bpb_list and heights:
                    thickness_mm = float(heights[0])
                    width_mm = float(width_mm)
                    volume = (thickness_mm / 1000) * (width_mm / 1000) * length_m
                    mass = volume * density
                    row[7].value = round(mass, 2)
            except Exception as e:
                print(f"❌ 计算旁路挡板质量失败: {e}")




        elif name == "中间挡板":
            vbaffles = pipe_data.get("VerticalBaffle", [])
            qty = len(vbaffles)
            row[6].value = qty

            try:
                # === 获取厚度和宽度（取第一个挡板）
                if vbaffles:
                    thickness_mm = float(vbaffles[0].get("Width", 0))  # mm
                    width_mm = float(vbaffles[0].get("Height", 0))  # mm
                else:
                    thickness_mm = width_mm = 0

                # === 获取长度（来自 jisuan_data）
                mid_baffle_length = get_param(
                    jisuan_data.get("DictOutDatas", {}).get("管束", {}).get("Datas", []),
                    "中间挡管/挡板长度"
                )
                length_m = float(mid_baffle_length) / 1000 if mid_baffle_length else 0

                # === 获取密度
                density = get_material_density("中间挡板", product_id)  # kg/m³

                # === 计算质量
                volume = (thickness_mm / 1000) * (width_mm / 1000) * length_m  # m³
                total_mass = volume * density * qty *1000
                row[7].value = round(total_mass, 2)
            except Exception as e:
                print(f"❌ 计算中间挡板质量失败: {e}")




        elif name == "螺柱（管箱法兰）" and luozhu_qty:
            row[6].value = luozhu_qty
            dh = get_value(jisuan_data, "管箱法兰", "螺栓公称直径")
            R = get_actual_diameter(dh)
            H = get_luozhu_length(jisuan_data, product_id)
            density = get_material_density("螺柱（管箱法兰）", product_id)
            print("R",R)
            print("H",H)
            print("density",density)
            if R and H and density:
                mass_luozhu = round((math.pi * (R / 1000) ** 2 / 4) * (H / 1000) * density*1000, 2)
                row[7].value = mass_luozhu



        elif name == "螺母（管箱法兰）" and luozhu_qty:
            row[6].value = luozhu_qty * 2
            # === 获取公称直径，查找质量 ===
            dia = get_value(jisuan_data, "管箱法兰", "螺栓公称直径")
            if dia:
                try:
                    conn3 = pymysql.connect(
                        host="localhost", user="root", password="123456",
                        database="材料库", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
                    )
                    with conn3.cursor() as cursor:
                        cursor.execute("""
                            SELECT `管法兰专用螺母` 
                            FROM `螺母近似质量表`
                            WHERE 规格 = %s
                            LIMIT 1
                        """, (str(dia),))
                        row_m = cursor.fetchone()
                        if row_m and row_m.get("管法兰专用螺母"):
                            mass_per_unit = float(row_m["管法兰专用螺母"])
                            row[7].value = mass_per_unit
                    conn3.close()
                except Exception as e:
                    print(f"❌ 查询螺母质量失败: {e}")
        elif name == "螺柱（管箱平盖）" and luozhu_qty2:
            row[6].value = luozhu_qty2
            dh = get_value(jisuan_data, "管箱法兰", "螺栓公称直径")
            R = get_actual_diameter(dh)
            H = get_luozhu_length(jisuan_data, product_id)
            density = get_material_density("螺柱（管箱平盖）", product_id)
            print("R",R)
            print("H",H)
            print("density",density)
            if R and H and density:
                mass_luozhu = round((math.pi * (R / 1000) ** 2 / 4) * (H / 1000) * density*1000, 2)
                row[7].value = mass_luozhu



        elif name == "螺母（管箱平盖）" and luozhu_qty2:
            row[6].value = luozhu_qty2 * 2
            # === 获取公称直径，查找质量 ===
            dia = get_value(jisuan_data, "管箱法兰", "螺栓公称直径")
            if dia:
                try:
                    conn3 = pymysql.connect(
                        host="localhost", user="root", password="123456",
                        database="材料库", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
                    )
                    with conn3.cursor() as cursor:
                        cursor.execute("""
                            SELECT `管法兰专用螺母` 
                            FROM `螺母近似质量表`
                            WHERE 规格 = %s
                            LIMIT 1
                        """, (str(dia),))
                        row_m = cursor.fetchone()
                        if row_m and row_m.get("管法兰专用螺母"):
                            mass_per_unit = float(row_m["管法兰专用螺母"])
                            row[7].value = mass_per_unit
                    conn3.close()
                except Exception as e:
                    print(f"❌ 查询螺母质量失败: {e}")
        elif name == "折流板" and baffle_R and baffle_t:
            density_zheliuban = get_material_density("折流板", product_id)

            row[7].value = calc_weight(baffle_R, baffle_t, density_zheliuban)
        # elif name == "防冲板":

        elif name == "支持板":
            if not row[6].value:
                row[6].value = 1
            if support_R and support_t:
                density_zhichiban = get_material_density("支持板", product_id)

                row[7].value = calc_weight(support_R, support_t, density_zhichiban)
        elif name == "挡管":
            # 获取挡管数量
            dummy_tubes = pipe_data.get("DummyTubesParam", [])
            dummy_count = len(dummy_tubes)
            row[6].value = dummy_count
            # 获取换热管质量
            uhx_data = jisuan_data.get("DictOutDatas", {}).get("固定管板", {}).get("Datas", [])
            uhx_mass = get_param(uhx_data, "单根换热管重量kg")
            uhx_mass = float(uhx_mass) if uhx_mass not in (None, "", "None") else None
            row[7].value = uhx_mass
        elif name in {"固定鞍座", "滑动鞍座"}:
            if not row[6].value:
                row[6].value = 1
            if saddle_mass:
                row[7].value = saddle_mass
        elif name in fixed_info_map:
            row[6].value, row[7].value = fixed_info_map[name]
        elif name == "防松支耳":
            # === 获取防松支耳数量配置 ===
            qty = None
            if dn_value:
                try:
                    conn2 = pymysql.connect(
                        host="localhost", user="root", password="123456",
                        database="配置库", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
                    )
                    with conn2.cursor() as cursor:
                        cursor.execute("SELECT value FROM user_config WHERE id = 2.16")
                        roww = cursor.fetchone()
                        if roww:
                            config = eval(roww["value"])
                            values = config[1][1:]
                            if dn_value < 800:
                                qty = values[0]
                            elif 800 <= dn_value <= 2000:
                                qty = values[1]
                            else:
                                qty = values[2]
                    conn2.close()
                except:
                    pass
            row[6].value = qty
        elif name == "带肩螺柱":
            row[6].value = qty
            row[7].value = mass_luozhu


def generate_material_list(product_id: str, output_path: str):
    template_path = os.path.join(os.getcwd(), "modules/wenbenshengcheng/设备材料清单.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError("未找到模板文件: 设备材料清单.xlsx")

    connection = pymysql.connect(
        host='localhost',
        port=3306,
        user='root',
        password='123456',
        database='产品设计活动库',
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )

    try:
        with connection.cursor() as cursor:
            sql = """
            SELECT 元件名称, 材料类型, 材料牌号, 材料标准, 供货状态
            FROM 产品设计活动表_元件材料表
            WHERE 产品ID = %s
            """
            cursor.execute(sql, (product_id,))
            rows = cursor.fetchall()
    finally:
        connection.close()

    if not rows:
        print(f"⚠️ 未找到产品ID {product_id} 的材料数据")
        return

    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active

    for idx, row in enumerate(rows):
        row_idx = 8 + idx
        sheet[f"A{row_idx}"] = idx + 1
        sheet[f"D{row_idx}"] = row["元件名称"]
        sheet[f"F{row_idx}"] = "/" if row["材料牌号"] == "见参数定义" else row["材料牌号"]
        sheet[f"K{row_idx}"] = "/" if row["材料类型"] == "见参数定义" else row["材料类型"]
        sheet[f"J{row_idx}"] = "/" if row["供货状态"] == "见参数定义" else row["供货状态"]

    config_path = os.path.expandvars(r"%APPDATA%\UDS\蓝滨数字化合作\data\config.ini")
    if not os.path.exists(config_path):
        print(f"❌ 配置文件未找到: {config_path}")
        wb.save(output_path)
        return

    with open(config_path, 'rb') as f:
        raw = f.read()
        encoding = chardet.detect(raw)['encoding'] or 'utf-8'
    config = configparser.ConfigParser()
    config.read_string(raw.decode(encoding))
    product_dir = os.path.normpath(config.get('ProjectInfo', 'product_directory', fallback=''))

    # 加载 JSON
    json_jisuan = load_json_file(os.path.join(os.getcwd(), "jisuan_output_new.json"))
    json_pipe = load_json_file(os.path.join(product_dir, "中间数据", "布管输出参数.json"))
    pipe_input_data = load_json_file(os.path.join(product_dir, "中间数据", "布管输入参数.json"))
    # 填写信息
    fill_quantity_weight(json_jisuan, sheet)
    fill_special_items(sheet, json_jisuan, json_pipe, pipe_input_data,product_id)

    # 保存
    wb.save(output_path)
    print(f"✅ 材料清单已生成：{output_path}")

def fill_quantity_by_relation(sheet):
    """
    根据其他元件的数量或默认规则，补充填写G列数量。
    """
    # 收集所有结构件 → 数量映射（G列）
    name_to_qty = {}
    for row in sheet.iter_rows(min_row=8):
        name_cell = row[3]  # D列
        qty_cell = row[6]  # G列
        if not name_cell.value:
            continue
        item_name = str(name_cell.value).strip()
        name_to_qty[item_name] = qty_cell.value

    # 定义依赖逻辑
    for row in sheet.iter_rows(min_row=8):
        name_cell = row[3]
        qty_cell = row[6]
        if not name_cell.value:
            continue
        item_name = str(name_cell.value).strip()

        # 仅在数量为空时填
        if qty_cell.value not in [None, ""] and qty_cell.value != 0:
            continue

        # 1. 与拉杆数量一致
        if item_name in {"螺母（拉杆）", "定距管"}:
            qty_cell.value = name_to_qty.get("拉杆", "")

        # 2. 螺柱 × 2
        elif item_name == "螺柱（管箱法兰）":
            val = name_to_qty.get("螺柱", "")
            if isinstance(val, (int, float)):
                qty_cell.value = val * 2

        # 3. 防松支耳 → 螺母（管箱法兰）
        elif item_name == "螺母（管箱法兰）":
            qty_cell.value = name_to_qty.get("防松支耳", "")

        # 4. 一些元件固定数量为 1
        elif item_name in {
            "管箱垫片", "支持板", "管箱侧垫片", "固定鞍座", "滑动鞍座","铭牌支架","铭牌板"
        }:
            qty_cell.value = 1
        elif item_name in {
            "铆钉"
        }:
            qty_cell.value = 8

    print("✅ 已填写依赖关系数量（如与拉杆相同、固定为1等）")


def fill_additional_quantities(sheet, path_to_json):
    try:
        with open(path_to_json, "r", encoding="utf-8") as f:
            pipe_data = json.load(f)
    except Exception as e:
        print(f"❌ 无法读取布管输出参数文件: {e}")
        return

    # 计数函数：获取含特征字段的数组元素数量
    def count_valid_items(array_key, required_field):
        items = pipe_data.get(array_key, [])
        if not isinstance(items, list):
            return 0
        return sum(1 for item in items if isinstance(item, dict) and required_field in item)

    quantity_map = {
        "旁路挡板": count_valid_items("BPBs", "BPBHeight"),
        "拉杆": count_valid_items("TieRodsParam", "Postion"),
        "滑道": count_valid_items("SlipWays", "P1"),
        "中间挡板": count_valid_items("DummyTubesParam", "CenterPt"),
        "防冲板": 1 if isinstance(pipe_data.get("ImpingementPlate"), dict) else 0
    }

    for row in sheet.iter_rows(min_row=8):
        name_cell = row[3]  # D列：元件名称
        qty_cell = row[6]   # G列：数量

        if not name_cell.value:
            continue

        item_name = str(name_cell.value).strip()
        if item_name in quantity_map:
            if qty_cell.value in [None, ""]:
                qty_cell.value = quantity_map[item_name]

    print("✅ 已从布管输出参数中填写附加数量（修正字段匹配）")



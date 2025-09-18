import json
import os
import re

import chardet
import configparser
import openpyxl
import pandas as pd
import pymysql

from modules.buguan.buguan_ziyong.My_Piping import create_product_connection
from modules.condition_input.funcs.db_cnt import get_connection
from openpyxl.styles import Alignment, Border, Side, Font
product_id = None


def on_product_id_changed(new_id):
    print(f"Received new PRODUCT_ID: {new_id}")
    global product_id
    product_id = new_id
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center')
font_10 = Font(size=10)

db_config1 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': '产品设计活动库'
}

# === 读取 JSON 数据 ===
def load_json_data(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)


# === 从 JSON 中提取指定 section + 名称 的值 ===
def get_value(data, section, name):
    for section_name, section_data in data.get("DictOutDatas", {}).items():
        if section_name == section:
            for item in section_data.get("Datas", []):
                if item.get("Name") == name:
                    try:
                        return float(item["Value"])
                    except:
                        return item["Value"]
    return None


# === 定义各结构件规格的生成逻辑 ===
def generate_spec(component_name, data, product_id=None):
    """
    根据元件名称返回其规格字符串，如：EHA500X10;h=8
    如果无法生成返回 None
    """
    print('component_name',component_name)
    if component_name == "管箱封头":
        dh = get_value(data, "管箱封头", "封头类型代号")
        d = get_value(data, "管箱封头", "椭圆形封头计算内径")
        t = get_value(data, "管箱封头", "椭圆形封头名义厚度")
        t_plus = get_value(data, "管箱封头", "椭圆形封头最小成型厚度")
        h = get_value(data, "管箱封头", "椭圆形封头直边高度")
        if None not in (dh,d, t, h):
            return f"{dh}{d}×{t}({t_plus});h={h}"
    elif component_name == "外头盖封头":
        dh = get_value(data, "外头盖封头", "封头类型代号")
        d = get_value(data, "外头盖封头", "椭圆形封头计算内径")
        t = get_value(data, "外头盖封头", "椭圆形封头名义厚度")
        t_plus = get_value(data, "外头盖封头", "椭圆形封头最小成型厚度")
        h = get_value(data, "外头盖封头", "椭圆形封头直边高度")
        if None not in (dh,d, t, h):
            return f"{dh}{d}×{t}({t_plus});h={h}"
    elif component_name == "管箱圆筒":
        id_ = get_value(data, "管箱圆筒", "圆筒内径")
        t = get_value(data, "管箱圆筒", "圆筒名义厚度")
        l = get_value(data, "管箱圆筒", "圆筒长度")
        if None not in (id_, t, l):
            return f"ID{id_}×{t};L={l}"
    elif component_name == "外头盖圆筒":
        id_ = get_value(data, "外头盖圆筒", "圆筒内径")
        t = get_value(data, "外头盖圆筒", "圆筒名义厚度")
        l = get_value(data, "外头盖圆筒", "圆筒长度")
        if None not in (id_, t, l):
            return f"ID{id_}×{t};L={l}"
    elif component_name == "管箱法兰":
        w = get_value(data, "管箱法兰", "法兰名义外径")
        n = get_value(data, "管箱法兰", "法兰名义内径")
        h = get_value(data, "管箱法兰", "法兰颈部高度")+get_value(data, "管箱法兰", "法兰名义厚度")
        if None not in (w, n, h):
            return f"Ø{w}/Ø{n}；H={h}"

    elif component_name == "外头盖法兰":
        w = get_value(data, "外头盖法兰", "法兰名义外径")
        n = get_value(data, "外头盖法兰", "法兰名义内径")
        h = get_value(data, "外头盖法兰", "法兰颈部高度")+get_value(data, "外头盖法兰", "法兰名义厚度")
        if None not in (w, n, h):
            return f"Ø{w}/Ø{n}；H={h}"
    elif component_name == "外头盖侧法兰":
        w = get_value(data, "外头盖侧法兰", "法兰名义外径")
        n = get_value(data, "外头盖侧法兰", "法兰名义内径")
        h = get_value(data, "外头盖侧法兰", "法兰颈部高度")+get_value(data, "外头盖侧法兰", "法兰名义厚度")
        if None not in (w, n, h):
            return f"Ø{w}/Ø{n}；H={h}"
    elif component_name == "浮头法兰":
        w = get_value(data, "浮头法兰", "浮头法兰名义外径(含覆层厚度)")
        n = get_value(data, "浮头法兰", "垫片名义内径")
        h1 = get_value(data, "浮头法兰", "法兰颈部高度") or 0
        h2 = get_value(data, "浮头法兰", "浮头法兰名义厚度") or 0
        h = h1 + h2
        if None not in (w, n, h):
            return f"Ø{w}/Ø{n}；H={h}"

    elif component_name == "分程隔板":
        t = get_value(data, "管箱分程隔板", "管箱分程隔板名义厚度")
        if t is not None:
            return f"δ={t}"
    elif component_name == "内导流筒":
        t = get_value(data, "浮头管束", "导流筒厚度")
        if t is not None:
            return f"δ={t}"
    elif component_name == "浮动管板":
        t = get_value(data, "浮头法兰", "浮动管板名义外径")
        if t is not None:
            return f"δ={t}"
    elif component_name == "隔板":
        t = get_value(data, "管箱分程隔板", "管箱分程隔板名义厚度")
        if t is not None:
            return f"δ={t}"
    elif component_name == "管箱垫片":
        w = get_value(data, "管箱法兰", "垫片名义外径")
        n = get_value(data, "管箱法兰", "垫片名义内径")
        if None not in (w, n):
            return f"Ø{w}/Ø{n}"
    elif component_name == "外头盖垫片":
        w = get_value(data, "外头盖法兰", "垫片名义外径")
        n = get_value(data, "外头盖法兰", "垫片名义内径")
        if None not in (w, n):
            return f"Ø{w}/Ø{n}"
    elif component_name == "浮头垫片":
        w = get_value(data, "浮头法兰", "垫片名义外径")
        n = get_value(data, "浮头法兰", "垫片名义内径")
        if None not in (w, n):
            return f"Ø{w}/Ø{n}"

    elif component_name == "U形换热管":
        w = get_value(data, "固定管板", "换热管外径")
        b = get_value(data, "固定管板", "换热管壁厚")
        l = get_pipe_param_value(product_id,"换热管公称长度LN")
        if None not in (w, b, l):
            return f"Ø{w}×Ø{b};L={l}"
    elif component_name == "换热管":
        w = get_value(data, "固定管板", "换热管外径")
        b = get_value(data, "固定管板", "换热管壁厚")
        l = get_pipe_param_value(product_id,"换热管公称长度LN")
        if None not in (w, b, l):
            return f"Ø{w}×Ø{b};L={l}"
    elif component_name == "旁路挡板":
        w = get_pipe_param_value(product_id,"LB_BaffleThick")
        if w not in(None,"Null","null"):
            return f"δ={w}"
    elif component_name == "固定管板":
        w = get_value(data, "固定管板", "管板名义厚度")
        if w is not None:
            return f"δ={w}"
    elif component_name == "定距管":
        # w = get_value(data, "管束", "换热管外径")
        # n = get_value(data, "管束", "换热管壁厚")
        # val1 = get_value(data, "管束", "定距管长度1")
        # val2 = get_value(data, "管束", "定距管长度2")
        # l = max(val1, val2)
        # if None not in (w,n,l):
        #     return f"Ø{w}×{n};L={l}"
        w = get_value(data, "固定管板", "换热管外径")
        b = get_value(data, "固定管板", "换热管壁厚")
        l = get_pipe_param_value(product_id,"换热管公称长度LN")
        if None not in (w, b, l):
            return f"Ø{w}×Ø{b};L={l}"
    elif component_name == "折流板":
        w = get_value(data, "管束", "折流板厚度")
        if w is None:
            w = get_value(data, "浮头管束", "折流板厚度")
        if w is not None:
            return f"δ={w}"
    elif component_name == "内折流板":
        w = get_value(data, "管束", "折流板厚度")
        if w is None:
            w = get_value(data, "浮头管束", "折流板厚度")
        if w is not None:
            return f"δ={w}"
    elif component_name == "异形折流板":
        w = get_value(data, "管束", "折流板厚度")
        if w is None:
            w = get_value(data, "浮头管束", "折流板厚度")
        if w is not None:
            return f"δ={w}"
    elif component_name == "弓形折流板":
        w = get_value(data, "管束", "折流板厚度")
        if w is None:
            w = get_value(data, "浮头管束", "折流板厚度")
        if w is not None:
            return f"δ={w}"
    elif component_name == "防冲板":
        w = get_pipe_param_value(product_id,"LB_BPBThick")
        if w is not None:
            return f"δ={w}"
    elif component_name == "滑道":
        w = get_pipe_param_value(product_id,"LB_SlipWayThick")
        if w is not None:
            return f"δ={w}"
    elif component_name == "支持板":
        w = get_value(data, "管束", "支持板厚度")
        if w is None:
            w = get_value(data, "浮头管束", "支持板厚度")
        if w is not None:
            return f"δ={w}"

    elif component_name == "挡管":
        w = get_value(data, "固定管板", "换热管外径")
        b = get_value(data, "固定管板", "换热管壁厚")
        l = get_value(data, "管束", "中间挡管/挡板长度")
        if l is None:
            l = get_value(data, "浮头管束", "中间挡管/挡板长度")
        if None not in (w, b, l):
            return f"Ø{w}×{b};L={l}"

    elif component_name == "拉杆":
        val1 = get_value(data, "管束", "拉杆长度1")
        if val1 is None:
            val1 = get_value(data, "浮头管束", "拉杆长度1")

        val2 = get_value(data, "管束", "拉杆长度2")
        if val2 is None:
            val2 = get_value(data, "浮头管束", "拉杆长度2")

        w = max(val1, val2) if None not in (val1, val2) else None
        l = get_value(data, "固定管板", "换热管外径")
        if None not in (w, l):
            return f"Ø{w};L={l}"

        if l is not None:
            try:
                l = float(l)
                if 10 <= l <= 14:
                    rod_diameter = 10
                elif 14 < l < 25:
                    rod_diameter = 12
                elif 25 <= l <= 32:
                    rod_diameter = 16
                elif 32 < l <= 57:
                    rod_diameter = 27
                else:
                    rod_diameter = "[超出范围]"
                return f"Ø{rod_diameter},L={w}"
            except:
                return ""

    elif component_name == "螺母（拉杆）":
        w = get_value(data, "固定管板", "换热管外径")
        if w is not None:
            try:
                w = float(w)
                if 10 <= w <= 14:
                    rod_diameter = 10
                elif 14 < w < 25:
                    rod_diameter = 12
                elif 25 <= w <= 32:
                    rod_diameter = 16
                elif 32 < w <= 57:
                    rod_diameter = 27
                else:
                    rod_diameter = "[超出范围]"
                return f"{rod_diameter}"
            except:
                return ""

    elif component_name == "管箱侧垫片":
        w = get_value(data, "管箱法兰", "垫片名义外径")
        n = get_value(data, "管箱法兰", "垫片名义内径")
        if None not in (w, n):
            return f"Ø{w}/Ø{n}"
    elif component_name == "头盖法兰":
        w = get_value(data, "头盖法兰", "法兰名义外径")
        n = get_value(data, "头盖法兰", "法兰名义内径")
        h = get_value(data, "头盖法兰", "法兰颈部高度")+get_value(data, "壳体法兰", "法兰名义厚度")
        if None not in (w, n, h):
            return f"Ø{w}/Ø{n}；H={h}"
    elif component_name == "管箱平盖":
        w = get_value(data, "管箱平盖", "法兰名义外径")
        h = get_value(data, "壳体法兰", "法兰名义厚度")
        if None not in (w, h):
            return f"Ø{w}；H={h}"
    elif component_name == "平盖垫片":
        w = get_value(data, "头盖法兰", "垫片名义外径")
        n = get_value(data, "头盖法兰", "垫片名义内径")
        if None not in (w, n):
            return f"Ø{w}/Ø{n}"
    elif component_name == "壳体法兰":
        w = get_value(data, "壳体法兰", "法兰名义外径")
        n = get_value(data, "壳体法兰", "法兰名义内径")
        h = get_value(data, "壳体法兰", "法兰颈部高度")+get_value(data, "壳体法兰", "法兰名义厚度")
        if None not in (w, n, h):
            return f"Ø{w}/Ø{n}；H={h}"

    elif component_name == "壳体圆筒":
        id_ = get_value(data, "壳体圆筒", "圆筒内径")
        t = get_value(data, "壳体圆筒", "圆筒名义厚度")
        l = get_value(data, "壳体圆筒", "圆筒长度")
        if None not in (id_, t, l):
            return f"ID{id_}×{t};L={l}"
    elif component_name == "壳体封头":
        dh = get_value(data, "壳体封头", "封头类型代号")
        d = get_value(data, "壳体封头", "椭圆形封头计算内径")
        t = get_value(data, "壳体封头", "椭圆形封头名义厚度")
        t_plus = get_value(data, "壳体封头", "椭圆形封头最小成型厚度")
        h = get_value(data, "壳体封头", "椭圆形封头直边高度")
        if None not in (dh,d, t, h):
            return f"{dh}{d}×{t}({t_plus});h={h}"
    elif component_name == "球冠形封头":
        dh = get_value(data, "外头盖封头", "封头类型代号")
        d = get_value(data, "浮头法兰", "球冠形封头内半径")
        t = get_value(data, "浮头法兰", "球冠形封头名义厚度")
        t_plus = get_value(data, "外头盖封头", "椭圆形封头最小成型厚度")
        h = get_value(data, "外头盖封头", "椭圆形封头直边高度")
        if None not in (dh,d, t, h):
            return f"{dh}{d}×{t}({t_plus});h={h}"

    elif component_name == "固定鞍座":
        conn = get_connection(**db_config1)
        cursor = conn.cursor()
        # 获取鞍座型式代号（dh）
        cursor.execute("""
            SELECT 参数值 
            FROM 产品设计活动表_元件附加参数表 
            WHERE 产品ID = %s AND 元件名称 = '固定鞍座' AND 参数名称 = '鞍座型式代号'
            LIMIT 1
        """, (product_id,))
        row_dh = cursor.fetchone()
        dh = row_dh["参数值"] if row_dh and row_dh.get("参数值") not in (None, "", "None") else None
        # 获取鞍座高度h
        cursor.execute("""
            SELECT 参数值 
            FROM 产品设计活动表_元件附加参数表 
            WHERE 产品ID = %s AND 元件名称 = '固定鞍座' AND 参数名称 = '鞍座高度h'
            LIMIT 1
        """, (product_id,))
        row_h = cursor.fetchone()
        h = row_h["参数值"] if row_h and row_h.get("参数值") not in (None, "", "None") else None
        if dh is not None and h is not None:
            return f"{dh},h={h}"
        elif dh is not None:
            return f"{dh}"
        elif h is not None:
            return f"h={h}"
        else:
            return ""

    elif component_name == "滑动鞍座":
        conn = get_connection(**db_config1)
        cursor = conn.cursor()
        # 获取鞍座型式代号（dh）
        cursor.execute("""
            SELECT 参数值 
            FROM 产品设计活动表_元件附加参数表 
            WHERE 产品ID = %s AND 元件名称 = '滑动鞍座' AND 参数名称 = '鞍座型式代号'
            LIMIT 1
        """, (product_id,))
        row_dh = cursor.fetchone()
        dh = row_dh["参数值"] if row_dh and row_dh.get("参数值") not in (None, "", "None") else None
        # 获取鞍座高度h
        cursor.execute("""
            SELECT 参数值 
            FROM 产品设计活动表_元件附加参数表 
            WHERE 产品ID = %s AND 元件名称 = '滑动鞍座' AND 参数名称 = '鞍座高度h'
            LIMIT 1
        """, (product_id,))
        row_h = cursor.fetchone()
        h = row_h["参数值"] if row_h and row_h.get("参数值") not in (None, "", "None") else None
        if dh is not None and h is not None:
            return f"{dh},h={h}"
        elif dh is not None:
            return f"{dh}"
        elif h is not None:
            return f"h={h}"
        else:
            return ""
    elif component_name == "螺柱（管箱法兰）":
        dh = get_value(data, "管箱法兰", "螺栓公称直径")
        if dh is None:
            return None

        try:
            dh_val = int(re.search(r'\d+', str(dh)).group())
        except:
            dh_val = 0

        flange_thk_1 = get_value(data, "管箱法兰", "法兰名义厚度") or 0
        gasket_thk_1 = get_value(data, "管箱法兰", "垫片厚度") or 0
        flange_thk_2 = get_value(data, "壳体法兰", "法兰名义厚度") or 0
        gasket_thk_2 = get_value(data, "壳体法兰", "垫片厚度") or 0
        ttgd = get_ttgd_from_db(product_id) or 0

        l = 20 + 2 * dh_val + flange_thk_1 + gasket_thk_1 + flange_thk_2 + gasket_thk_2 - 2 * ttgd

        return f"{dh}x{l}"
    elif component_name == "螺柱（浮头法兰）":
        dh = get_value(data, "浮头法兰", "螺栓公称直径")
        if dh is None:
            return None
        try:
            dh_val = int(re.search(r'\d+', str(dh)).group())
        except:
            dh_val = 0
        flange_thk_1 = get_value(data, "管箱法兰", "法兰名义厚度") or 0
        gasket_thk_1 = get_value(data, "管箱法兰", "垫片厚度") or 0
        flange_thk_2 = get_value(data, "壳体法兰", "法兰名义厚度") or 0
        gasket_thk_2 = get_value(data, "壳体法兰", "垫片厚度") or 0
        ttgd = get_ttgd_from_db(product_id) or 0

        l = 20 + 2 * dh_val + flange_thk_1 + gasket_thk_1 + flange_thk_2 + gasket_thk_2 - 2 * ttgd

        return f"{dh}x{l}"
    elif component_name == "螺母（管箱法兰）":
        dh = get_value(data, "管箱法兰", "螺栓公称直径")

        if dh is not None:
            return f"{dh}"
    elif component_name == "螺母（浮头法兰）":
        dh = get_value(data, "浮头法兰", "螺栓公称直径")
        if dh is not None:
            return f"{dh}"
    elif component_name == "螺柱（管箱平盖）":
        dh = get_value(data, "管箱平盖", "螺栓公称直径")

        if dh is None:
            return None

        try:
            dh_val = int(re.search(r'\d+', str(dh)).group())
        except:
            dh_val = 0

        flange_thk_1 = get_value(data, "管箱平盖", "法兰名义厚度") or 0
        gasket_thk_1 = get_value(data, "管箱平盖", "垫片厚度") or 0
        flange_thk_2 = get_value(data, "头盖法兰", "法兰名义厚度") or 0
        gasket_thk_2 = get_value(data, "头盖法兰", "垫片厚度") or 0
        ttgd = get_ttgd_from_db(product_id) or 0

        l = 20 + 2 * dh_val + flange_thk_1 + gasket_thk_1 + flange_thk_2 + gasket_thk_2 - 2 * ttgd

        return f"{dh}x{l}"
    elif component_name == "螺母（管箱平盖）":
        dh = get_value(data, "管箱平盖", "螺栓公称直径")
        if dh is not None:
            return f"{dh}"
    # elif component_name == "接管(钢管)":
    #     dh = get_value(data, "管程入口接管", "接管外径")
    #     bh = get_value(data, "管程入口接管", "接管外径")
    #     l = get_value(data, "管程入口接管", "接管实际外伸长度")+get_value(data, "管程入口接管", "接管实际内伸长度")
    #     if None not in (dh, bh):
    #         return f"OD{dh}×{bh};L={l}"
    # elif component_name == "接管(钢管)":
    #     dh = get_value(data, "管程入口接管", "接管外径")
    #     bh = get_value(data, "管程入口接管", "接管名义厚度")
    #     l = get_value(data, "管程入口接管", "接管实际外伸长度")+get_value(data, "管程入口接管", "接管实际内伸长度")
    #     if None not in (dh, bh):
    #         return f"OD{dh}×{bh};L={l}"
    # elif component_name == "接管(钢板)":
    #     dh = get_value(data, "管程入口接管", "接管外径")
    #     bh = get_value(data, "管程入口接管", "接管名义厚度")
    #     l = get_value(data, "管程入口接管", "接管实际外伸长度")+get_value(data, "管程入口接管", "接管实际内伸长度")
    #     if None not in (dh, bh):
    #         return f"OD{dh}×{bh};L={l}"
    # elif component_name == "接管(钢锻件)":
    #     dh = get_value(data, "管程入口接管", "接管外径")
    #     bh = get_value(data, "管程入口接管", "接管内径")
    #     l = get_value(data, "管程入口接管", "接管实际外伸长度")+get_value(data, "管程入口接管", "接管实际内伸长度")
    #     if None not in (dh, bh):
    #         return f"Ø{dh}/Ø{bh}；L={l}"
    # 你可以在此添加更多规则：
    # elif component_name == "其他元件名称":
    #     return "你定义的规格格式"
    elif component_name == "铭牌支架":
        return "δ=5"
    elif component_name == "铭牌板":
        return "δ=2"
    elif component_name == "铆钉":
        return "Ø3×14"
    elif component_name.endswith("接管"):
        print(component_name)
        od = get_value(data, component_name, "接管大端外径")
        thick = get_value(data, component_name, "接管大端壁厚")
        l1 = get_value(data, component_name, "接管实际外伸长度") or 0
        l2 = get_value(data, component_name, "接管实际内伸长度") or 0
        if None not in (od, thick):
            return f"OD{od}×{thick};L={l1 + l2}"

    return None  # 未匹配或数据缺失


# === 写入规格到 Excel ===
def write_spec_to_excel(data, excel_path, sheet_name, product_id):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]

    for row in sheet.iter_rows(min_row=8):  # 从第8行开始
        if len(row) < 5:
            continue
        d_cell = row[3]  # D列
        e_cell = row[4]  # E列

        if d_cell.value:
            name = str(d_cell.value).strip()
            print(name)
            spec = generate_spec(name, data, product_id)
            if spec is not None:
                e_cell.value = spec
            else:
                print(f"⚠️ 无法生成规格：{name}")

    wb.save(excel_path)
    print(f"✅ 已输入规格列至 Excel：{excel_path}")

def get_pipe_param_value(product_id, field_name):
    """
    从产品设计活动表_布管输入表中获取指定字段值
    特殊逻辑：当 field_name = "换热管公称长度" 时，从布管参数表获取 "热交换器公称（换热管）长度 L"
    """
    conn = create_product_connection()
    if conn is None:
        return None

    try:
        cursor = conn.cursor()

        # 特殊逻辑
        if field_name == "换热管公称长度LN":
            sql = """
                SELECT `参数值`
                FROM 产品设计活动表_布管参数表
                WHERE 产品ID = %s AND 参数名 = '热交换器公称（换热管）长度 L'
                LIMIT 1
            """
            cursor.execute(sql, (product_id,))
            row = cursor.fetchone()
            return row["参数值"] if row else None

        # 普通逻辑：从布管输入表读取
        sql = """
            SELECT `key`, `value`
            FROM 产品设计活动表_布管输入表
            WHERE 产品ID = %s
        """
        cursor.execute(sql, (product_id,))
        rows = cursor.fetchall()

        data = {row["key"]: row["value"] for row in rows}
        return data.get(field_name)

    except Exception as e:
        print(f"❌ 获取参数 `{field_name}` 失败: {e}")
        return None
    finally:
        cursor.close()
        conn.close()
def get_ttgd_from_db(product_id):
    try:
        conn = get_connection(**db_config1)
        cursor = conn.cursor()
        sql = """
            SELECT 参数值
            FROM 产品设计活动表_元件附加参数表
            WHERE 产品ID = %s AND 元件名称 = '固定管板' AND 参数名称 = '管板凸台高度'
        """
        cursor.execute(sql, (product_id,))
        row = cursor.fetchone()
        conn.close()
        if row and "参数值" in row:
            return float(row["参数值"])
    except Exception as e:
        print(f"❌ 获取管板凸台高度失败: {e}")
    return 0  # 默认值为0，避免None参与计算出错

def insert_jiaguan_falan_rows(sheet, product_id, json_data):
    """
    在“管口”行后插入接管法兰行。
    - C列：法兰标准
    - D列：管口功能 + 接管法兰
    - E列：规格
    - F列：材料牌号（来自接管材料牌号）
    - H列：质量
    - J列：供货状态
    - K列：材料类型
    """

    nps_to_dn = {
        "1/2": "15", "3/4": "20", "1": "25", "1-1/4": "32", "1-1/2": "40", "2": "50",
        "2-1/2": "65", "3": "80", "4": "100", "5": "125", "6": "150", "8": "200",
        "10": "250", "12": "300", "14": "350", "16": "400", "18": "450",
        "20": "500", "24": "600"
    }

    try:
        conn = get_connection(**db_config1)
        cursor = conn.cursor()

        # 1️⃣ 查询接管法兰主参数
        sql_main = """
            SELECT 法兰标准, 管口功能, 公称尺寸, 压力等级, 法兰型式, 密封面型式
            FROM 产品设计活动表_管口表
            WHERE 产品ID = %s
        """
        cursor.execute(sql_main, (product_id,))
        rows = cursor.fetchall()

        if not rows:
            print("⚠️ 数据库中未找到接管法兰数据")
            conn.close()
            return

        conn.close()

        # 2️⃣ 定位“管口”行
        insert_index = None
        for idx, row in enumerate(sheet.iter_rows(min_row=8), start=8):
            d_val = str(row[3].value).strip()
            if d_val == "管口":
                insert_index = idx + 1
                break

        if insert_index is None:
            print("❌ 未找到“管口”行，无法插入接管法兰")
            return

        # 3️⃣ 倒序插入并输入
        for data in reversed(rows):
            sheet.insert_rows(insert_index)

            standard = str(data.get("法兰标准", "")).strip()
            function = str(data.get("管口功能", "")).strip()
            dn = str(data.get("公称尺寸", "")).strip()
            pn = str(data.get("压力等级", "")).strip()
            flange_type = str(data.get("法兰型式", "")).strip()
            face_type = str(data.get("密封面型式", "")).strip()

            # 🔍 从 JSON 中提取焊端规格
            handuan_type = ""
            jiaguan_key = function + "接管"
            try:
                datas = json_data.get("DictOutDatas", {}).get(jiaguan_key, {}).get("Datas", [])
                for item in datas:
                    if item.get("Name") == "接管与管法兰或外部连接端壁厚（焊端规格）":
                        handuan_type = str(item.get("Value", "")).strip()
                        break
            except Exception as e:
                print(f"⚠️ 获取 {jiaguan_key} 焊端规格失败: {e}")

            # 替换公称尺寸为 DN（若符合）
            dn = nps_to_dn.get(dn, dn)

            # C列
            sheet.cell(row=insert_index, column=3).value = standard
            # D列
            sheet.cell(row=insert_index, column=4).value = f"{function}接管法兰"

            # E列：规格
            if standard in ("HG/T 20615-2009", "HG/T 20592-2009"):
                spec = f"{flange_type} {dn}-{pn} {face_type} s={handuan_type}mm"
            else:
                spec = f"{dn}-{pn} {flange_type} {face_type}"
            sheet.cell(row=insert_index, column=5).value = spec

            # G列：数量
            sheet.cell(row=insert_index, column=7).value = 1

            # 🔎 获取材料牌号/类型（元件计算结果表 + 附加参数表）
            conn = get_connection(**db_config1)
            cursor = conn.cursor()

            # 从 元件计算结果表 里取
            cursor.execute("""
                SELECT Name, Value
                FROM 产品设计活动表_元件计算结果表
                WHERE 产品ID = %s AND 元件名称 = %s
            """, (product_id, jiaguan_key))
            calc_rows = cursor.fetchall()
            cursor.close()

            mat_type = ""
            mat_grade = ""
            for r in calc_rows:
                if r["Name"] == "接管材料类型":
                    mat_type = r["Value"]
                elif r["Name"] == "接管材料牌号":
                    mat_grade = r["Value"]

            # 再查 附加参数表 找供货状态
            supply_status = ""
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 参数名称, 参数值
                FROM 产品设计活动表_管口附加参数表
                WHERE 产品ID = %s
            """, (product_id,))
            param_rows = cursor.fetchall()
            cursor.close()
            conn.close()

            # 参数表转成字典方便查找
            param_map = {r["参数名称"]: r["参数值"] for r in param_rows}

            # 遍历序号 1,2,3，找到和接管对应的材料类型、牌号
            for idx in (1, 2, 3):
                t_key = f"接管材料类型{idx}"
                g_key = f"接管材料牌号{idx}"
                s_key = f"接管供货状态{idx}"

                t_val = param_map.get(t_key, "")
                g_val = param_map.get(g_key, "")
                print("t_val",t_val)
                if t_val == mat_type and g_val == mat_grade:
                    supply_status = param_map.get(s_key, "")
                    break

            # 🔎 同时再去拿接管法兰材料类型/牌号（保持和接管编号一致）
            for idx in (1, 2, 3):
                ft_key = f"接管法兰材料类型{idx}"
                fg_key = f"接管法兰材料牌号{idx}"
                fs_key = f"接管法兰供货状态{idx}"

                # 如果对应的接管编号就是这个 idx，那就取接管法兰的信息
                t_val = param_map.get(f"接管材料类型{idx}", "")
                g_val = param_map.get(f"接管材料牌号{idx}", "")
                if t_val == mat_type and g_val == mat_grade:
                    mat_type = param_map.get(ft_key, mat_type)  # 替换为接管法兰的类型
                    mat_grade = param_map.get(fg_key, mat_grade)  # 替换为接管法兰的牌号
                    supply_status = param_map.get(fs_key, supply_status)  # 替换为接管法兰的供货状态
                    break

            # F列：材料牌号
            sheet.cell(row=insert_index, column=6).value = mat_grade
            # J列：供货状态
            sheet.cell(row=insert_index, column=10).value = supply_status
            # K列：材料类型
            sheet.cell(row=insert_index, column=11).value = mat_type

        print(f"✅ 已插入接管法兰 {len(rows)} 条，含材料信息")

        # === 质量写入逻辑（保持原有 Step1–Step4） ===
        try:
            conn1 = pymysql.connect(
                host="localhost", user="root", password="123456",
                database="产品设计活动库", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
            )
            conn2 = pymysql.connect(
                host="localhost", user="root", password="123456",
                database="材料库", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
            )

            try:
                cursor = conn1.cursor()
                cursor.execute("""
                    SELECT 公称尺寸类型, 公称压力类型 
                    FROM 产品设计活动表_管口类型选择表 
                    WHERE 产品ID = %s LIMIT 1
                """, (product_id,))
                config = cursor.fetchone()
                size_type = config.get("公称尺寸类型", "DN").strip()
                press_type = config.get("公称压力类型", "PN").strip()
                cursor.close()

                cursor = conn1.cursor()
                cursor.execute("""
                    SELECT 管口代号, 管口功能, 公称尺寸, 压力等级, 法兰型式 
                    FROM 产品设计活动表_管口表 
                    WHERE 产品ID = %s
                """, (product_id,))
                kou_rows = cursor.fetchall()
                cursor.close()

                flange_mass_map = {}
                cursor2 = conn2.cursor()
                for row in kou_rows:
                    kou_id = row["管口代号"]
                    size = str(row["公称尺寸"]).strip()
                    pressure = str(row["压力等级"]).strip()
                    flange_type = row["法兰型式"].strip()

                    standard = "20592" if press_type == "PN" else "20615"
                    size_col = "DN" if size_type == "DN" else "NPS"
                    press_col = "PN" if press_type == "PN" else "Class"

                    cursor2.execute(f"""
                        SELECT 质量 FROM 管法兰质量表
                        WHERE 标准 = %s AND 法兰型式代号 = %s AND `{size_col}` = %s AND `{press_col}` = %s
                        LIMIT 1
                    """, (standard, flange_type, size, pressure))
                    res = cursor2.fetchone()
                    flange_mass_map[kou_id] = float(res["质量"]) if res and res.get("质量") else 0.0

                cursor2.close()
            finally:
                conn1.close()
                conn2.close()

            print("✅ flange_mass_map =", flange_mass_map)

            for row in sheet.iter_rows(min_row=2):
                part_name = str(row[3].value).strip()
                for kou in kou_rows:
                    kou_id = kou["管口代号"]
                    kou_func = kou.get("管口功能", "").strip()
                    expected_name = f"{kou_func}接管法兰"
                    if part_name == expected_name:
                        row[7].value = flange_mass_map.get(kou_id, 0)  # H列写质量
                        break

        except Exception as e:
            print(f"❌ 获取接管法兰质量或写入 Excel 失败: {e}")

    except Exception as e:
        print(f"❌ insert_jiaguan_falan_rows 失败: {e}")

import json

def insert_jiaguan_rows(sheet, product_id, data, jisuan_json_path):
    """
    在“管口”行后插入接管行。
    每行包括：
    - D列：管口功能接管
    - E列：规格（依据材料类型判断格式）
    - F列：材料牌号
    - G列：数量（默认为 1）
    - H列：接管重量
    - J列：供货状态
    - K列：材料类型
    """
    import json

    # === 读取计算结果 JSON 文件 ===
    try:
        with open(jisuan_json_path, "r", encoding="utf-8") as f:
            jisuan_data = json.load(f)
            dict_out = jisuan_data.get("DictOutDatas", {})
    except Exception as e:
        print(f"❌ 无法读取计算结果 JSON: {e}")
        dict_out = {}

    conn = get_connection(**db_config1)
    cursor = conn.cursor()

    # === 找到“管口”行 ===
    insert_index = None
    for idx, row in enumerate(sheet.iter_rows(min_row=8), start=8):
        if str(row[3].value).strip() == "管口":
            insert_index = idx + 1
            break
    if insert_index is None:
        print("❌ 未找到“管口”行，无法插入接管")
        return

    # === 获取所有接管名称 ===
    cursor.execute("""
        SELECT DISTINCT 元件名称
        FROM 产品设计活动表_元件计算结果表
        WHERE 产品ID = %s AND 元件名称 LIKE %s
    """, (product_id, '%接管'))
    rows = cursor.fetchall()
    jieguan_names = [row["元件名称"] for row in rows if row["元件名称"]]

    # === 预读取附加参数表 ===
    cursor.execute("""
        SELECT 参数名称, 参数值
        FROM 产品设计活动表_管口附加参数表
        WHERE 产品ID = %s
    """, (product_id,))
    extra_params = {row["参数名称"]: row["参数值"] for row in cursor.fetchall()}

    conn.close()

    # === 倒序插入接管 ===
    for name in reversed(jieguan_names):
        spec = generate_spec(name, data) or ""

        # ⛳ 从 JSON 提取该接管的重量
        mass = ""
        module = dict_out.get(name, {})
        datas = module.get("Datas", [])
        for item in datas:
            if item.get("Name", "").strip() == "接管重量":
                mass = item.get("Value", "")
            if item.get("Name", "").strip() == "接管材料类型":
                mat_type = item.get("Value", "")
            if item.get("Name", "").strip() == "接管材料牌号":
                mat_grade = item.get("Value", "")

        # === 匹配供货状态 ===
        supply_status = ""
        for i in range(1, 4):
            t_key = f"接管材料类型{i}"
            g_key = f"接管材料牌号{i}"
            s_key = f"接管供货状态{i}"
            if extra_params.get(t_key) == mat_type and extra_params.get(g_key) == mat_grade:
                supply_status = extra_params.get(s_key, "")
                break

        # === 插入行 ===
        sheet.insert_rows(insert_index)
        sheet.cell(row=insert_index, column=4).value = name        # D列：接管名称
        sheet.cell(row=insert_index, column=5).value = spec        # E列：规格
        sheet.cell(row=insert_index, column=6).value = mat_grade   # F列：材料牌号
        sheet.cell(row=insert_index, column=7).value = 1           # G列：数量
        sheet.cell(row=insert_index, column=8).value = mass        # H列：重量
        sheet.cell(row=insert_index, column=10).value = supply_status  # J列：供货状态
        sheet.cell(row=insert_index, column=11).value = mat_type       # K列：材料类型




from openpyxl.styles import Alignment, Border, Side, Font

def clean_and_renumber(sheet,product_id):
    """
    删除指定结构件行（黑名单），
    以及灰名单中数据库无对应 name 的行。
    """

    # 黑名单：一定删除
    names_to_remove = {
        "螺母（保温支撑）", "螺柱（保温支撑）",
        "底板（固定鞍座）", "腹板（固定鞍座）", "筋板（固定鞍座）", "垫板（固定鞍座）",
        "底板（滑动鞍座）", "腹板（滑动鞍座）", "筋板（滑动鞍座）", "垫板（滑动鞍座）",
        "支撑板（保温支撑）", "支撑环（保温支撑）", "支撑条（保温支撑）",
        "环首螺钉", "管口",
        "顶丝", "顶板", "堵板", "破涡器",
        "尾部支撑", "防冲板", "纵向隔板"
    }

    # 灰名单：如果数据库里没有匹配则删除
    gray_names = {"外头盖吊耳", "接地端子", "接地板","壳体吊耳"}
    conn = pymysql.connect(
        host="localhost",
        port=3306,
        user="root",
        password="123456",
        database="产品设计活动库",
        charset="utf8mb4"
    )
    cursor = conn.cursor()
    # 数据库查询
    valid_names = set()
    with conn.cursor() as cursor:
        cursor.execute("""
                SELECT name
                FROM 产品设计活动表_元件计算结果表
                WHERE 产品ID = %s
            """, (product_id,))
        rows = cursor.fetchall()
        for row in rows:
            # 假设 cursor 返回 dict 类型（DictCursor），否则要 row[0]
            val = row["name"] if isinstance(row, dict) else row[0]
            if val:
                valid_names.add(str(val).strip())

    # 收集要删除的行
    rows_to_delete = []
    for idx, row in enumerate(sheet.iter_rows(min_row=8), start=8):
        d_val = str(row[3].value).strip() if row[3].value else ""

        # 黑名单：必删
        if d_val in names_to_remove:
            rows_to_delete.append(idx)
            continue

        # 灰名单：数据库无包含关系 → 删
        if any(g in d_val for g in gray_names):
            if not any(valid in d_val for valid in valid_names):
                rows_to_delete.append(idx)

    # 删除行
    for idx in reversed(rows_to_delete):
        sheet.delete_rows(idx)

    # 重新编号和格式化...
    # （下面保持不变）



# === 主函数入口 ===
def main(json_file_path, excel_file_path, sheet_name, product_id):
    import openpyxl
    from openpyxl.cell.cell import MergedCell

    data = load_json_data(json_file_path)
    write_spec_to_excel(data, excel_file_path, sheet_name, product_id)

    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb[sheet_name]
    insert_jiaguan_falan_rows(sheet, product_id,data)
    insert_jiaguan_rows(sheet, product_id, data, "jisuan_output_new.json")
    clean_and_renumber(sheet,product_id)

    # ✅ 填充 I 列：G * H（即第7、8列），仅限 D 列有值的行
    for row in sheet.iter_rows(min_row=8):
        if isinstance(row[8], MergedCell):
            continue  # 跳过合并单元格

        d_val = row[3].value
        g_val = row[6].value
        h_val = row[7].value
        i_cell = row[8]

        if d_val and i_cell.value in (None, "", "None"):  # D 列有值且 I 列没填过
            try:
                g = float(g_val) if g_val not in (None, "", "None") else 0
                h = float(h_val) if h_val not in (None, "", "None") else 0
                i_cell.value = round(g * h, 3)
            except:
                i_cell.value = 0

    # ✅ 删除指定名称的无效零件行，并重新编号 A 列
    remove_names = {"旁路挡板", "中间挡板", "防冲板", "挡管"}
    rows_to_delete = []

    for i, row in enumerate(sheet.iter_rows(min_row=8), start=8):
        d_val = str(row[3].value).strip() if row[3].value else ""
        g_val = row[6].value
        if d_val in remove_names:
            if g_val in (None, "", "None", 0, 0.0, "0"):
                rows_to_delete.append(i)

    # 倒序删除以避免索引错乱
    for i in reversed(rows_to_delete):
        sheet.delete_rows(i)

    # ✅ 重排 A 列序号直到 D 列为空
    current_index = 1
    for row in sheet.iter_rows(min_row=8):
        d_val = row[3].value
        if d_val in (None, "", "None"):
            break
        row[0].value = current_index
        current_index += 1
        # ✅ 写入管箱法兰、固定管板、壳体法兰的质量（L-Q列）
        name_field_map = {
            "管箱法兰": ("管箱法兰", "法兰成型质量"),
            "固定管板": ("固定管板", "管板重量-成品"),
            "壳体法兰": ("壳体法兰", "法兰成型质量"),
            "头盖法兰": ("头盖法兰", "法兰成型质量"),
            "管箱平盖": ("管箱平盖", "法兰成型质量"),
            "外头盖法兰": ("外头盖法兰", "法兰成型质量"),
            "外头盖侧法兰": ("外头盖侧法兰", "法兰成型质量"),

        }

        for row in sheet.iter_rows(min_row=8):
            part_name = str(row[3].value).strip() if row[3].value else ""
            if part_name in name_field_map:
                module, key = name_field_map[part_name]
                try:
                    datas = data.get("DictOutDatas", {}).get(module, {}).get("Datas", [])
                    for item in datas:
                        if item.get("Name") == key:
                            val = item.get("Value", "")
                            # 写入 L 列（即 index 11），合并单元格区域 L-Q 只写 L 即可
                            row[11].value = "成型重量："+val
                            break
                except Exception as e:
                    print(f"⚠️ 处理 {part_name} 时出错：{e}")
    wb.save(excel_file_path)



# === 示例调用 ===
if __name__ == "__main__":
    main("jisuan_output_new.json", "材料清单_已填.xlsx", "Sheet1")
